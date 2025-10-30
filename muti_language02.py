import re
import openpyxl
from openpyxl.styles import PatternFill, Font
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from datetime import datetime
import os



class CompleteDualDisplayTool:
    def __init__(self, root):
        self.root = root
        self.progress = None
        self.progress_label = None
        self.setup_ui()

    def setup_ui(self):
        """设置用户界面"""
        self.root.title("多语言文件对比工具")
        self.root.geometry("750x600")

        # 样式
        pad_y = 10
        btn_width = 35

        # 标题
        tk.Label(self.root, text="多语言文件对比工具", font=('Arial', 16, 'bold')).pack(pady=15)
        tk.Label(self.root, text="结果将按指定语言顺序排列对比", fg="blue").pack()
        tk.Label(self.root,
                 text="差异内容标记红色 | 一致内容标记绿色",
                 fg="red").pack(pady=5)

        # 进度显示区域
        progress_frame = tk.Frame(self.root)
        progress_frame.pack(pady=15, fill=tk.X, padx=30)

        self.progress_label = tk.Label(progress_frame, text="准备就绪", anchor=tk.W, font=('Arial', 10))
        self.progress_label.pack(fill=tk.X)

        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, mode='determinate')
        self.progress.pack(fill=tk.X)

        # 代码转Excel部分
        frame_convert = tk.Frame(self.root)
        frame_convert.pack(pady=10)
        tk.Label(frame_convert, text="第一步: 将代码文件转换为Excel", font=('Arial', 12)).pack()
        tk.Button(
            frame_convert, text="转换代码文件为Excel",
            command=self.start_conversion_thread, width=btn_width, height=2
        ).pack(pady=5)

        # 对比部分
        frame_compare = tk.Frame(self.root)
        frame_compare.pack(pady=10)
        tk.Label(frame_compare, text="第二步: 对比Excel文件", font=('Arial', 12)).pack()

        # 对比顺序说明
        tk.Label(frame_compare,
                 text="对比顺序:\n1. 按键名匹配行\n2. 按指定语言顺序对比\n3. 标记差异(红)和一致(绿)",
                 justify=tk.LEFT).pack()

        tk.Button(
            frame_compare, text="开始对比",
            command=self.start_comparison_thread, width=btn_width, height=2
        ).pack(pady=5)

        # 退出按钮
        tk.Button(self.root, text="退出", command=self.root.quit, width=btn_width - 5).pack(pady=15)

    def get_output_filename(self, suffix):
        """生成带日期时间的输出文件名"""
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{now}_{suffix}.xlsx"

    def update_progress(self, value, message):
        """更新进度条和标签"""
        self.progress['value'] = value
        self.progress_label['text'] = message
        self.root.update_idletasks()

    def start_conversion_thread(self):
        """启动转换线程"""
        thread = threading.Thread(target=self.convert_code_to_excel)
        thread.start()

    def start_comparison_thread(self):
        """启动对比线程"""
        thread = threading.Thread(target=self.compare_excel_files)
        thread.start()

    def convert_code_to_excel(self):
        """将C语言多语言代码转换为Excel"""
        try:
            self.update_progress(0, "正在选择代码文件...")

            code_file = filedialog.askopenfilename(
                title="选择代码文件",
                filetypes=[("C files", "*.c *.h"), ("All files", "*.*")]
            )
            if not code_file:
                self.update_progress(0, "操作已取消")
                return

            output_filename = self.get_output_filename("mutilanguagetest")
            output_excel = filedialog.asksaveasfilename(
                title="保存转换后的Excel文件",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_filename
            )
            if not output_excel:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(20, "正在读取代码文件...")
            with open(code_file, 'r', encoding='utf-8') as f:
                content = f.read()

            self.update_progress(30, "正在解析代码结构...")
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Translations"

            # 添加表头
            headers = ["Key"] + [
                "中文（CN）",
                "英文（EN）English",
                "德语(DE)Deutsch",
                "西语（ES）Español",
                "法语(FR)Français",
                "意大利语(IT)Italiano",
                "巴西葡语(BR)Português",
                "俄语（Pyc）Русский",
                "土耳其语(TR)Turkish",
                "日语(JP)日本語",
                "韩语(KR)한국어",
                "繁体中文",
                "阿拉伯语عربية"
            ]
            ws.append(headers)

            # 匹配多语言数组
            pattern = re.compile(
                r'const char \*(.*?)\[MAX_LANGUAGE\]\s*=\s*\{([^}]*)\};',
                re.DOTALL
            )

            matches = list(pattern.finditer(content))
            total_matches = len(matches)

            self.update_progress(40, f"发现{total_matches}个多语言条目...")

            for i, match in enumerate(matches):
                key = match.group(1).strip()
                values_block = match.group(2)

                # 提取各语言字符串（处理可能的多行情况）
                values = re.findall(r'"(.*?)"', values_block.replace("\n", ""))

                # 写入Excel行
                row = [key] + values[:len(headers) - 1]  # 确保不超过语言列数
                ws.append(row)

                progress = 40 + (i / total_matches) * 50
                self.update_progress(progress, f"正在处理 {i + 1}/{total_matches}...")

            self.update_progress(95, "正在保存Excel文件...")
            wb.save(output_excel)

            self.update_progress(100, "转换完成！")
            messagebox.showinfo("成功", f"转换成功:\n{os.path.basename(output_excel)}")

        except Exception as e:
            self.update_progress(0, f"错误: {str(e)}")
            messagebox.showerror("错误", f"转换失败:\n{str(e)}")
        finally:
            self.root.after(2000, lambda: self.update_progress(0, "准备就绪"))



    def compare_excel_files(self):
        """对比两个Excel文件，按指定语言顺序排列结果"""
        try:
            self.update_progress(0, "正在选择源文件...")
            source_file = filedialog.askopenfilename(
                title="选择源Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not source_file:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(10, "正在选择翻译文件...")
            trans_file = filedialog.askopenfilename(
                title="选择翻译Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not trans_file:
                self.update_progress(0, "操作已取消")
                return

            output_filename = self.get_output_filename("comparison_result")
            output_file = filedialog.asksaveasfilename(
                title="保存对比结果",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_filename
            )
            if not output_file:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(20, "正在加载文件...")

            # 调试信息
            print(f"源文件路径: {source_file}")
            print(f"翻译文件路径: {trans_file}")

            # 加载工作簿
            source_wb = openpyxl.load_workbook(source_file)
            trans_wb = openpyxl.load_workbook(trans_file)
            output_wb = openpyxl.Workbook()

            # 调试信息
            print(f"源文件工作表: {source_wb.sheetnames}")
            print(f"翻译文件工作表: {trans_wb.sheetnames}")

            source_ws = source_wb.active
            trans_ws = trans_wb.active
            output_ws = output_wb.active
            output_ws.title = "对比结果"

            # 调试信息
            print("源文件标题行:", [cell.value for cell in source_ws[1]])
            print("翻译文件标题行:", [cell.value for cell in trans_ws[1]])

            # 定义颜色和样式
            RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            GREEN_FILL = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            HEADER_FONT = Font(bold=True, size=12)
            DIFF_FONT = Font(bold=True, color='FF0000')
            MATCH_FONT = Font(color='006400')
            HEADER_FILL = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')

            # 设置列宽
            output_ws.column_dimensions['A'].width = 25  # 键名
            output_ws.column_dimensions['B'].width = 20  # 语言
            output_ws.column_dimensions['C'].width = 30  # 源文件内容
            output_ws.column_dimensions['D'].width = 30  # 翻译文件内容
            output_ws.column_dimensions['E'].width = 15  # 对比结果

            # 添加表头
            headers = ["键名", "语言", "源文件内容", "翻译文件内容", "对比结果"]
            output_ws.append(headers)

            # 设置表头样式
            for col in range(1, 6):
                cell = output_ws.cell(row=1, column=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL

            # 定义标准语言顺序
            LANGUAGE_ORDER = [
                "中文（CN）",
                "英文（EN）English",
                "德语(DE)Deutsch",
                "西语（ES）Español",
                "法语(FR)Français",
                "意大利语(IT)Italiano",
                "巴西葡语(BR)Português",
                "俄语（Pyc）Русский",
                "土耳其语(TR)Turkish",
                "日语(JP)日本語",
                "韩语(KR)한국어",
                "繁体中文",
                "阿拉伯语عربية"
            ]

            # 1. 找到中文列的索引
            def find_chinese_col(ws, file_type):
                chinese_names = ["中文（CN）", "中文(CN)", "中文", "Chinese", "CN"]
                print(f"\n正在查找{file_type}文件的中文列...")
                for col in range(1, ws.max_column + 1):
                    header = str(ws.cell(row=1, column=col).value).strip()
                    print(f"列{col}: '{header}'")
                    if any(name in header for name in chinese_names):
                        print(f"找到中文列: 第{col}列 '{header}'")
                        return col
                    if re.search(r'中[文国]|汉', header):
                        print(f"通过正则找到中文列: 第{col}列 '{header}'")
                        return col
                print("警告: 未找到明确的中文列，将使用第一列作为默认")
                return 1

            source_cn_col = find_chinese_col(source_ws, "源")
            trans_cn_col = find_chinese_col(trans_ws, "翻译")

            # 2. 构建键映射
            def build_key_map(ws, cn_col, file_type):
                print(f"\n构建{file_type}文件键映射...")
                key_map = {}
                for row in range(2, ws.max_row + 1):
                    key = ws.cell(row=row, column=cn_col).value
                    print(f"行{row} 键值: {key} (类型: {type(key)})")
                    if key is not None:
                        try:
                            key_str = str(key).strip()
                            if key_str:
                                key_map[key_str] = row
                                print(f"有效键: '{key_str}' -> 行{row}")
                            else:
                                print("警告: 空键值，已跳过")
                        except Exception as e:
                            print(f"键值转换错误: {e}，已跳过")
                    else:
                        print("警告: None键值，已跳过")
                print(f"{file_type}文件共找到{len(key_map)}个有效键")
                return key_map

            source_key_map = build_key_map(source_ws, source_cn_col, "源")
            trans_key_map = build_key_map(trans_ws, trans_cn_col, "翻译")

            # 3. 收集所有键
            print("\n收集所有键...")
            all_keys = set()

            print("源文件键:", source_key_map.keys())
            for k in source_key_map:
                try:
                    all_keys.add(str(k).strip())
                    print(f"添加源键: '{k}'")
                except Exception as e:
                    print(f"源键添加错误: {e}")

            print("\n翻译文件键:", trans_key_map.keys())
            for k in trans_key_map:
                try:
                    all_keys.add(str(k).strip())
                    print(f"添加翻译键: '{k}'")
                except Exception as e:
                    print(f"翻译键添加错误: {e}")

            print(f"\n所有键集合: {all_keys}")

            # 4. 安全排序
            print("\n开始排序键...")
            try:
                sample_keys = list(all_keys)[:5]
                print(f"示例键 (前5个): {sample_keys}")
                print(f"键类型: {[type(k) for k in sample_keys]}")

                def safe_key_func(k):
                    try:
                        return str(k).lower()
                    except Exception as e:
                        print(f"键 '{k}' 排序出错: {e}")
                        return ""

                sorted_keys = sorted(all_keys, key=safe_key_func)
                print("排序成功完成")
            except Exception as e:
                print(f"排序出错: {e}")
                print("将使用未排序的键列表")
                sorted_keys = list(all_keys)

            # 5. 定义语言列映射
            def build_lang_col_map(ws):
                lang_map = {}
                for col in range(1, ws.max_column + 1):
                    header = str(ws.cell(row=1, column=col).value).strip()
                    lang_map[header] = col
                return lang_map

            source_lang_map = build_lang_col_map(source_ws)
            trans_lang_map = build_lang_col_map(trans_ws)

            # 6. 执行对比
            differences_found = False
            output_row = 2  # 从第2行开始输出

            for i, cn_key in enumerate(sorted_keys):
                src_row = source_key_map.get(cn_key)
                trans_row = trans_key_map.get(cn_key)

                # 获取两文件中都存在的语言列，并按照标准语言顺序排序
                common_langs = set(source_lang_map.keys()) & set(trans_lang_map.keys())
                # 按照LANGUAGE_ORDER排序，同时保留不在顺序表中的语言
                common_langs_sorted = sorted(common_langs,
                                             key=lambda x: LANGUAGE_ORDER.index(x) if x in LANGUAGE_ORDER else len(
                                                 LANGUAGE_ORDER))

                for lang in common_langs_sorted:
                    if lang == "中文（CN）":
                        continue  # 跳过中文列，因为它是我们的键

                    # 获取值
                    src_val = str(source_ws.cell(
                        row=src_row,
                        column=source_lang_map[lang]
                    ).value) if src_row else ""

                    trans_val = str(trans_ws.cell(
                        row=trans_row,
                        column=trans_lang_map[lang]
                    ).value) if trans_row else ""

                    # 写入键名（只在该语言组的第一行写入）
                    if lang == common_langs_sorted[0]:
                        output_ws.cell(row=output_row, column=1).value = cn_key
                        output_ws.cell(row=output_row, column=1).font = Font(bold=True)

                    # 写入语言标签
                    output_ws.cell(row=output_row, column=2).value = lang

                    # 写入源内容
                    output_ws.cell(row=output_row, column=3).value = src_val

                    # 写入翻译内容
                    output_ws.cell(row=output_row, column=4).value = trans_val

                    # 对比并标记结果
                    if src_val.strip() == trans_val.strip():
                        output_ws.cell(row=output_row, column=5).value = "✓ 一致"
                        output_ws.cell(row=output_row, column=5).font = MATCH_FONT
                        output_ws.cell(row=output_row, column=3).fill = GREEN_FILL
                        output_ws.cell(row=output_row, column=4).fill = GREEN_FILL
                    else:
                        output_ws.cell(row=output_row, column=5).value = "✗ 差异"
                        output_ws.cell(row=output_row, column=5).font = DIFF_FONT
                        output_ws.cell(row=output_row, column=3).fill = RED_FILL
                        output_ws.cell(row=output_row, column=4).fill = RED_FILL
                        differences_found = True

                    output_row += 1

                # 添加空行分隔不同键
                output_row += 1

                # 更新进度
                progress = 40 + (i / len(sorted_keys)) * 50
                self.update_progress(progress, f"正在对比 {i + 1}/{len(sorted_keys)}...")

            # 添加总结信息
            summary_row = output_row + 2
            output_ws.cell(row=summary_row, column=1).value = "对比总结"
            output_ws.cell(row=summary_row, column=1).font = HEADER_FONT

            if differences_found:
                output_ws.cell(row=summary_row + 1, column=1).value = "发现差异内容已用红色标记"
                output_ws.cell(row=summary_row + 1, column=1).font = DIFF_FONT
            else:
                output_ws.cell(row=summary_row + 1, column=1).value = "所有内容完全一致"
                output_ws.cell(row=summary_row + 1, column=1).font = MATCH_FONT

            # 添加颜色说明
            output_ws.cell(row=summary_row + 3, column=1).value = "颜色说明:"
            output_ws.cell(row=summary_row + 3, column=1).font = Font(bold=True)

            # 绿色说明
            green_cell = output_ws.cell(row=summary_row + 4, column=1)
            green_cell.value = "绿色: 内容完全一致"
            green_cell.fill = GREEN_FILL

            # 红色说明
            red_cell = output_ws.cell(row=summary_row + 5, column=1)
            red_cell.value = "红色: 内容存在差异"
            red_cell.fill = RED_FILL

            self.update_progress(95, "正在保存结果...")
            output_wb.save(output_file)

            msg = "对比完成！\n\n"
            msg += "结果文件格式说明:\n"
            msg += "- 第1列: 键名\n"
            msg += "- 第2列: 语言标签\n"
            msg += "- 第3列: 源文件内容\n"
            msg += "- 第4列: 翻译文件内容\n"
            msg += "- 第5列: 对比结果\n\n"
            msg += f"结果已保存到:\n{os.path.basename(output_file)}"

            self.update_progress(100, "对比完成！")
            messagebox.showinfo("完成", msg)

        except Exception as e:
            import traceback
            error_msg = f"对比失败:\n{str(e)}\n\n调试信息:\n"
            error_msg += f"错误类型: {type(e)}\n"
            error_msg += f"调用栈:\n{traceback.format_exc()}"

            try:
                error_msg += f"\n\n源文件行数: {source_ws.max_row}"
                error_msg += f"\n翻译文件行数: {trans_ws.max_row}"
                error_msg += f"\n最后处理的键: {cn_key if 'cn_key' in locals() else '未开始'}"
            except:
                pass

            messagebox.showerror("错误", error_msg)
            self.update_progress(0, f"错误: {str(e)}")
        finally:
            self.root.after(2000, lambda: self.update_progress(0, "准备就绪"))

if __name__ == "__main__":
    root = tk.Tk()
    app = CompleteDualDisplayTool(root)
    root.mainloop()