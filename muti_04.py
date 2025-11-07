import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from datetime import datetime
import os
import logging
from difflib import SequenceMatcher


class CompleteDualDisplayTool:
    def __init__(self, root):
        self.root = root
        self.progress = None
        self.progress_label = None
        self.selected_languages = []
        self.language_checkboxes = []

        # 设置日志
        self.setup_logging()
        self.setup_ui()

    def setup_logging(self):
        """设置日志系统"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('language_tool.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def setup_ui(self):
        """设置用户界面"""
        self.root.title("多语言文件对比工具 v2.0")
        self.root.geometry("900x800")
        self.root.configure(bg='#f0f0f0')

        # 创建主框架
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # 标题
        title_frame = tk.Frame(main_frame, bg='#f0f0f0')
        title_frame.pack(pady=(0, 15))

        tk.Label(title_frame, text="多语言文件对比工具",
                 font=('Arial', 18, 'bold'), bg='#f0f0f0', fg='#2c3e50').pack()

        # 说明标签
        info_frame = tk.Frame(main_frame, bg='#f0f0f0')
        info_frame.pack(pady=(0, 10))

        tk.Label(info_frame, text="支持三种独立的多语言对比功能",
                 font=('Arial', 11), bg='#f0f0f0', fg='#3498db').pack()
        tk.Label(info_frame, text="差异内容标记红色 | 一致内容标记绿色 | 缺失内容标记黄色",
                 font=('Arial', 10), bg='#f0f0f0', fg='#e74c3c').pack()

        # 进度显示区域
        self.setup_progress_section(main_frame)

        # 语言选择部分
        self.setup_language_section(main_frame)

        # 功能按钮区域 - 使用Notebook标签页
        self.setup_function_tabs(main_frame)

    def setup_progress_section(self, parent):
        """设置进度显示区域"""
        progress_frame = tk.LabelFrame(parent, text="进度", font=('Arial', 10, 'bold'),
                                       bg='#f0f0f0', padx=10, pady=10)
        progress_frame.pack(fill=tk.X, pady=10)

        self.progress_label = tk.Label(progress_frame, text="准备就绪",
                                       anchor=tk.W, font=('Arial', 9), bg='#f0f0f0')
        self.progress_label.pack(fill=tk.X)

        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL,
                                        mode='determinate')
        self.progress.pack(fill=tk.X, pady=(5, 0))

    def setup_language_section(self, parent):
        """设置语言选择区域"""
        lang_frame = tk.LabelFrame(parent, text="选择对比语言",
                                   font=('Arial', 10, 'bold'), padx=15, pady=15)
        lang_frame.pack(fill=tk.X, pady=10)

        # 标准语言顺序
        self.LANGUAGE_ORDER = [
            "中文（CN）", "英文（EN）English", "德语(DE)Deutsch",
            "西语（ES）Español", "法语(FR)Français", "意大利语(IT)Italiano",
            "巴西葡语(BR)Português", "俄语（Pyc）Русский", "土耳其语(TR)Turkish",
            "日语(JP)日本語", "韩语(KR)한국어", "繁体中文", "阿拉伯语عربية"
        ]

        # 创建语言选择网格
        self.language_vars = []
        inner_frame = tk.Frame(lang_frame, bg='#f0f0f0')
        inner_frame.pack(fill=tk.X)

        for i, lang in enumerate(self.LANGUAGE_ORDER):
            var = tk.IntVar(value=1)
            self.language_vars.append(var)
            cb = tk.Checkbutton(inner_frame, text=lang, variable=var,
                                font=('Arial', 9), bg='#f0f0f0', anchor='w')
            cb.grid(row=i // 5, column=i % 5, sticky=tk.W + tk.E,
                    padx=5, pady=2)
            inner_frame.columnconfigure(i % 5, weight=1)

        # 全选/取消全选按钮
        btn_frame = tk.Frame(lang_frame, bg='#f0f0f0')
        btn_frame.pack(pady=(10, 0))

        tk.Button(btn_frame, text="全选", command=self.select_all_languages,
                  font=('Arial', 9), width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="取消全选", command=self.deselect_all_languages,
                  font=('Arial', 9), width=10).pack(side=tk.LEFT, padx=5)

    def setup_function_tabs(self, parent):
        """设置功能标签页"""
        # 创建Notebook
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)

        # 功能一：代码转Excel
        tab1 = ttk.Frame(notebook)
        notebook.add(tab1, text="功能一: 代码转Excel")
        self.setup_tab1_content(tab1)

        # 功能二：Excel对比
        tab2 = ttk.Frame(notebook)
        notebook.add(tab2, text="功能二: Excel对比")
        self.setup_tab2_content(tab2)

        # 功能三：错误码校对
        tab3 = ttk.Frame(notebook)
        notebook.add(tab3, text="功能三: 错误码校对")
        self.setup_tab3_content(tab3)

    def setup_tab1_content(self, parent):
        """设置功能一内容"""
        content_frame = tk.Frame(parent, padx=20, pady=20)
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 功能说明
        desc_frame = tk.Frame(content_frame)
        desc_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(desc_frame, text="功能说明:", font=('Arial', 11, 'bold')).pack(anchor=tk.W)
        tk.Label(desc_frame, text="• 将C语言多语言代码文件转换为Excel格式",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 自动解析代码中的多语言数组",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 生成标准的多语言Excel文件",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)

        # 按钮
        btn_frame = tk.Frame(content_frame)
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="开始转换代码文件",
                  command=self.start_conversion_thread,
                  font=('Arial', 12), width=25, height=2,
                  bg='#2ecc71', fg='white').pack()

    def setup_tab2_content(self, parent):
        """设置功能二内容"""
        content_frame = tk.Frame(parent, padx=20, pady=20)
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 功能说明
        desc_frame = tk.Frame(content_frame)
        desc_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(desc_frame, text="功能说明:", font=('Arial', 11, 'bold')).pack(anchor=tk.W)
        tk.Label(desc_frame, text="• 对比两个Excel文件的多语言内容",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 按键名匹配行，按语言顺序对比",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 标记差异(红)、一致(绿)、缺失(黄)",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 以源文件为准，标记对比文件中缺失的内容",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)

        # 按钮
        btn_frame = tk.Frame(content_frame)
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="开始Excel文件对比",
                  command=self.start_comparison_thread,
                  font=('Arial', 12), width=25, height=2,
                  bg='#e67e22', fg='white').pack()

    def setup_tab3_content(self, parent):
        """设置功能三内容"""
        content_frame = tk.Frame(parent, padx=20, pady=20)
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 功能说明
        desc_frame = tk.Frame(content_frame)
        desc_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(desc_frame, text="功能说明:", font=('Arial', 11, 'bold')).pack(anchor=tk.W)
        tk.Label(desc_frame, text="• 通过错误码数字进行模糊匹配",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 源文件格式: 支持AC1234 (2字母+4数字)和纯4位数字",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 对比文件格式: 支持多种格式的模糊匹配",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 以源文件为准，按选择语言对比所有数据",
                 font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=2)
        tk.Label(desc_frame, text="• 使用上方语言选择区域选择要对比的语言",
                 font=('Arial', 10, 'bold'), justify=tk.LEFT, fg='#e74c3c').pack(anchor=tk.W, pady=2)

        # 按钮
        btn_frame = tk.Frame(content_frame)
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="开始错误码校对",
                  command=self.start_error_code_check_thread,
                  font=('Arial', 12), width=25, height=2,
                  bg='#9b59b6', fg='white').pack()

        # 退出按钮
        exit_frame = tk.Frame(content_frame)
        exit_frame.pack(side=tk.BOTTOM, pady=20)

        tk.Button(exit_frame, text="退出程序",
                  command=self.root.quit,
                  font=('Arial', 10), width=15, height=1,
                  bg='#95a5a6', fg='white').pack()

    # ==================== 功能一：代码转Excel ====================
    def convert_code_to_excel(self):
        """将C语言多语言代码转换为Excel"""
        try:
            self.update_progress(0, "正在选择代码文件...")

            code_file = filedialog.askopenfilename(
                title="选择代码文件",
                filetypes=[("C files", "*.c *.h"), ("All files", "*.*")]
            )
            if not code_file:
                self.update_progress(0, "操作已取消")
                return

            output_filename = self.get_output_filename("mutilanguagetest")
            output_excel = filedialog.asksaveasfilename(
                title="保存转换后的Excel文件",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_filename
            )
            if not output_excel:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(20, "正在读取代码文件...")
            with open(code_file, 'r', encoding='utf-8') as f:
                content = f.read()

            self.update_progress(30, "正在解析代码结构...")
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Translations"

            # 添加表头
            headers = ["Key"] + self.LANGUAGE_ORDER
            ws.append(headers)

            # 匹配多语言数组
            pattern = re.compile(
                r'const char \*(.*?)\[MAX_LANGUAGE\]\s*=\s*\{([^}]*)\};',
                re.DOTALL
            )

            matches = list(pattern.finditer(content))
            total_matches = len(matches)

            self.update_progress(40, f"发现{total_matches}个多语言条目...")

            for i, match in enumerate(matches):
                key = match.group(1).strip()
                values_block = match.group(2)

                # 提取各语言字符串（处理可能的多行情况）
                values = re.findall(r'"(.*?)"', values_block.replace("\n", ""))

                # 写入Excel行 - 如果键名包含反斜杠，保持原始格式不分割
                if '\\' in key:
                    # 键名包含反斜杠，保持原始格式
                    row = [key] + values[:len(headers) - 1]  # 确保不超过语言列数
                    ws.append(row)
                    # 调整列宽以适应内容
                    ws.column_dimensions['A'].width = max(25, len(key) + 2)
                else:
                    # 普通键名处理
                    row = [key] + values[:len(headers) - 1]  # 确保不超过语言列数
                    ws.append(row)

                progress = 40 + (i / total_matches) * 50
                self.update_progress(progress, f"正在处理 {i + 1}/{total_matches}...")

            self.update_progress(95, "正在保存Excel文件...")
            wb.save(output_excel)

            self.update_progress(100, "转换完成！")
            messagebox.showinfo("成功", f"转换成功:\n{os.path.basename(output_excel)}")

        except Exception as e:
            self.update_progress(0, f"错误: {str(e)}")
            messagebox.showerror("错误", f"转换失败:\n{str(e)}")
        finally:
            self.root.after(2000, lambda: self.update_progress(0, "准备就绪"))

    # ==================== 功能二：Excel文件对比 ====================
    def compare_excel_files(self):
        """对比两个Excel文件，按指定语言顺序排列结果"""
        try:
            # 获取用户选择的语言
            selected_languages = self.get_selected_languages()
            self.update_progress(0, f"已选择语言: {', '.join(selected_languages)}")

            self.update_progress(5, "正在选择源文件...")
            source_file = filedialog.askopenfilename(
                title="选择源Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not source_file:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(10, "正在选择翻译文件...")
            trans_file = filedialog.askopenfilename(
                title="选择翻译Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not trans_file:
                self.update_progress(0, "操作已取消")
                return

            output_filename = self.get_output_filename("comparison_result")
            output_file = filedialog.asksaveasfilename(
                title="保存对比结果",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_filename
            )
            if not output_file:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(20, "正在加载文件...")

            # 加载工作簿
            source_wb = openpyxl.load_workbook(source_file)
            trans_wb = openpyxl.load_workbook(trans_file)
            output_wb = openpyxl.Workbook()

            source_ws = source_wb.active
            trans_ws = trans_wb.active
            output_ws = output_wb.active
            output_ws.title = "对比结果"

            # 定义颜色和样式
            RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            GREEN_FILL = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            YELLOW_FILL = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            ORANGE_FILL = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')  # 橙色，用于需要人工确认
            HEADER_FONT = Font(bold=True, size=12)
            DIFF_FONT = Font(bold=True, color='FF0000')
            MATCH_FONT = Font(color='006400')
            MISSING_FONT = Font(bold=True, color='FFA500')
            CONFIRM_FONT = Font(bold=True, color='FF8C00')  # 深橙色，用于人工确认
            HEADER_FILL = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')

            # 设置列宽
            output_ws.column_dimensions['A'].width = 25
            output_ws.column_dimensions['B'].width = 25
            output_ws.column_dimensions['C'].width = 20
            output_ws.column_dimensions['D'].width = 30
            output_ws.column_dimensions['E'].width = 30
            output_ws.column_dimensions['F'].width = 15

            # 添加表头
            headers = ["源文件键名", "翻译文件键名", "语言", "源文件内容", "翻译文件内容", "对比结果"]
            output_ws.append(headers)

            # 设置表头样式
            for col in range(1, 7):
                cell = output_ws.cell(row=1, column=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL

            def debug_sheet_info(ws, name):
                """调试函数：打印工作表信息"""
                print(f"\n=== {name} 工作表信息 ===")
                print(f"总行数: {ws.max_row}")
                print(f"总列数: {ws.max_column}")

                # 打印前几行的内容
                for row in range(1, min(4, ws.max_row + 1)):
                    row_data = []
                    for col in range(1, min(ws.max_column + 1, 15)):  # 最多打印15列
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(f"'{cell_value}'" if cell_value else "None")
                    print(f"行 {row}: [{', '.join(row_data)}]")

            # 调试信息
            debug_sheet_info(source_ws, "源文件")
            debug_sheet_info(trans_ws, "翻译文件")

            def build_key_map(ws, file_type):
                """构建键名到行号的映射 - 增强版本"""
                key_map = {}
                print(f"\n=== 开始构建 {file_type} 键映射 ===")

                # 首先确定键名所在的列
                key_column = 1  # 默认第一列

                # 检查第一行，看是否有"Key"或"键名"等标识
                header_row = 1
                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=header_row, column=col).value
                    if header_value and any(keyword in str(header_value).lower()
                                            for keyword in ['key', '键名', '键', '标识', 'id']):
                        key_column = col
                        print(f"发现键名列: 第{col}列 ('{header_value}')")
                        break

                for row in range(2, ws.max_row + 1):
                    key_cell = ws.cell(row=row, column=key_column)
                    key_value = key_cell.value

                    if key_value is not None:
                        try:
                            key_str = str(key_value).strip()
                            if key_str:
                                key_map[key_str] = row
                                if len(key_map) <= 5:  # 只打印前5个键用于调试
                                    print(f"{file_type} 键 {len(key_map)}: '{key_str}' 在行 {row}")
                        except Exception as e:
                            print(f"处理{file_type}文件行{row}时出错: {e}")

                print(f"=== {file_type} 总共找到 {len(key_map)} 个键 ===")
                return key_map, key_column

            # 构建键映射
            source_key_map, source_key_col = build_key_map(source_ws, "源")
            trans_key_map, trans_key_col = build_key_map(trans_ws, "翻译")

            def build_lang_col_map(ws, file_type):
                """构建语言列映射 - 完全匹配实际表头"""
                lang_map = {}
                print(f"\n=== 构建 {file_type} 语言列映射 ===")

                # 收集所有表头
                all_headers = []
                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col).value
                    if header_value is not None:
                        header_str = str(header_value).strip()
                        all_headers.append((col, header_str))
                        print(f"{file_type} 列 {col}: '{header_str}'")

                # 基于实际表头构建映射
                header_mapping = {
                    # 源文件的实际表头映射
                    "中文（CN）": "中文（CN）",
                    "英文（EN）English": "英文（EN）English",

                    # 翻译文件的实际表头映射
                    "中文（CN）": "中文（CN）",
                    "英文（EN）English": "英文（EN）English",
                    "德语(DE)Deutsch": "德语(DE)Deutsch",
                    "西语（ES）Español": "西语（ES）Español",
                    "法语(FR)Français": "法语(FR)Français",
                    "意大利语(IT)Italiano": "意大利语(IT)Italiano",
                    "巴西葡语(BR)Português": "巴西葡语(BR)Português",
                    "俄语（Pyc）Русский": "俄语（Pyc）Русский",
                    "土耳其语(TR)Turkish": "土耳其语(TR)Turkish",
                    "日语(JP)日本語": "日语(JP)日本語",
                    "韩语(KR)한국어": "韩语(KR)한국어",
                    "繁体中文": "繁体中文",
                    "阿拉伯语عربية": "阿拉伯语عربية"
                }

                for col, header in all_headers:
                    # 直接精确匹配
                    if header in header_mapping:
                        lang_name = header_mapping[header]
                        lang_map[lang_name] = col
                        print(f"{file_type} 匹配: '{header}' -> '{lang_name}' 在第 {col} 列")
                    else:
                        # 尝试模糊匹配
                        header_lower = header.lower()
                        for lang_name in header_mapping.values():
                            if any(keyword in header_lower for keyword in [lang_name[:2].lower(),
                                                                           lang_name.split('（')[
                                                                               0].lower() if '（' in lang_name else lang_name[
                                                                               :2].lower()]):
                                if lang_name not in lang_map:
                                    lang_map[lang_name] = col
                                    print(f"{file_type} 模糊匹配: '{header}' -> '{lang_name}' 在第 {col} 列")
                                break

                print(f"{file_type} 最终语言映射: {lang_map}")
                return lang_map

            source_lang_map = build_lang_col_map(source_ws, "源文件")
            trans_lang_map = build_lang_col_map(trans_ws, "翻译文件")

            # 构建对比语言列表 - 基于两个文件都存在的语言
            comparison_langs = []
            for lang in selected_languages:
                if lang in source_lang_map and lang in trans_lang_map:
                    comparison_langs.append(lang)
                    print(
                        f"将对比语言: {lang} (源文件第 {source_lang_map[lang]} 列, 翻译文件第 {trans_lang_map[lang]} 列)")
                else:
                    if lang not in source_lang_map:
                        print(f"警告: 源文件中未找到语言列: {lang}")
                    if lang not in trans_lang_map:
                        print(f"警告: 翻译文件中未找到语言列: {lang}")

            print(f"最终对比语言列表: {comparison_langs}")

            # 先尝试键名精确匹配
            def find_translation_by_key(source_key, trans_key_map):
                """通过键名在翻译文件中查找匹配的行"""
                return trans_key_map.get(source_key)

            # 如果键名不匹配，再基于内容进行模糊匹配
            def find_translation_by_content(source_text, trans_ws, trans_lang_map, lang):
                """通过内容在翻译文件中查找匹配的翻译，使用相似度匹配"""
                if not source_text or not source_text.strip():
                    return None, None, 0.0

                trans_col = trans_lang_map.get(lang)
                if not trans_col:
                    return None, None, 0.0

                source_text_clean = source_text.strip()
                best_match = None
                best_row = None
                best_similarity = 0.0

                # 在翻译文件中搜索最佳匹配的内容
                for row in range(2, trans_ws.max_row + 1):
                    try:
                        trans_cell = trans_ws.cell(row=row, column=trans_col)
                        trans_value = str(trans_cell.value) if trans_cell.value is not None else ""
                        trans_value_clean = trans_value.strip()

                        if trans_value_clean:
                            # 计算相似度
                            similarity = SequenceMatcher(None, source_text_clean, trans_value_clean).ratio()

                            # 记录最佳匹配
                            if similarity > best_similarity:
                                best_similarity = similarity
                                best_row = row
                                key_cell = trans_ws.cell(row=row, column=1)  # 第一列是Key
                                best_match = str(key_cell.value) if key_cell.value is not None else ""
                    except Exception as e:
                        continue

                return best_match, best_row, best_similarity

            # 执行对比
            differences_found = False
            output_row = 2
            total_comparisons = 0
            content_matches = 0
            no_matches = 0
            need_confirm = 0

            if not comparison_langs:
                output_ws.cell(row=2, column=1).value = "错误：未找到可对比的语言列"
                output_ws.cell(row=2, column=1).font = Font(bold=True, color='FF0000')
            else:
                for i, source_key in enumerate(source_key_map.keys()):
                    src_row = source_key_map[source_key]

                    # 记录这个键是否进行了任何对比
                    key_compared = False

                    for lang in comparison_langs:
                        # 检查语言列在源文件中是否存在
                        source_col = source_lang_map.get(lang)
                        if not source_col:
                            continue

                        # 获取源文件值
                        src_val = ""
                        if src_row:
                            try:
                                src_cell = source_ws.cell(row=src_row, column=source_col)
                                src_val = str(src_cell.value) if src_cell.value is not None else ""
                            except Exception as e:
                                src_val = "【读取错误】"

                        # 先尝试键名精确匹配
                        trans_row = find_translation_by_key(source_key, trans_key_map)
                        trans_key = source_key if trans_row else None
                        key_match_type = "键名匹配" if trans_row else "键名不匹配"

                        # 如果键名不匹配，再尝试内容模糊匹配
                        if not trans_row:
                            trans_key, trans_row, similarity = find_translation_by_content(src_val, trans_ws,
                                                                                           trans_lang_map, lang)
                            key_match_type = "内容匹配" if trans_row else "无匹配"

                        # 获取翻译文件值
                        trans_val = ""
                        trans_col = trans_lang_map.get(lang)

                        # 计算内容相似度（无论键名是否匹配，都进行内容分析）
                        content_similarity = 0.0
                        if trans_row and trans_col and src_val.strip():
                            try:
                                trans_cell = trans_ws.cell(row=trans_row, column=trans_col)
                                trans_content = str(trans_cell.value) if trans_cell.value is not None else ""
                                if trans_content.strip():
                                    content_similarity = SequenceMatcher(None, src_val.strip(),
                                                                         trans_content.strip()).ratio()
                            except Exception as e:
                                pass

                        if trans_row and trans_col:
                            try:
                                trans_cell = trans_ws.cell(row=trans_row, column=trans_col)
                                trans_val = str(trans_cell.value) if trans_cell.value is not None else ""

                                # 根据匹配类型和内容相似度判断结果
                                if key_match_type == "键名匹配":
                                    # 键名匹配成功
                                    if content_similarity >= 0.8:
                                        content_matches += 1
                                        trans_val_status = "一致"
                                    elif content_similarity >= 0.5:
                                        # 键名匹配但内容需要确认
                                        trans_val = f"【键名匹配，内容相似度{content_similarity:.0%}需确认】{trans_val}"
                                        trans_val_status = "键名匹配，内容需确认"
                                        need_confirm += 1
                                    else:
                                        # 键名匹配但内容差异大
                                        trans_val = f"【键名匹配，内容相似度{content_similarity:.0%}不匹配】{trans_val}"
                                        trans_val_status = "键名匹配，内容不匹配"
                                else:
                                    # 内容匹配的情况
                                    if content_similarity >= 0.8:
                                        content_matches += 1
                                        trans_val_status = "一致"
                                    elif content_similarity >= 0.5:
                                        # 50%-80%相似度需要人工确认
                                        trans_val = f"【内容相似度{content_similarity:.0%}需人工确认】{trans_val}"
                                        trans_val_status = "需人工确认"
                                        need_confirm += 1
                                    else:
                                        # 低于50%视为不匹配
                                        trans_val = f"【内容相似度{content_similarity:.0%}不匹配】{trans_val}"
                                        trans_val_status = "不匹配"
                            except Exception as e:
                                trans_val = "【读取错误】"
                                trans_val_status = "读取错误"
                        elif not trans_row:
                            trans_val = "【内容未匹配】"
                            trans_val_status = "内容未匹配"
                            no_matches += 1
                        elif not trans_col:
                            trans_val = "【对比文件缺失语言列】"
                            trans_val_status = "缺失语言列"
                        else:
                            trans_val = "【未知错误】"
                            trans_val_status = "未知错误"

                        # 写入键名（只在该语言组的第一行写入）
                        if not key_compared:
                            output_ws.cell(row=output_row, column=1).value = source_key
                            output_ws.cell(row=output_row, column=1).font = Font(bold=True)
                            output_ws.cell(row=output_row, column=2).value = trans_key if trans_key else "无匹配键名"
                            key_compared = True

                        # 写入语言标签
                        output_ws.cell(row=output_row, column=3).value = lang

                        # 写入源内容
                        output_ws.cell(row=output_row, column=4).value = src_val

                        # 写入翻译内容
                        output_ws.cell(row=output_row, column=5).value = trans_val

                        # 对比并标记结果
                        if "需人工确认" in trans_val:
                            output_ws.cell(row=output_row, column=6).value = "⚠ 需人工确认"
                            output_ws.cell(row=output_row, column=6).font = CONFIRM_FONT
                            output_ws.cell(row=output_row, column=4).fill = ORANGE_FILL
                            output_ws.cell(row=output_row, column=5).fill = ORANGE_FILL
                            differences_found = True
                        elif "不匹配" in trans_val:
                            output_ws.cell(row=output_row, column=6).value = "✗ 不匹配"
                            output_ws.cell(row=output_row, column=6).font = DIFF_FONT
                            output_ws.cell(row=output_row, column=4).fill = RED_FILL
                            output_ws.cell(row=output_row, column=5).fill = RED_FILL
                            differences_found = True
                        elif "内容未匹配" in trans_val:
                            output_ws.cell(row=output_row, column=6).value = "⚠ 内容未匹配"
                            output_ws.cell(row=output_row, column=6).font = MISSING_FONT
                            output_ws.cell(row=output_row, column=4).fill = YELLOW_FILL
                            differences_found = True
                        elif "缺失语言列" in trans_val:
                            output_ws.cell(row=output_row, column=6).value = "⚠ 语言列缺失"
                            output_ws.cell(row=output_row, column=6).font = MISSING_FONT
                            output_ws.cell(row=output_row, column=4).fill = YELLOW_FILL
                            differences_found = True
                        elif src_val == trans_val:
                            output_ws.cell(row=output_row, column=6).value = "✓ 一致"
                            output_ws.cell(row=output_row, column=6).font = MATCH_FONT
                            output_ws.cell(row=output_row, column=4).fill = GREEN_FILL
                            output_ws.cell(row=output_row, column=5).fill = GREEN_FILL
                        else:
                            output_ws.cell(row=output_row, column=6).value = "✗ 差异"
                            output_ws.cell(row=output_row, column=6).font = DIFF_FONT
                            output_ws.cell(row=output_row, column=4).fill = RED_FILL
                            output_ws.cell(row=output_row, column=5).fill = RED_FILL
                            differences_found = True

                        output_row += 1
                        total_comparisons += 1

                    # 如果这个键进行了对比，添加空行分隔
                    if key_compared:
                        output_row += 1

                    # 更新进度
                    progress = 40 + (i / len(source_key_map)) * 50
                    if i % 10 == 0 or i == len(source_key_map) - 1:
                        self.update_progress(progress, f"正在对比 {i + 1}/{len(source_key_map)}...")

            # 添加总结信息
            if total_comparisons > 0:
                summary_row = output_row + 2
                output_ws.cell(row=summary_row, column=1).value = "对比总结"
                output_ws.cell(row=summary_row, column=1).font = HEADER_FONT

                summary_text = [
                    f"源文件键数: {len(source_key_map)}",
                    f"完成对比: {total_comparisons} 次",
                    f"内容匹配: {content_matches} 个",
                    f"需人工确认: {need_confirm} 个",
                    f"未匹配: {no_matches} 个"
                ]

                if differences_found:
                    summary_text.append("发现差异内容")
                else:
                    summary_text.append("所有内容完全一致")

                for i, text in enumerate(summary_text):
                    output_ws.cell(row=summary_row + 1 + i, column=1).value = text

            else:
                output_ws.cell(row=2, column=1).value = "未找到可对比的数据"
                output_ws.cell(row=2, column=1).font = Font(bold=True, color='FF0000')

            self.update_progress(95, "正在保存结果...")
            output_wb.save(output_file)

            msg = "对比完成！\n\n"
            msg += f"结果已保存到:\n{os.path.basename(output_file)}"
            msg += f"\n\n统计信息:"
            msg += f"\n- 源文件键数: {len(source_key_map)}"
            msg += f"\n- 完成对比: {total_comparisons} 次"
            msg += f"\n- 内容匹配: {content_matches} 个"
            msg += f"\n- 未匹配: {no_matches} 个"

            self.update_progress(100, "对比完成！")
            messagebox.showinfo("完成", msg)

        except Exception as e:
            self.logger.error(f"对比过程中出错: {e}")
            import traceback
            traceback.print_exc()
            self.update_progress(0, f"错误: {str(e)}")
            messagebox.showerror("错误", f"对比失败:\n{str(e)}")
        finally:
            self.root.after(2000, lambda: self.update_progress(0, "准备就绪"))

    # ==================== 功能三：错误码校对 ====================
    def start_error_code_check_thread(self):
        """启动错误码校对线程"""
        thread = threading.Thread(target=self.error_code_check, daemon=True)
        thread.start()

    def extract_source_error_code(self, key):
        """从源文件提取错误码 - 只提取4位纯数字，不管有没有Key前缀"""
        if not key:
            return None

        key_str = str(key).strip()

        # 从任意位置提取4位连续数字
        match = re.search(r'\d{4}', key_str)
        if not match:
            return None

        digits = match.group(0)

        # 返回原始数字和数值
        return {
            'original': digits,
            'numeric': int(digits)
        }

    def extract_trans_error_code(self, key):
        """从对比文件提取错误码 - 只提取带Key前缀的3-4位数字"""
        if not key:
            return None

        key_str = str(key).strip()

        # 只处理带Key前缀的键（不区分大小写）
        if not re.match(r'^Key\d+', key_str, re.IGNORECASE):
            return None

        # 从Key开头提取连续的3-4位数字（允许001-9999）
        match = re.match(r'^[^\d]*(\d{3,4})', key_str)
        if not match:
            return None

        digits = match.group(1)

        # 返回原始数字、数值和补零版本（用于3位数字补零）
        return {
            'original': digits,
            'numeric': int(digits),
            'padded': digits.zfill(4)  # 确保4位，前面补0
        }

    def extract_error_code_flexible(self, key):
        """兼容提取错误码的方法，避免旧代码引用缺失导致异常。
        - 优先提取以Key前缀的3-4位数字（大小写不敏感）
        - 若无Key前缀，尝试提取任意位置的4位数字
        - 若仍无4位数字，尝试提取独立的3位数字并补零到4位
        返回: {original, numeric, padded} 或 None
        """
        if not key:
            return None

        key_str = str(key).strip()

        # 1) Key前缀 3-4 位数字
        m_key = re.match(r'^[Kk]ey[^\d]*(\d{3,4})', key_str)
        if m_key:
            digits = m_key.group(1)
            return {
                'original': digits,
                'numeric': int(digits),
                'padded': digits.zfill(4)
            }

        # 2) 任意位置 4 位数字
        m4 = re.search(r'\d{4}', key_str)
        if m4:
            digits = m4.group(0)
            return {
                'original': digits,
                'numeric': int(digits),
                'padded': digits  # 已是4位
            }

        # 3) 独立的 3 位数字，补零到4位
        m3 = re.search(r'\b(\d{3})\b', key_str)
        if m3:
            digits = m3.group(1)
            return {
                'original': digits,
                'numeric': int(digits),
                'padded': digits.zfill(4)
            }

        return None

    def find_matching_trans_key_improved(self, source_key, trans_keys):
        """改进的错误码匹配 - 源文件4位数字与对比文件Key前缀4位数字做模糊匹配"""
        # 源文件只提取4位纯数字
        source_digits_info = self.extract_source_error_code(source_key)
        if not source_digits_info:
            return None, []

        source_numeric = source_digits_info['numeric']

        best_match = None
        best_score = 0
        all_matches = []  # 存储所有匹配结果

        self.logger.debug(f"为源文件键 '{source_key}' 查找匹配，数字: {source_numeric}")

        for trans_key in trans_keys:
            # 对比文件只提取带Key前缀的4位数字
            trans_digits_info = self.extract_trans_error_code(trans_key)
            if not trans_digits_info:
                continue

            trans_numeric = trans_digits_info['numeric']
            trans_padded = trans_digits_info['padded']

            # 数字模糊匹配：源文件数字与对比文件Key数字（3位自动补0）匹配
            match_found = (
                # 数值完全匹配
                    source_numeric == trans_numeric or
                    # 对比文件3位数字补0匹配（如101补成0101）
                    source_numeric == int(trans_padded)
            )

            if match_found:
                # 计算键名相似度
                score = SequenceMatcher(None, source_key.lower(), trans_key.lower()).ratio()

                # 记录所有匹配结果
                match_info = {
                    'trans_key': trans_key,
                    'score': score,
                    'source_digits': str(source_numeric),
                    'trans_digits': f"{trans_numeric}({trans_padded})"
                }
                all_matches.append(match_info)

                self.logger.debug(
                    f"候选匹配: {source_key} -> {trans_key} (数字: {source_numeric} == {trans_numeric}({trans_padded}), 相似度: {score:.2f})")

                if score > best_score:
                    best_score = score
                    best_match = trans_key

        if all_matches:
            self.logger.info(f"找到 {len(all_matches)} 个匹配: {source_key} -> {[m['trans_key'] for m in all_matches]}")
            # 返回最佳匹配和所有匹配结果
            return best_match, all_matches
        else:
            self.logger.debug(f"未找到匹配: {source_key}")
            return None, []

    def error_code_check(self):
        """错误码多语言校对功能 - 支持多语言对比"""
        try:
            # 获取用户选择的语言
            selected_languages = self.get_selected_languages()
            self.update_progress(0, f"错误码校对 - 已选择语言: {', '.join(selected_languages)}")

            self.update_progress(5, "正在选择源文件...")
            source_file = filedialog.askopenfilename(
                title="选择源Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not source_file:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(10, "正在选择翻译文件...")
            trans_file = filedialog.askopenfilename(
                title="选择翻译Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not trans_file:
                self.update_progress(0, "操作已取消")
                return

            output_filename = self.get_output_filename("error_code_check")
            output_file = filedialog.asksaveasfilename(
                title="保存错误码校对结果",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_filename
            )
            # 若用户取消保存对话框，使用默认文件名保存在当前工作目录
            if not output_file:
                self.logger.warning("未选择保存路径，使用默认文件名保存到当前目录。")
                output_file = os.path.join(os.getcwd(), output_filename)
            if not output_file:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(20, "正在加载文件...")

            # 加载工作簿
            source_wb = openpyxl.load_workbook(source_file)
            trans_wb = openpyxl.load_workbook(trans_file)
            output_wb = openpyxl.Workbook()

            source_ws = source_wb.active
            trans_ws = trans_wb.active
            output_ws = output_wb.active
            output_ws.title = "错误码校对结果"

            # 定义颜色和样式
            RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            GREEN_FILL = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            YELLOW_FILL = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            ORANGE_FILL = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')  # 橙色，用于需要人工确认
            HEADER_FONT = Font(bold=True, size=12)
            DIFF_FONT = Font(bold=True, color='FF0000')
            MATCH_FONT = Font(color='006400')
            MISSING_FONT = Font(bold=True, color='FFA500')
            CONFIRM_FONT = Font(bold=True, color='FF8C00')  # 深橙色，用于人工确认
            HEADER_FILL = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')

            # 设置列宽
            column_widths = {
                'A': 25,  # 源文件键名
                'B': 25,  # 翻译文件键名
                'C': 15,  # 错误码
                'D': 15,  # 匹配类型
                'E': 20,  # 语言
                'F': 40,  # 源文件内容
                'G': 40,  # 翻译文件内容
                'H': 15  # 对比结果
            }

            for col, width in column_widths.items():
                output_ws.column_dimensions[col].width = width

            # 添加表头
            headers = ["源文件键名", "翻译文件键名", "错误码", "匹配类型", "语言", "源文件内容", "翻译文件内容",
                       "对比结果"]
            output_ws.append(headers)

            # 设置表头样式
            for col in range(1, len(headers) + 1):
                cell = output_ws.cell(row=1, column=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL

            # 构建键映射 - 使用第一列作为键名
            def build_key_map(ws, file_type):
                """构建键名到行号的映射"""
                key_map = {}

                for row in range(2, ws.max_row + 1):
                    key_cell = ws.cell(row=row, column=1)
                    key_value = key_cell.value

                    if key_value is not None:
                        try:
                            key_str = str(key_value).strip()
                            if key_str:
                                key_map[key_str] = row
                        except Exception as e:
                            self.logger.warning(f"处理{file_type}文件行{row}时出错: {e}")

                return key_map

            # 直接映射表：源文件列名 -> 标准语言名
            def build_lang_col_map(ws, file_type):
                """构建语言列映射 - 使用直接映射表"""
                lang_map = {}
                all_headers = []

                # 收集所有表头信息
                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col).value
                    if header_value is not None:
                        header_str = str(header_value).strip()
                        all_headers.append(f"第{col}列: '{header_str}'")

                self.logger.info(f"{file_type}文件表头: {', '.join(all_headers)}")

                # 直接映射表：列名 -> 标准语言名
                header_mapping = {
                    # 源文件列名映射
                    "中文(CN)": "中文（CN）",
                    "英文（EN）English": "英文（EN）English",
                    "德语(DE)": "德语(DE)Deutsch",
                    "西语（ES）": "西语（ES）Español",
                    "法语(FR)": "法语(FR)Français",
                    "意大利语(IT)": "意大利语(IT)Italiano",
                    "葡萄牙语(BR)": "巴西葡语(BR)Português",
                    "俄语（Pyc）": "俄语（Pyc）Русский",
                    "土耳其语(TR)": "土耳其语(TR)Turkish",
                    "日语(JP)": "日语(JP)日本語",
                    "韩语(KR)": "韩语(KR)한국어",
                    "阿拉伯语(xx)": "阿拉伯语عربية",
                    "繁体中文": "繁体中文",

                    # 对比文件列名映射（保持不变）
                    "中文（CN）": "中文（CN）",
                    "英文（EN）English": "英文（EN）English",
                    "德语(DE)Deutsch": "德语(DE)Deutsch",
                    "西语（ES）Español": "西语（ES）Español",
                    "法语(FR)Français": "法语(FR)Français",
                    "意大利语(IT)Italiano": "意大利语(IT)Italiano",
                    "巴西葡语(BR)Português": "巴西葡语(BR)Português",
                    "俄语（Pyc）Русский": "俄语（Pyc）Русский",
                    "土耳其语(TR)Turkish": "土耳其语(TR)Turkish",
                    "日语(JP)日本語": "日语(JP)日本語",
                    "韩语(KR)한국어": "韩语(KR)한국어",
                    "繁体中文": "繁体中文",
                    "阿拉伯语عربية": "阿拉伯语عربية"
                }

                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col).value
                    if header_value is None:
                        continue

                    header = str(header_value).strip()

                    # 直接查找映射
                    if header in header_mapping:
                        lang_name = header_mapping[header]
                        if lang_name not in lang_map:
                            lang_map[lang_name] = col
                            self.logger.info(f"{file_type}文件: 列 '{header}' -> 语言 '{lang_name}' 在第 {col} 列")

                return lang_map

            # 构建键映射
            source_key_map = build_key_map(source_ws, "源文件")
            trans_key_map = build_key_map(trans_ws, "对比文件")

            # 构建语言列映射
            source_lang_map = build_lang_col_map(source_ws, "源文件")
            trans_lang_map = build_lang_col_map(trans_ws, "对比文件")

            # 构建对比语言列表 - 只使用用户选择的语言
            comparison_langs = []
            for lang in selected_languages:
                if lang in source_lang_map:
                    comparison_langs.append(lang)
                    self.logger.info(f"将对比语言: {lang}")
                else:
                    self.logger.warning(f"源文件中未找到语言列: {lang}")

            # 强制确保中文参与对比（如果存在对应列且未被选择）
            try:
                if "中文（CN）" in source_lang_map and "中文（CN）" not in comparison_langs:
                    comparison_langs.insert(0, "中文（CN）")
                    self.logger.info("已自动添加中文（CN）到对比语言列表")
            except Exception:
                pass

            self.logger.info(f"源文件键数: {len(source_key_map)}")
            self.logger.info(f"对比文件键数: {len(trans_key_map)}")
            self.logger.info(f"最终对比语言: {comparison_langs}")

            # 执行错误码匹配和多语言对比
            output_row = 2
            exact_matches = 0
            no_matches = 0
            need_confirm = 0
            total_comparisons = 0
            total_source_keys = len(source_key_map)

            # 将每次语言写入封装为函数，确保每语言占独立一行，避免覆盖
            def write_language_row(row_idx, write_key_info, source_key_val, trans_key_val, source_err_disp,
                                   match_type_val, lang_label, src_text, trans_text):
                # 写入键相关信息（只在该键的第一行写入）
                if write_key_info:
                    if source_key_val and "\\" in source_key_val:
                        output_ws.cell(row=row_idx, column=1).value = source_key_val
                        output_ws.column_dimensions['A'].width = max(20, len(source_key_val) * 1.2)
                    else:
                        output_ws.cell(row=row_idx, column=1).value = source_key_val
                    output_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                    output_ws.cell(row=row_idx, column=2).value = trans_key_val
                    output_ws.cell(row=row_idx, column=3).value = source_err_disp
                    output_ws.cell(row=row_idx, column=4).value = match_type_val

                # 写入语言及内容
                output_ws.cell(row=row_idx, column=5).value = lang_label
                output_ws.cell(row=row_idx, column=6).value = src_text
                output_ws.cell(row=row_idx, column=7).value = trans_text

                # 计算相似度与结果
                content_similarity = 0.0
                if trans_text not in ["【对比文件缺失该键】", "【语言列缺失】", "【读取错误】"] and src_text and trans_text:
                    content_similarity = SequenceMatcher(None, src_text, trans_text).ratio()

                result_text_local = ""
                if trans_text == "【对比文件缺失该键】":
                    result_text_local = "⚠ 缺失"
                    output_ws.cell(row=row_idx, column=8).font = MISSING_FONT
                    output_ws.cell(row=row_idx, column=6).fill = YELLOW_FILL
                    output_ws.cell(row=row_idx, column=7).fill = YELLOW_FILL
                elif trans_text == "【语言列缺失】":
                    result_text_local = "⚠ 语言列缺失"
                    output_ws.cell(row=row_idx, column=8).font = MISSING_FONT
                    output_ws.cell(row=row_idx, column=6).fill = YELLOW_FILL
                    output_ws.cell(row=row_idx, column=7).fill = YELLOW_FILL
                elif src_text == trans_text:
                    result_text_local = "✓ 一致"
                    output_ws.cell(row=row_idx, column=8).font = MATCH_FONT
                    output_ws.cell(row=row_idx, column=6).fill = GREEN_FILL
                    output_ws.cell(row=row_idx, column=7).fill = GREEN_FILL
                elif content_similarity >= 0.8:
                    result_text_local = "✓ 高相似度匹配"
                    output_ws.cell(row=row_idx, column=8).font = MATCH_FONT
                    output_ws.cell(row=row_idx, column=6).fill = GREEN_FILL
                    output_ws.cell(row=row_idx, column=7).fill = GREEN_FILL
                elif content_similarity >= 0.5:
                    result_text_local = f"⚠ 需人工确认（相似度{content_similarity:.0%})"
                    output_ws.cell(row=row_idx, column=8).font = CONFIRM_FONT
                    output_ws.cell(row=row_idx, column=6).fill = ORANGE_FILL
                    output_ws.cell(row=row_idx, column=7).fill = ORANGE_FILL
                else:
                    result_text_local = f"✗ 不匹配（相似度{content_similarity:.0%})"
                    output_ws.cell(row=row_idx, column=8).font = DIFF_FONT
                    output_ws.cell(row=row_idx, column=6).fill = RED_FILL
                    output_ws.cell(row=row_idx, column=7).fill = RED_FILL

                output_ws.cell(row=row_idx, column=8).value = result_text_local
                self.logger.info(f"已写入第 {row_idx} 行，语言={lang_label}")
                return row_idx + 1

            for i, source_key in enumerate(source_key_map.keys()):
                progress = 30 + (i / total_source_keys) * 60
                self.update_progress(progress, f"正在处理 {i + 1}/{total_source_keys}...")

                # 使用改进的错误码匹配
                trans_key_match, all_matches = self.find_matching_trans_key_improved(source_key, trans_key_map.keys())

                # 获取源文件行号
                source_row = source_key_map[source_key]

                # 提取错误码信息用于显示
                source_digits_info = self.extract_source_error_code(source_key)
                source_error_display = str(source_digits_info['numeric']) if source_digits_info else "无错误码"

                # 处理多个匹配结果
                if all_matches and len(all_matches) > 1:
                    # 有多个匹配结果，需要显示所有匹配
                    self.logger.info(f"源键 '{source_key}' 找到 {len(all_matches)} 个匹配结果")

                    # 为每个匹配结果创建对比行
                    for match_idx, match_info in enumerate(all_matches):
                        trans_key = match_info['trans_key']
                        trans_row = trans_key_map[trans_key]

                        # 记录这个键是否进行了任何对比
                        key_compared = False

                        # 对每个选中的语言进行对比
                        for lang in comparison_langs:
                            # 检查语言列在源文件中是否存在
                            source_col = source_lang_map.get(lang)
                            trans_col = trans_lang_map.get(lang)

                            if not source_col:
                                self.logger.warning(f"源文件中未找到语言列: {lang}")
                                continue

                            # 获取源文件内容
                            source_content = ""
                            if source_row:
                                try:
                                    source_cell = source_ws.cell(row=source_row, column=source_col)
                                    source_content = str(source_cell.value) if source_cell.value is not None else ""
                                except Exception as e:
                                    source_content = "【读取错误】"
                                    self.logger.error(f"读取源文件内容错误: {e}")

                            # 获取翻译文件内容
                            trans_content = ""
                            if trans_row and trans_col:
                                try:
                                    trans_cell = trans_ws.cell(row=trans_row, column=trans_col)
                                    trans_content = str(trans_cell.value) if trans_cell.value is not None else ""
                                except Exception as e:
                                    trans_content = "【读取错误】"
                                    self.logger.error(f"读取翻译文件内容错误: {e}")
                            elif trans_row and not trans_col:
                                trans_content = "【语言列缺失】"
                                self.logger.warning(f"对比文件中未找到语言列: {lang}")
                            else:
                                trans_content = "【对比文件缺失该键】"

                            # 记录本次语言处理的简要信息，便于定位中文未输出问题
                            try:
                                sc_preview = (source_content or "")[0:30]
                                tc_preview = (trans_content or "")[0:30]
                                self.logger.info(
                                    f"语言处理: lang='{lang}', 源列={source_col}, 对比列={trans_col}, 源预览='{sc_preview}', 对比预览='{tc_preview}'"
                                )
                            except Exception:
                                pass

                            # 写入一行（独立行）
                            match_display = f"{trans_key} (匹配{match_idx + 1}/{len(all_matches)})"
                            output_row = write_language_row(
                                output_row,
                                not key_compared,
                                source_key,
                                match_display,
                                source_error_display,
                                f"错误码匹配 ({match_info['score']:.0%})",
                                lang,
                                source_content,
                                trans_content
                            )
                            exact_matches += 1 if not key_compared else 0
                            key_compared = True
                            total_comparisons += 1

                        # 在每个匹配结果之间添加分隔行
                        if key_compared and match_idx < len(all_matches) - 1:
                            output_row += 1
                else:
                    # 只有一个匹配结果或没有匹配，逐语言逐行写入
                    trans_row = trans_key_map[trans_key_match] if trans_key_match else None

                    # 标记此键是否已在任何语言中写入过键信息
                    key_compared = False

                    # 对每个选中的语言进行对比并写入独立行
                    for lang in comparison_langs:
                        # 检查语言列在源文件中是否存在
                        source_col = source_lang_map.get(lang)
                        trans_col = trans_lang_map.get(lang)

                        if not source_col:
                            self.logger.warning(f"源文件中未找到语言列: {lang}")
                            continue

                        # 获取源文件内容
                        source_content = ""
                        if source_row:
                            try:
                                source_cell = source_ws.cell(row=source_row, column=source_col)
                                source_content = str(source_cell.value) if source_cell.value is not None else ""
                            except Exception as e:
                                source_content = "【读取错误】"
                                self.logger.error(f"读取源文件内容错误: {e}")

                        # 获取翻译文件内容
                        trans_content = ""
                        if trans_row and trans_col:
                            try:
                                trans_cell = trans_ws.cell(row=trans_row, column=trans_col)
                                trans_content = str(trans_cell.value) if trans_cell.value is not None else ""
                            except Exception as e:
                                trans_content = "【读取错误】"
                                self.logger.error(f"读取翻译文件内容错误: {e}")
                        elif trans_row and not trans_col:
                            trans_content = "【语言列缺失】"
                            self.logger.warning(f"对比文件中未找到语言列: {lang}")
                        else:
                            trans_content = "【对比文件缺失该键】"
                            no_matches += 1
                            # 特殊处理：如果键名包含反斜杠，保留原始格式
                            if source_key and "\\" in source_key:
                                # 保留原始格式，不强制分行显示
                                pass

                        # 记录本次语言处理的简要信息，便于定位中文未输出问题
                        try:
                            sc_preview = (source_content or "")[0:30]
                            tc_preview = (trans_content or "")[0:30]
                            self.logger.info(
                                f"语言处理: lang='{lang}', 源列={source_col}, 对比列={trans_col}, 源预览='{sc_preview}', 对比预览='{tc_preview}'"
                            )
                        except Exception:
                            pass

                        # 写入一行（独立行），仅在该键的第一行写入键相关信息
                        match_type = "错误码匹配" if trans_key_match else "无匹配"
                        output_row = write_language_row(
                            output_row,
                            not key_compared,
                            source_key,
                            trans_key_match if trans_key_match else "无匹配",
                            source_error_display,
                            match_type,
                            lang,
                            source_content,
                            trans_content
                        )

                        if trans_key_match and not key_compared:
                            exact_matches += 1
                        key_compared = True
                        total_comparisons += 1

                # 如果这个键进行了对比，添加空行分隔
                if key_compared:
                    output_row += 1

            # 添加总结信息
            summary_row = output_row + 2
            output_ws.cell(row=summary_row, column=1).value = "错误码校对总结"
            output_ws.cell(row=summary_row, column=1).font = HEADER_FONT

            summary_data = [
                f"源文件总键数: {total_source_keys}",
                f"错误码匹配: {exact_matches}",
                f"文案需人工确认: {need_confirm}",
                f"无匹配: {no_matches}",
                f"对比语言: {', '.join(comparison_langs)}",
                f"完成对比: {total_comparisons} 次"
            ]

            for i, text in enumerate(summary_data):
                output_ws.cell(row=summary_row + 1 + i, column=1).value = text

            # 统计语言写入分布，帮助定位缺失情况
            try:
                from collections import Counter
                lang_counter = Counter()
                for r in range(2, output_row):
                    val = output_ws.cell(row=r, column=5).value
                    if val:
                        lang_counter[val] += 1
                self.logger.info(f"语言写入分布: {dict(lang_counter)}")
            except Exception as e:
                self.logger.warning(f"统计语言分布时出错: {e}")

            self.update_progress(95, "正在保存错误码校对结果...")
            self.logger.info(f"保存错误码校对结果到: {output_file}")
            output_wb.save(output_file)
            self.logger.info("错误码校对结果已成功保存。")

            msg = "错误码校对完成！\n\n"
            msg += f"统计信息:\n"
            msg += f"- 源文件总键数: {total_source_keys}\n"
            msg += f"- 错误码匹配: {exact_matches}\n"
            msg += f"- 文案需人工确认: {need_confirm}\n"
            msg += f"- 无匹配: {no_matches}\n"
            msg += f"- 对比语言: {', '.join(comparison_langs)}\n"
            msg += f"- 完成对比: {total_comparisons} 次\n\n"
            msg += f"结果已保存到:\n{os.path.basename(output_file)}"

            self.update_progress(100, "错误码校对完成！")
            messagebox.showinfo("完成", msg)

        except Exception as e:
            self.logger.error(f"错误码校对过程中出错: {e}")
            import traceback
            traceback.print_exc()
            self.update_progress(0, f"错误: {str(e)}")
            messagebox.showerror("错误", f"错误码校对失败:\n{str(e)}")
        finally:
            self.root.after(2000, lambda: self.update_progress(0, "准备就绪"))

    # ==================== 通用方法 ====================
    def select_all_languages(self):
        """全选所有语言"""
        for var in self.language_vars:
            var.set(1)

    def deselect_all_languages(self):
        """取消全选所有语言"""
        for var in self.language_vars:
            var.set(0)

    def get_selected_languages(self):
        """获取用户选择的语言列表"""
        selected = []
        for i, var in enumerate(self.language_vars):
            if var.get() == 1:
                selected.append(self.LANGUAGE_ORDER[i])
        return selected if selected else self.LANGUAGE_ORDER

    def get_output_filename(self, suffix):
        """生成带日期时间的输出文件名"""
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{now}_{suffix}.xlsx"

    def update_progress(self, value, message):
        """更新进度条和标签"""
        self.progress['value'] = value
        self.progress_label['text'] = message
        self.root.update_idletasks()

    def start_conversion_thread(self):
        """启动转换线程"""
        thread = threading.Thread(target=self.convert_code_to_excel, daemon=True)
        thread.start()

    def start_comparison_thread(self):
        """启动对比线程"""
        thread = threading.Thread(target=self.compare_excel_files, daemon=True)
        thread.start()


if __name__ == "__main__":
    root = tk.Tk()
    app = CompleteDualDisplayTool(root)
    root.mainloop()