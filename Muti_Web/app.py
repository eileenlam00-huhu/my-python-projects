from flask import Flask, render_template, request, jsonify, send_file
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import os
import logging
from difflib import SequenceMatcher
import tempfile
import threading
import uuid
import shutil
import socket
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = os.path.join(tempfile.gettempdir(), 'multilang_tool')
app.config['MAX_FILES'] = 100  # 最大保留文件数

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 标准语言顺序
LANGUAGE_ORDER = [
    "中文（CN）", "英文（EN）English", "德语(DE)Deutsch",
    "西语（ES）Español", "法语(FR)Français", "意大利语(IT)Italiano",
    "巴西葡语(BR)Português", "俄语（Pyc）Русский", "土耳其语(TR)Turkish",
    "日语(JP)日本語", "韩语(KR)한국어", "阿拉伯语عربية", "繁体中文",
    "波兰语（PL）Polski","越南语（VI）Tiếng Việt","印尼语（ID）Bahasa Indonesia",
    "泰语（TH）ไทย","马来语（MS）Bahasa Melayu","希伯来语（HE）עברית","南非语（AF）Afrikaans"
]

# 存储任务状态
tasks = {}


def get_local_ip():
    """获取本机局域网IP"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception as e:
        logger.warning(f"获取局域网IP失败: {e}")
        return '127.0.0.1'


def cleanup_old_files():
    """清理旧文件，避免占用太多空间"""
    try:
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            return

        files = [os.path.join(app.config['UPLOAD_FOLDER'], f)
                 for f in os.listdir(app.config['UPLOAD_FOLDER'])]

        # 过滤出文件（排除目录）
        files = [f for f in files if os.path.isfile(f)]
        files.sort(key=os.path.getctime)

        # 如果文件数量超过限制，删除最旧的
        while len(files) > app.config['MAX_FILES']:
            try:
                oldest_file = files.pop(0)
                os.remove(oldest_file)
                logger.info(f"已清理旧文件: {os.path.basename(oldest_file)}")
            except Exception as e:
                logger.warning(f"清理文件失败: {e}")
                continue

    except Exception as e:
        logger.error(f"清理文件过程出错: {e}")


@app.route('/')
def index():
    """渲染主页面"""
    return render_template('index.html', languages=LANGUAGE_ORDER)


@app.route('/api/convert-code', methods=['POST'])
def convert_code():
    """功能一：代码转Excel"""
    temp_file = None
    try:
        if 'code_file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400

        file = request.files['code_file']
        if file.filename == '':
            return jsonify({'error': '未选择文件'}), 400

        # 保存上传的文件
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{uuid.uuid4().hex}_{filename}")
        file.save(temp_path)

        # 读取文件内容
        with open(temp_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Translations"

        # 添加表头
        headers = ["Key"] + LANGUAGE_ORDER
        ws.append(headers)

        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # 设置列宽
        ws.column_dimensions['A'].width = 30
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # 改进的多语言数组匹配模式
        # 匹配: const char *key[MAX_LANGUAGE] = {"值1", "值2", ...};
        pattern = re.compile(
            r'const\s+char\s*\*\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\[\s*MAX_LANGUAGE\s*\]\s*=\s*\{([^}]*)\}\s*;',
            re.DOTALL | re.MULTILINE
        )

        # 匹配单字符串数组（需要排除的）
        # 匹配: const char *key[MAX_LANGUAGE] = "值";
        single_pattern = re.compile(
            r'const\s+char\s*\*\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\[\s*MAX_LANGUAGE\s*\]\s*=\s*"([^"]*)"\s*;',
            re.DOTALL | re.MULTILINE
        )

        # 记录所有匹配
        matches = []
        single_matches = []

        # 先找出所有单字符串定义（要排除的）
        for single_match in single_pattern.finditer(content):
            key = single_match.group(1).strip()
            single_matches.append(key)
            logger.info(f"排除单字符串定义: {key} = {single_match.group(2)}")

        # 找出多语言数组定义
        for match in pattern.finditer(content):
            key = match.group(1).strip()

            # 跳过单字符串定义
            if key in single_matches:
                logger.info(f"跳过 {key} (已定义为单字符串)")
                continue

            values_block = match.group(2).strip()

            # 检查是否真的包含多个值（至少有一个逗号分隔）
            if ',' not in values_block:
                # 只有一个值的情况，可能是误匹配
                # 进一步检查是否包含多个引号字符串
                values = re.findall(r'"([^"\\]*(?:\\.[^"\\]*)*)"', values_block)
                if len(values) <= 1:
                    logger.info(f"跳过 {key} (只有一个值)")
                    continue

            matches.append(match)

        converted_count = 0

        for i, match in enumerate(matches):
            key = match.group(1).strip()
            values_block = match.group(2)

            # 更精确地提取各语言字符串（处理转义字符）
            values = re.findall(r'"((?:[^"\\]|\\.)*)"', values_block)

            # 如果提取的值少于预期，补充空字符串
            if len(values) < len(headers) - 1:
                values.extend([''] * (len(headers) - 1 - len(values)))
            elif len(values) > len(headers) - 1:
                values = values[:len(headers) - 1]

            # 写入Excel行
            row = [key] + values
            ws.append(row)
            converted_count += 1

            logger.info(f"已转换: {key} -> {len(values)} 个语言值")

        # 如果没有找到任何多语言数组，尝试另一种模式
        if converted_count == 0:
            # 备选模式：匹配更通用的多语言数组定义
            alt_pattern = re.compile(
                r'const\s+char\s*\*\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\[\s*\d+\s*\]\s*=\s*\{([^}]*)\}\s*;',
                re.DOTALL | re.MULTILINE
            )

            for match in alt_pattern.finditer(content):
                key = match.group(1).strip()
                values_block = match.group(2).strip()

                # 检查是否有多个值
                if ',' in values_block:
                    values = re.findall(r'"((?:[^"\\]|\\.)*)"', values_block)
                    if len(values) > 1:
                        if len(values) < len(headers) - 1:
                            values.extend([''] * (len(headers) - 1 - len(values)))
                        elif len(values) > len(headers) - 1:
                            values = values[:len(headers) - 1]

                        row = [key] + values
                        ws.append(row)
                        converted_count += 1
                        logger.info(f"备选模式转换: {key} -> {len(values)} 个语言值")

        # 保存到文件
        output_filename = f"converted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        wb.save(output_path)

        # 清理临时文件
        try:
            os.remove(temp_path)
        except:
            pass

        # 清理旧文件
        cleanup_old_files()

        return jsonify({
            'success': True,
            'message': f'转换成功！共处理 {converted_count} 个多语言条目',
            'filename': output_filename,
            'download_url': f'/download/{output_filename}'
        })

    except UnicodeDecodeError:
        return jsonify({'error': '文件编码错误，请确保文件是UTF-8编码'}), 400
    except Exception as e:
        logger.error(f"转换失败: {str(e)}")
        return jsonify({'error': f'转换失败: {str(e)}'}), 500


@app.route('/api/compare-excel', methods=['POST'])
def compare_excel():
    """功能二：Excel对比 - 启动异步任务"""
    try:
        if 'source_file' not in request.files or 'trans_file' not in request.files:
            return jsonify({'error': '请上传两个Excel文件'}), 400

        source_file = request.files['source_file']
        trans_file = request.files['trans_file']

        if source_file.filename == '' or trans_file.filename == '':
            return jsonify({'error': '请选择有效的文件'}), 400

        # 获取选中的语言
        selected_languages = request.form.getlist('languages[]')
        if not selected_languages:
            selected_languages = LANGUAGE_ORDER  # 默认全选

        # 保存上传的文件
        source_filename = secure_filename(source_file.filename)
        trans_filename = secure_filename(trans_file.filename)

        source_path = os.path.join(app.config['UPLOAD_FOLDER'], f"source_{uuid.uuid4().hex}_{source_filename}")
        trans_path = os.path.join(app.config['UPLOAD_FOLDER'], f"trans_{uuid.uuid4().hex}_{trans_filename}")

        source_file.save(source_path)
        trans_file.save(trans_path)

        # 创建任务
        task_id = uuid.uuid4().hex
        tasks[task_id] = {
            'progress': 0,
            'status': 'processing',
            'message': '任务已创建'
        }

        # 启动后台线程处理
        def process_task():
            try:
                tasks[task_id]['progress'] = 10
                tasks[task_id]['message'] = '正在加载Excel文件...'

                result = compare_excel_files(
                    source_path,
                    trans_path,
                    selected_languages,
                    task_id
                )

                if result['success']:
                    tasks[task_id].update({
                        'progress': 100,
                        'status': 'completed',
                        'message': '对比完成',
                        'filename': result['filename'],
                        'download_url': f'/download/{result["filename"]}'
                    })
                else:
                    tasks[task_id].update({
                        'progress': 0,
                        'status': 'error',
                        'message': result['error']
                    })

            except Exception as e:
                logger.error(f"处理任务失败: {e}")
                tasks[task_id].update({
                    'progress': 0,
                    'status': 'error',
                    'message': str(e)
                })
            finally:
                # 清理临时文件，添加重试机制
                import time
                time.sleep(1)  # 等待文件释放

                for retry in range(3):  # 最多重试3次
                    try:
                        if os.path.exists(source_path):
                            os.remove(source_path)
                            logger.info(f"已删除临时文件: {source_path}")
                        if os.path.exists(trans_path):
                            os.remove(trans_path)
                            logger.info(f"已删除临时文件: {trans_path}")
                        break
                    except Exception as e:
                        if retry < 2:  # 最后一次不等待
                            logger.warning(f"删除临时文件失败，{retry + 1}/3 次重试: {e}")
                            time.sleep(2)  # 等待2秒后重试
                        else:
                            logger.error(f"最终无法删除临时文件: {e}")

        thread = threading.Thread(target=process_task)
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'task_id': task_id
        })

    except Exception as e:
        logger.error(f"启动任务失败: {e}")
        return jsonify({'error': f'启动任务失败: {str(e)}'}), 500


def compare_excel_files(source_path, trans_path, selected_languages, task_id):
    """执行Excel对比的核心逻辑"""
    source_wb = None
    trans_wb = None
    output_wb = None

    try:
        tasks[task_id]['progress'] = 20
        tasks[task_id]['message'] = '正在加载Excel文件...'

        # 加载工作簿
        source_wb = openpyxl.load_workbook(source_path, data_only=True)
        trans_wb = openpyxl.load_workbook(trans_path, data_only=True)
        output_wb = openpyxl.Workbook()

        source_ws = source_wb.active
        trans_ws = trans_wb.active
        output_ws = output_wb.active
        output_ws.title = "对比结果"

        tasks[task_id]['progress'] = 30
        tasks[task_id]['message'] = '正在分析文件结构...'

        # 定义颜色样式
        GREEN_FILL = PatternFill(start_color='8FBC8F', end_color='8FBC8F', fill_type='solid')
        RED_FILL = PatternFill(start_color='CD0000', end_color='CD0000', fill_type='solid')
        YELLOW_FILL = PatternFill(start_color='FFF68F', end_color='FFF68F', fill_type='solid')
        HEADER_FONT = Font(bold=True, size=12)
        HEADER_FILL = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        BOLD_FONT = Font(bold=True)

        # 设置列宽
        output_ws.column_dimensions['A'].width = 30  # 源文件键名
        output_ws.column_dimensions['B'].width = 30  # 翻译文件键名
        output_ws.column_dimensions['C'].width = 20  # 语言
        output_ws.column_dimensions['D'].width = 40  # 源文件内容
        output_ws.column_dimensions['E'].width = 40  # 翻译文件内容
        output_ws.column_dimensions['F'].width = 15  # 对比结果

        # 添加表头
        headers = ["源文件键名", "翻译文件键名", "语言", "源文件内容", "翻译文件内容", "对比结果"]
        output_ws.append(headers)

        # 设置表头样式
        for col in range(1, 7):
            cell = output_ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        # 打印文件信息以便调试
        logger.info("=" * 50)
        logger.info("源文件信息:")
        logger.info(f"最大行: {source_ws.max_row}, 最大列: {source_ws.max_column}")

        # 打印源文件表头
        source_headers = []
        for col in range(1, source_ws.max_column + 1):
            header = source_ws.cell(row=1, column=col).value
            source_headers.append(f"列{col}: {header}")
        logger.info(f"源文件表头: {', '.join(source_headers)}")

        logger.info("\n翻译文件信息:")
        logger.info(f"最大行: {trans_ws.max_row}, 最大列: {trans_ws.max_column}")

        # 打印翻译文件表头
        trans_headers = []
        for col in range(1, trans_ws.max_column + 1):
            header = trans_ws.cell(row=1, column=col).value
            trans_headers.append(f"列{col}: {header}")
        logger.info(f"翻译文件表头: {', '.join(trans_headers)}")
        logger.info("=" * 50)

        # 构建键映射 - 使用第一列作为键名
        def build_key_map(ws, file_name):
            key_map = {}
            key_to_row = {}
            logger.info(f"\n开始构建 {file_name} 键映射:")

            # 明确使用第一列作为键名
            key_column = 1
            logger.info(f"使用第1列作为键名列")

            # 收集所有键名
            keys_found = []
            for row in range(2, ws.max_row + 1):
                key_cell = ws.cell(row=row, column=key_column)
                if key_cell.value:
                    key_str = str(key_cell.value).strip()
                    if key_str:
                        key_map[key_str] = row
                        key_to_row[key_str] = row
                        keys_found.append(key_str)

            # 打印前10个键名用于调试
            logger.info(f"{file_name} 前10个键名: {keys_found[:10]}")
            logger.info(f"{file_name} 共找到 {len(key_map)} 个键")
            return key_map, key_column, key_to_row

        source_key_map, source_key_col, source_key_to_row = build_key_map(source_ws, "源文件")
        trans_key_map, trans_key_col, trans_key_to_row = build_key_map(trans_ws, "翻译文件")

        tasks[task_id]['progress'] = 40
        tasks[task_id]['message'] = f'源文件发现 {len(source_key_map)} 个键，翻译文件发现 {len(trans_key_map)} 个键'

        # 构建语言列映射 - 严格区分中文和繁体中文
        def build_lang_col_map(ws, file_name):
            lang_map = {}
            logger.info(f"\n开始构建 {file_name} 语言列映射:")

            # 首先处理特殊情况：中文和繁体中文必须严格区分
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and isinstance(header, str):
                    header_str = header.strip()

                    # 1. 先检查是否是繁体中文（优先级最高）
                    if header_str == "繁体中文" or "繁体" in header_str:
                        lang_map["繁体中文"] = col
                        logger.info(f"✓ 繁体中文匹配: '{header_str}' -> 繁体中文 (第{col}列)")
                        continue

                    # 2. 检查是否是中文（CN）- 必须包含"CN"或明确的中文标识
                    if (header_str == "中文（CN）" or
                            header_str == "中文(CN)" or
                            (("中文" in header_str or "Chinese" in header_str) and "CN" in header_str) or
                            (header_str == "中文" and ws.max_column > 2)):  # 如果只有"中文"且有多列，假设第一列是中文
                        lang_map["中文（CN）"] = col
                        logger.info(f"✓ 中文(CN)匹配: '{header_str}' -> 中文（CN） (第{col}列)")
                        continue

                    # 3. 检查其他语言
                    header_lower = header_str.lower()

                    # 英文
                    if "英文" in header_str or "english" in header_lower or "en" == header_lower:
                        lang_map["英文（EN）English"] = col
                        logger.info(f"✓ 英文匹配: '{header_str}' -> 英文（EN）English (第{col}列)")

                    # 德语
                    elif "德语" in header_str or "german" in header_lower or "de" == header_lower:
                        lang_map["德语(DE)Deutsch"] = col
                        logger.info(f"✓ 德语匹配: '{header_str}' -> 德语(DE)Deutsch (第{col}列)")

                    # 西语
                    elif "西语" in header_str or "spanish" in header_lower or "es" == header_lower:
                        lang_map["西语（ES）Español"] = col
                        logger.info(f"✓ 西语匹配: '{header_str}' -> 西语（ES）Español (第{col}列)")

                    # 法语
                    elif "法语" in header_str or "french" in header_lower or "fr" == header_lower:
                        lang_map["法语(FR)Français"] = col
                        logger.info(f"✓ 法语匹配: '{header_str}' -> 法语(FR)Français (第{col}列)")

                    # 意大利语
                    elif "意大利语" in header_str or "italian" in header_lower or "it" == header_lower:
                        lang_map["意大利语(IT)Italiano"] = col
                        logger.info(f"✓ 意大利语匹配: '{header_str}' -> 意大利语(IT)Italiano (第{col}列)")

                    # 巴西葡语
                    elif "巴西葡语" in header_str or "portuguese" in header_lower or "br" in header_lower:
                        lang_map["巴西葡语(BR)Português"] = col
                        logger.info(f"✓ 巴西葡语匹配: '{header_str}' -> 巴西葡语(BR)Português (第{col}列)")

                    # 俄语
                    elif "俄语" in header_str or "russian" in header_lower or "ru" in header_lower:
                        lang_map["俄语（Pyc）Русский"] = col
                        logger.info(f"✓ 俄语匹配: '{header_str}' -> 俄语（Pyc）Русский (第{col}列)")

                    # 土耳其语
                    elif "土耳其语" in header_str or "turkish" in header_lower or "tr" in header_lower:
                        lang_map["土耳其语(TR)Turkish"] = col
                        logger.info(f"✓ 土耳其语匹配: '{header_str}' -> 土耳其语(TR)Turkish (第{col}列)")

                    # 日语
                    elif "日语" in header_str or "japanese" in header_lower or "jp" in header_lower:
                        lang_map["日语(JP)日本語"] = col
                        logger.info(f"✓ 日语匹配: '{header_str}' -> 日语(JP)日本語 (第{col}列)")

                    # 韩语
                    elif "韩语" in header_str or "korean" in header_lower or "kr" in header_lower:
                        lang_map["韩语(KR)한국어"] = col
                        logger.info(f"✓ 韩语匹配: '{header_str}' -> 韩语(KR)한국어 (第{col}列)")

                    # 阿拉伯语
                    elif "阿拉伯语" in header_str or "arabic" in header_lower or "ar" in header_lower:
                        lang_map["阿拉伯语عربية"] = col
                        logger.info(f"✓ 阿拉伯语匹配: '{header_str}' -> 阿拉伯语عربية (第{col}列)")

                    # 波兰语（新增）
                    elif "波兰语" in header_str or "polish" in header_lower or "pl" == header_lower:
                        lang_map["波兰语（PL）Polski"] = col
                        logger.info(f"✓ 波兰语匹配: '{header_str}' -> 波兰语（PL）Polski (第{col}列)")

                    # 越南语（新增）
                    elif "越南语" in header_str or "vietnamese" in header_lower or "vi" == header_lower:
                        lang_map["越南语（VI）Tiếng Việt"] = col
                        logger.info(f"✓ 越南语匹配: '{header_str}' -> 越南语（VI）Tiếng Việt (第{col}列)")

                    # 印尼语（新增）
                    elif "印尼语" in header_str or "indonesian" in header_lower or "id" == header_lower:
                        lang_map["印尼语（ID）Bahasa Indonesia"] = col
                        logger.info(f"✓ 印尼语匹配: '{header_str}' -> 印尼语（ID）Bahasa Indonesia (第{col}列)")

                    # 泰语（新增）
                    elif "泰语" in header_str or "thai" in header_lower or "th" == header_lower:
                        lang_map["泰语（TH）ไทย"] = col
                        logger.info(f"✓ 泰语匹配: '{header_str}' -> 泰语（TH）ไทย (第{col}列)")

                    # 马来语（新增）
                    elif "马来语" in header_str or "malay" in header_lower or "ms" == header_lower:
                        lang_map["马来语（MS）Bahasa Melayu"] = col
                        logger.info(f"✓ 马来语匹配: '{header_str}' -> 马来语（MS）Bahasa Melayu (第{col}列)")

                    # 希伯来语（新增）
                    elif "希伯来语" in header_str or "hebrew" in header_lower or "he" == header_lower:
                        lang_map["希伯来语（HE）עברית"] = col
                        logger.info(f"✓ 希伯来语匹配: '{header_str}' -> 希伯来语（HE）עברית (第{col}列)")

                    # 南非语（新增）
                    elif "南非语" in header_str or "afrikaans" in header_lower or "af" == header_lower:
                        lang_map["南非语（AF）Afrikaans"] = col
                        logger.info(f"✓ 南非语匹配: '{header_str}' -> 南非语（AF）Afrikaans (第{col}列)")

            logger.info(f"{file_name} 最终语言映射: {lang_map}")
            return lang_map

        source_lang_map = build_lang_col_map(source_ws, "源文件")
        trans_lang_map = build_lang_col_map(trans_ws, "翻译文件")

        # 筛选可对比的语言
        comparison_langs = []
        for lang in selected_languages:
            if lang in source_lang_map:
                comparison_langs.append(lang)
                logger.info(f"将对比语言: {lang} (源文件第{source_lang_map[lang]}列)")
                if lang in trans_lang_map:
                    logger.info(f"  - 翻译文件对应列: 第{trans_lang_map[lang]}列")
                else:
                    logger.warning(f"  - 警告: 翻译文件中未找到此语言列")
            else:
                logger.warning(f"源文件中未找到语言列: {lang}")

        # 如果某些选中的语言在源文件中没有找到，尝试使用列顺序映射
        missing_langs = [lang for lang in selected_languages if lang not in comparison_langs]
        if missing_langs:
            logger.warning(f"以下语言未找到匹配列: {missing_langs}，尝试使用列顺序映射")

            # 列顺序映射表
            col_order = [
                "中文（CN）",
                "英文（EN）English",
                "德语(DE)Deutsch",
                "西语（ES）Español",
                "法语(FR)Français",
                "意大利语(IT)Italiano",
                "巴西葡语(BR)Português",
                "俄语（Pyc）Русский",
                "土耳其语(TR)Turkish",
                "日语(JP)日本語",
                "韩语(KR)한국어",
                "阿拉伯语عربية",
                "繁体中文",
                "波兰语（PL）Polski",
                "越南语（VI）Tiếng Việt",
                "印尼语（ID）Bahasa Indonesia",
                "泰语（TH）ไทย",
                "马来语（MS）Bahasa Melayu",
                "希伯来语（HE）עברית",
                "南非语（AF）Afrikaans"
            ]

            # 根据列顺序为缺失的语言分配列号
            next_col = 2  # 从第2列开始（第1列是键名）
            for lang in col_order:
                if lang in missing_langs and next_col <= source_ws.max_column:
                    source_lang_map[lang] = next_col
                    comparison_langs.append(lang)
                    logger.info(f"默认将第{next_col}列映射为 {lang}")
                next_col += 1

        # 特别检查：确保繁体中文没有被错误映射到中文
        if "繁体中文" in selected_languages and "繁体中文" in source_lang_map:
            # 确认繁体中文的列不是中文的列
            if "中文（CN）" in source_lang_map and source_lang_map["中文（CN）"] == source_lang_map["繁体中文"]:
                logger.warning("繁体中文和中文映射到了同一列，尝试重新映射")
                # 尝试找到真正的繁体中文列
                for col in range(source_ws.max_column, 1, -1):  # 从最后一列往前找
                    header = source_ws.cell(row=1, column=col).value
                    if header and isinstance(header, str):
                        if "繁体" in header or "traditional" in header.lower():
                            source_lang_map["繁体中文"] = col
                            logger.info(f"重新映射繁体中文到第{col}列")
                            break

        tasks[task_id]['progress'] = 50
        tasks[task_id]['message'] = f'开始对比 {len(source_key_map)} 个键，{len(comparison_langs)} 种语言...'

        # 打印最终的语言映射，用于调试
        logger.info("=" * 50)
        logger.info("最终语言映射:")
        for lang in comparison_langs:
            source_col = source_lang_map.get(lang, "None")
            trans_col = trans_lang_map.get(lang, "None")
            logger.info(f"{lang}: 源文件第{source_col}列, 翻译文件第{trans_col}列")
        logger.info("=" * 50)

        # 执行对比
        output_row = 2
        matched_count = 0
        total_keys = len(source_key_map)
        total_comparisons = 0
        key_match_count = 0

        # 首先找出所有匹配的键名（忽略大小写和空格）
        source_keys_normalized = {k.strip().lower(): k for k in source_key_map.keys()}
        trans_keys_normalized = {k.strip().lower(): k for k in trans_key_map.keys()}

        # 创建翻译键名的反向映射（原始键名 -> 行号）
        trans_key_row_map = {k: v for k, v in trans_key_map.items()}

        # 找出匹配的键名
        matching_keys = set(source_keys_normalized.keys()) & set(trans_keys_normalized.keys())
        key_match_count = len(matching_keys)

        # 创建源键名到翻译键名的映射
        source_to_trans_key = {}
        for norm_key in matching_keys:
            source_original = source_keys_normalized[norm_key]
            trans_original = trans_keys_normalized[norm_key]
            source_to_trans_key[source_original] = trans_original

        logger.info(f"键名匹配数量: {key_match_count}/{total_keys}")
        logger.info(f"匹配的键名示例: {list(source_to_trans_key.items())[:5]}")

        # 存储每个键的起始行和结束行，用于后续合并
        key_ranges = []

        # 第一次遍历：收集每个键的数据行范围
        current_key = None
        key_start_row = output_row
        key_lang_count = 0
        current_trans_key = None

        # 按源文件键名顺序处理
        source_keys_list = list(source_key_map.keys())

        for i, source_key in enumerate(source_keys_list):
            source_row = source_key_map[source_key]

            # 更新进度
            progress = 50 + (i / total_keys) * 40
            if i % 5 == 0:
                tasks[task_id]['progress'] = int(progress)
                tasks[task_id]['message'] = f'正在对比 {i + 1}/{total_keys}...'

            # 查找匹配的翻译键
            trans_key = source_to_trans_key.get(source_key)
            trans_row = trans_key_row_map.get(trans_key) if trans_key else None

            # 记录这个键的起始行
            if current_key != source_key:
                if current_key is not None and key_lang_count > 0:
                    # 保存上一个键的范围
                    key_ranges.append({
                        'source_key': current_key,
                        'trans_key': current_trans_key,
                        'start_row': key_start_row,
                        'end_row': output_row - 1,
                        'lang_count': key_lang_count
                    })
                # 开始新键
                current_key = source_key
                current_trans_key = trans_key if trans_key else "未匹配"
                key_start_row = output_row
                key_lang_count = 0

            # 对每个选中的语言进行对比
            for lang in comparison_langs:
                source_col = source_lang_map.get(lang)

                if not source_col:
                    continue

                # 获取源文件内容
                source_content = ""
                try:
                    source_cell = source_ws.cell(row=source_row, column=source_col)
                    source_content = str(source_cell.value) if source_cell.value is not None else ""
                except Exception as e:
                    source_content = f"【读取错误】"
                    logger.warning(f"读取源文件内容错误: {e}")

                # 获取翻译文件内容
                trans_content = ""
                trans_col = trans_lang_map.get(lang) if lang in trans_lang_map else None

                if trans_row and trans_col:
                    try:
                        trans_cell = trans_ws.cell(row=trans_row, column=trans_col)
                        trans_content = str(trans_cell.value) if trans_cell.value is not None else ""
                    except Exception as e:
                        trans_content = f"【读取错误】"
                        logger.warning(f"读取翻译文件内容错误: {e}")
                elif not trans_row:
                    trans_content = "【键名未匹配】"
                elif not trans_col:
                    trans_content = f"【翻译文件缺少{lang}列】"

                # 先不写入源文件键名和翻译文件键名，留空
                output_ws.cell(row=output_row, column=1).value = ""  # 源文件键名留空，后续合并
                output_ws.cell(row=output_row, column=2).value = ""  # 翻译文件键名留空，后续合并
                output_ws.cell(row=output_row, column=3).value = lang
                output_ws.cell(row=output_row, column=4).value = source_content
                output_ws.cell(row=output_row, column=5).value = trans_content

                # 判断对比结果并设置样式
                if not trans_row:
                    result = "键名缺失"
                    output_ws.cell(row=output_row, column=6).fill = YELLOW_FILL
                elif trans_content.startswith("【翻译文件缺少"):
                    result = "语言列缺失"
                    output_ws.cell(row=output_row, column=6).fill = YELLOW_FILL
                elif source_content and trans_content and source_content == trans_content:
                    result = "一致"
                    output_ws.cell(row=output_row, column=4).fill = GREEN_FILL
                    output_ws.cell(row=output_row, column=5).fill = GREEN_FILL
                    matched_count += 1
                elif not source_content and not trans_content:
                    result = "均为空"
                    output_ws.cell(row=output_row, column=6).fill = GREEN_FILL
                else:
                    result = "不一致"
                    output_ws.cell(row=output_row, column=4).fill = RED_FILL
                    output_ws.cell(row=output_row, column=5).fill = RED_FILL

                output_ws.cell(row=output_row, column=6).value = result
                output_row += 1
                key_lang_count += 1
                total_comparisons += 1

        # 保存最后一个键的范围
        if current_key is not None and key_lang_count > 0:
            key_ranges.append({
                'source_key': current_key,
                'trans_key': current_trans_key,
                'start_row': key_start_row,
                'end_row': output_row - 1,
                'lang_count': key_lang_count
            })

        logger.info(f"收集到 {len(key_ranges)} 个键的数据范围")

        # 第二次遍历：合并单元格并写入键名
        for key_info in key_ranges:
            start_row = key_info['start_row']
            end_row = key_info['end_row']

            if start_row <= end_row:
                try:
                    # 合并源文件键名单元格
                    if end_row > start_row:
                        output_ws.merge_cells(start_row=start_row, start_column=1,
                                              end_row=end_row, end_column=1)

                    # 写入源文件键名
                    source_cell = output_ws.cell(row=start_row, column=1)
                    source_cell.value = key_info['source_key']
                    source_cell.font = BOLD_FONT
                    source_cell.alignment = Alignment(vertical='center')

                    # 合并翻译文件键名单元格
                    if end_row > start_row:
                        output_ws.merge_cells(start_row=start_row, start_column=2,
                                              end_row=end_row, end_column=2)

                    # 写入翻译文件键名
                    trans_cell = output_ws.cell(row=start_row, column=2)
                    trans_cell.value = key_info['trans_key']
                    trans_cell.font = BOLD_FONT
                    trans_cell.alignment = Alignment(vertical='center')

                    logger.info(
                        f"写入键名: 源='{key_info['source_key']}', 翻译='{key_info['trans_key']}', 行范围 {start_row}-{end_row}")

                except Exception as e:
                    logger.warning(f"合并单元格失败: {e}")

        tasks[task_id]['progress'] = 90
        tasks[task_id]['message'] = '正在生成总结报告...'

        # 添加总结
        summary_row = output_row + 2
        output_ws.cell(row=summary_row, column=1).value = "对比总结"
        output_ws.cell(row=summary_row, column=1).font = HEADER_FONT

        # 统计各语言的数据量
        lang_stats = {}
        for lang in comparison_langs:
            if lang in trans_lang_map:
                lang_stats[lang] = "有对应列"
            else:
                lang_stats[lang] = "无对应列"

        summary_data = [
            f"源文件总键数: {total_keys}",
            f"翻译文件总键数: {len(trans_key_map)}",
            f"键名匹配数: {key_match_count}",
            f"内容一致数: {matched_count}",
            f"对比语言数: {len(comparison_langs)}",
            f"总对比次数: {total_comparisons}",
            f"语言列情况: {lang_stats}"
        ]

        for i, text in enumerate(summary_data):
            output_ws.cell(row=summary_row + 1 + i, column=1).value = text

        # 先关闭所有工作簿
        if source_wb:
            source_wb.close()
        if trans_wb:
            trans_wb.close()

        tasks[task_id]['progress'] = 95
        tasks[task_id]['message'] = '正在保存结果文件...'

        # 保存结果
        output_filename = f"comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)

        # 确保目录存在
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

        # 保存并关闭输出文件
        output_wb.save(output_path)
        output_wb.close()

        # 验证文件是否成功保存
        if not os.path.exists(output_path):
            raise Exception("文件保存失败")

        tasks[task_id]['progress'] = 100
        tasks[task_id]['message'] = '对比完成'

        # 清理旧文件
        cleanup_old_files()

        logger.info("=" * 50)
        logger.info(f"对比完成，结果保存到: {output_path}")
        logger.info(f"统计: 总键数={total_keys}, 键名匹配={key_match_count}, 内容一致={matched_count}")
        logger.info("=" * 50)

        return {
            'success': True,
            'filename': output_filename
        }

    except Exception as e:
        logger.error(f"对比过程出错: {e}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e)
        }
    finally:
        # 确保所有工作簿都被关闭
        try:
            if source_wb:
                source_wb.close()
            if trans_wb:
                trans_wb.close()
            if output_wb:
                output_wb.close()
        except:
            pass


@app.route('/api/task-status/<task_id>')
def task_status(task_id):
    """获取任务状态"""
    if task_id in tasks:
        return jsonify(tasks[task_id])
    return jsonify({'status': 'not_found', 'message': '任务不存在'})


@app.route('/download/<filename>')
def download_file(filename):
    """下载文件"""
    try:
        # 安全检查，防止路径遍历
        if '..' in filename or filename.startswith('/') or '\\' in filename:
            return jsonify({'error': '无效的文件名'}), 400

        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        if not os.path.exists(file_path):
            return jsonify({'error': '文件不存在或已过期'}), 404

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"下载文件失败: {e}")
        return jsonify({'error': '下载失败'}), 500


# 添加错误码校对辅助类
class ErrorCodeMatcher:
    """错误码匹配器 - 优化版"""

    def __init__(self, threshold=90, ignore_case=True, ignore_special_chars=True, extract_numbers=True):
        self.threshold = threshold / 100.0  # 转换为0-1之间的小数
        self.ignore_case = ignore_case
        self.ignore_special_chars = ignore_special_chars
        self.extract_numbers = extract_numbers

    def normalize_error_code(self, code):
        """标准化错误码"""
        if not code or code == "":
            return ""

        code_str = str(code)

        if self.ignore_case:
            code_str = code_str.lower()

        if self.ignore_special_chars:
            # 移除非字母数字字符，但保留空格
            code_str = re.sub(r'[^\w\s]', '', code_str)

        return code_str.strip()

    def extract_numbers_only(self, code):
        """只提取数字部分"""
        if not code:
            return []

        code_str = str(code)
        # 提取所有连续的数字
        numbers = re.findall(r'\d+', code_str)
        # 转换为整数列表
        return [int(num) for num in numbers]

    def extract_prefix(self, code):
        """提取错误码的前缀（字母部分）"""
        if not code:
            return ""

        code_str = str(code)
        # 提取开头的字母部分（直到遇到第一个数字）
        match = re.match(r'^([A-Za-z]+)', code_str)
        if match:
            return match.group(1).lower() if self.ignore_case else match.group(1)
        return ""

    def calculate_similarity(self, code1, code2):
        """计算两个错误码的相似度 - 优化版"""
        if not code1 or not code2:
            return 0

        str1 = str(code1)
        str2 = str(code2)

        # 1. 首先检查是否完全相等
        if str1 == str2:
            return 100
        if self.ignore_case and str1.lower() == str2.lower():
            return 100

        # 2. 提取数字部分
        nums1 = self.extract_numbers_only(str1)
        nums2 = self.extract_numbers_only(str2)

        # 3. 提取前缀（字母部分）
        prefix1 = self.extract_prefix(str1)
        prefix2 = self.extract_prefix(str2)

        # 4. 数字必须存在且匹配的检查
        if nums1 and nums2:
            # 数字必须完全相等
            if nums1 != nums2:
                # 检查是否是包含关系（如 2000 和 2000A）
                if len(nums1) == 1 and len(nums2) == 1:
                    num1_str = str(nums1[0])
                    num2_str = str(nums2[0])
                    # 如果一个数字是另一个数字的开头部分
                    if num2_str.startswith(num1_str) or num1_str.startswith(num2_str):
                        # 检查前缀是否匹配
                        if prefix1 and prefix2 and prefix1 == prefix2:
                            # 相同前缀，数字包含关系，给70%基础分
                            base_score = 70
                        elif prefix1 and prefix2:
                            # 不同前缀，给40%基础分
                            base_score = 40
                        else:
                            # 无前缀或单边有前缀，给50%基础分
                            base_score = 50

                        # 比较剩余部分
                        text1 = re.sub(r'\d+', '', str1).strip()
                        text2 = re.sub(r'\d+', '', str2).strip()
                        if text1 and text2:
                            text_similarity = SequenceMatcher(None, text1, text2).ratio() * 100
                            return min(100, base_score + text_similarity * 0.3)
                        return base_score
                # 数字不匹配，直接返回0
                return 0

        # 5. 如果只有一边有数字
        elif nums1 and not nums2:
            # 有数字的一边和无数字的一边，返回0
            return 0
        elif not nums1 and nums2:
            return 0

        # 6. 都没有数字，使用文本相似度
        norm1 = self.normalize_error_code(str1)
        norm2 = self.normalize_error_code(str2)

        if not norm1 or not norm2:
            return 0

        similarity = SequenceMatcher(None, norm1, norm2).ratio() * 100
        return round(similarity, 2)

    def find_best_match(self, source_code, trans_codes_dict):
        """在翻译错误码中寻找最佳匹配 - 优化版"""
        best_match = None
        best_similarity = 0
        best_trans_key = None

        str_source = str(source_code)
        nums_source = self.extract_numbers_only(str_source)
        prefix_source = self.extract_prefix(str_source)

        # 如果有数字，先过滤出数字匹配的候选
        candidates = []
        if nums_source:
            for trans_key, trans_code in trans_codes_dict.items():
                str_trans = str(trans_code)
                nums_trans = self.extract_numbers_only(str_trans)

                if nums_trans:
                    # 数字必须完全相等
                    if nums_source == nums_trans:
                        candidates.append((trans_key, trans_code, 100))
                    else:
                        # 检查数字包含关系
                        if len(nums_source) == 1 and len(nums_trans) == 1:
                            num1_str = str(nums_source[0])
                            num2_str = str(nums_trans[0])
                            if num2_str.startswith(num1_str) or num1_str.startswith(num2_str):
                                # 检查前缀
                                prefix_trans = self.extract_prefix(str_trans)
                                if prefix_source and prefix_trans and prefix_source == prefix_trans:
                                    candidates.append((trans_key, trans_code, 70))
                                else:
                                    candidates.append((trans_key, trans_code, 40))
        else:
            # 没有数字，所有都是候选
            candidates = [(key, code, 0) for key, code in trans_codes_dict.items()]

        # 如果没有候选，返回空
        if not candidates:
            return None, None, 0

        # 在候选中计算相似度
        for trans_key, trans_code, base_score in candidates:
            similarity = self.calculate_similarity(source_code, trans_code)

            # 如果候选有基础分，确保相似度不低于基础分
            if base_score > 0:
                similarity = max(similarity, base_score)

            if similarity > best_similarity:
                best_similarity = similarity
                best_match = trans_code
                best_trans_key = trans_key

        # 如果最佳匹配超过阈值，返回匹配结果
        if best_similarity >= (self.threshold * 100):
            return best_trans_key, best_match, best_similarity
        else:
            return None, None, best_similarity


def extract_error_codes_from_excel(file_path, selected_languages):
    """从Excel文件中提取错误码和对应的翻译"""
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # 构建语言列映射
        lang_col_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and isinstance(header, str):
                header_str = header.strip()
                for lang in selected_languages:
                    if lang in header_str:
                        lang_col_map[lang] = col
                        break

        # 提取数据
        error_codes = []
        for row in range(2, ws.max_row + 1):
            # 获取键名（第一列）
            key_cell = ws.cell(row=row, column=1)
            if not key_cell.value:
                continue

            key = str(key_cell.value).strip()
            if not key:
                continue

            # 获取各语言的错误码
            translations = {}
            for lang, col in lang_col_map.items():
                cell = ws.cell(row=row, column=col)
                translations[lang] = str(cell.value) if cell.value else ""

            error_codes.append({
                'key': key,
                'translations': translations
            })

        return error_codes, lang_col_map

    finally:
        if wb:
            wb.close()


@app.route('/api/error-code-check', methods=['POST'])
def error_code_check():
    """功能三：错误码校对 - 完整实现"""
    try:
        if 'source_file' not in request.files or 'trans_file' not in request.files:
            return jsonify({'error': '请上传两个Excel文件'}), 400

        source_file = request.files['source_file']
        trans_file = request.files['trans_file']

        if source_file.filename == '' or trans_file.filename == '':
            return jsonify({'error': '请选择有效的文件'}), 400

        # 获取参数
        selected_languages = request.form.getlist('languages[]')
        if not selected_languages:
            selected_languages = LANGUAGE_ORDER

        threshold = int(request.form.get('threshold', 80))
        ignore_case = request.form.get('ignore_case', 'true').lower() == 'true'
        ignore_special_chars = request.form.get('ignore_special_chars', 'true').lower() == 'true'
        extract_numbers = request.form.get('extract_numbers', 'true').lower() == 'true'

        # 保存上传的文件
        source_filename = secure_filename(source_file.filename)
        trans_filename = secure_filename(trans_file.filename)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        source_path = os.path.join(app.config['UPLOAD_FOLDER'], f"error_source_{uuid.uuid4().hex}_{source_filename}")
        trans_path = os.path.join(app.config['UPLOAD_FOLDER'], f"error_trans_{uuid.uuid4().hex}_{trans_filename}")

        source_file.save(source_path)
        trans_file.save(trans_path)

        # 创建任务
        task_id = uuid.uuid4().hex
        tasks[task_id] = {
            'progress': 0,
            'status': 'processing',
            'message': '错误码校对任务已创建'
        }

        # 启动后台线程处理
        def process_error_task():
            try:
                tasks[task_id]['progress'] = 10
                tasks[task_id]['message'] = '正在加载源文件...'

                # 加载源文件
                wb_source = openpyxl.load_workbook(source_path, data_only=True)
                ws_source = wb_source.active

                # 加载翻译文件
                wb_trans = openpyxl.load_workbook(trans_path, data_only=True)
                ws_trans = wb_trans.active

                tasks[task_id]['progress'] = 20
                tasks[task_id]['message'] = '正在识别语言列...'

                # 构建语言列映射
                def build_lang_column_map(ws):
                    lang_map = {}
                    for col in range(1, ws.max_column + 1):
                        header = ws.cell(row=1, column=col).value
                        if header and isinstance(header, str):
                            header_str = header.strip()
                            for lang in selected_languages:
                                if lang in header_str:
                                    lang_map[lang] = col
                                    break
                    return lang_map

                source_lang_map = build_lang_column_map(ws_source)
                trans_lang_map = build_lang_column_map(ws_trans)

                # 找到中文列（用于键名匹配）
                source_chinese_col = None
                trans_chinese_col = None

                for lang in selected_languages:
                    if '中文' in lang:
                        source_chinese_col = source_lang_map.get(lang)
                        trans_chinese_col = trans_lang_map.get(lang)
                        break

                if not source_chinese_col or not trans_chinese_col:
                    raise Exception("无法找到中文列")

                tasks[task_id]['progress'] = 30
                tasks[task_id]['message'] = '正在构建源文件数据...'

                # 读取源文件数据
                source_data = []
                for row in range(2, ws_source.max_row + 1):
                    key_cell = ws_source.cell(row=row, column=1)
                    if not key_cell.value:
                        continue

                    source_key = str(key_cell.value).strip()
                    chinese_cell = ws_source.cell(row=row, column=source_chinese_col)
                    chinese_value = str(chinese_cell.value) if chinese_cell.value else ""

                    # 提取数字用于匹配
                    numbers = re.findall(r'\d+', chinese_value)

                    source_data.append({
                        'key': source_key,
                        'chinese': chinese_value,
                        'numbers': numbers,
                        'row': row,
                        'translations': {}
                    })

                    # 读取所有语言的内容
                    for lang, col in source_lang_map.items():
                        cell = ws_source.cell(row=row, column=col)
                        source_data[-1]['translations'][lang] = str(cell.value) if cell.value else ""

                tasks[task_id]['progress'] = 40
                tasks[task_id]['message'] = '正在构建翻译文件数据...'

                # 读取翻译文件数据
                trans_data = []
                trans_by_key = {}
                trans_by_chinese = {}  # 按中文内容索引

                for row in range(2, ws_trans.max_row + 1):
                    key_cell = ws_trans.cell(row=row, column=1)
                    if not key_cell.value:
                        continue

                    trans_key = str(key_cell.value).strip()
                    chinese_cell = ws_trans.cell(row=row, column=trans_chinese_col)
                    chinese_value = str(chinese_cell.value) if chinese_cell.value else ""

                    # 提取数字
                    numbers = re.findall(r'\d+', chinese_value)

                    item = {
                        'key': trans_key,
                        'chinese': chinese_value,
                        'numbers': numbers,
                        'row': row,
                        'translations': {}
                    }

                    # 读取所有语言的内容
                    for lang, col in trans_lang_map.items():
                        cell = ws_trans.cell(row=row, column=col)
                        item['translations'][lang] = str(cell.value) if cell.value else ""

                    trans_data.append(item)
                    trans_by_key[trans_key] = item

                    # 按中文内容建立索引
                    if chinese_value:
                        # 按完整内容索引
                        if chinese_value not in trans_by_chinese:
                            trans_by_chinese[chinese_value] = []
                        trans_by_chinese[chinese_value].append(item)

                tasks[task_id]['progress'] = 50
                tasks[task_id]['message'] = '正在创建匹配器...'

                # 创建匹配器
                matcher = ErrorCodeMatcher(
                    threshold=threshold,
                    ignore_case=ignore_case,
                    ignore_special_chars=ignore_special_chars,
                    extract_numbers=extract_numbers
                )

                tasks[task_id]['progress'] = 60
                tasks[task_id]['message'] = '正在进行键名+中文匹配...'

                # 第一步：按键名+中文进行匹配
                matches = []
                key_matches = {}  # 记录每个源键名匹配到的翻译键名
                current_row = 2

                for source_item in source_data:
                    source_key = source_item['key']
                    source_chinese = source_item['chinese']

                    # 记录起始行
                    if source_key not in key_matches:
                        key_matches[source_key] = {
                            'matched_key': '未匹配',
                            'match_type': '无匹配',
                            'similarity': 0,
                            'start_row': current_row,
                            'end_row': current_row + len(selected_languages) - 1
                        }

                    # 尝试匹配
                    best_match = None
                    best_similarity = 0

                    # 1. 先尝试精确匹配键名
                    if source_key in trans_by_key:
                        trans_item = trans_by_key[source_key]
                        trans_chinese = trans_item['chinese']

                        if source_chinese and trans_chinese:
                            similarity = matcher.calculate_similarity(source_chinese, trans_chinese)
                            if similarity >= threshold:
                                best_match = trans_item
                                best_similarity = similarity
                                match_type = '键名+中文精确匹配'

                    # 2. 如果键名不匹配，尝试通过中文内容匹配
                    if not best_match and source_chinese:
                        # 先找相同数字的
                        if source_item['numbers']:
                            for num in source_item['numbers']:
                                for trans_item in trans_data:
                                    if num in trans_item['numbers']:
                                        similarity = matcher.calculate_similarity(source_chinese, trans_item['chinese'])
                                        if similarity > best_similarity and similarity >= threshold:
                                            best_similarity = similarity
                                            best_match = trans_item
                                            match_type = '数字匹配'

                        # 如果没找到，找相似内容的
                        if not best_match:
                            for trans_item in trans_data:
                                if trans_item['chinese']:
                                    similarity = matcher.calculate_similarity(source_chinese, trans_item['chinese'])
                                    if similarity > best_similarity and similarity >= threshold:
                                        best_similarity = similarity
                                        best_match = trans_item
                                        match_type = '内容模糊匹配'

                    # 更新匹配结果
                    if best_match:
                        key_matches[source_key]['matched_key'] = best_match['key']
                        key_matches[source_key]['match_type'] = match_type
                        key_matches[source_key]['similarity'] = best_similarity
                        key_matches[source_key]['trans_item'] = best_match

                tasks[task_id]['progress'] = 70
                tasks[task_id]['message'] = '正在对比各语言内容...'

                # 第二步：根据匹配到的键名，对比所有语言
                exact_matches = 0
                content_matches = 0
                unmatched = 0

                for source_item in source_data:
                    source_key = source_item['key']
                    match_info = key_matches[source_key]
                    matched_key = match_info['matched_key']
                    trans_item = match_info.get('trans_item')

                    for lang in selected_languages:
                        source_value = source_item['translations'].get(lang, '')
                        trans_value = ''

                        if trans_item and lang in trans_item['translations']:
                            trans_value = trans_item['translations'][lang]

                        # 判断匹配类型
                        if matched_key != '未匹配' and trans_item:
                            if source_value and trans_value and source_value == trans_value:
                                match_type = f'内容一致'
                                exact_matches += 1
                            elif source_value and trans_value:
                                match_type = f'内容不一致'
                                content_matches += 1
                            elif source_value and not trans_value:
                                match_type = f'翻译缺失'
                                unmatched += 1
                            elif not source_value and trans_value:
                                match_type = f'源文件缺失'
                                unmatched += 1
                            else:
                                match_type = f'均为空'
                                unmatched += 1
                        else:
                            match_type = f'键名未匹配'
                            unmatched += 1

                        matches.append({
                            'source_key': source_key,
                            'matched_key': matched_key,
                            'lang': lang,
                            'source_value': source_value,
                            'trans_value': trans_value,
                            'match_type': match_type,
                            'row': current_row
                        })
                        current_row += 1

                tasks[task_id]['progress'] = 85
                tasks[task_id]['message'] = '正在生成结果文件...'

                # 创建结果Excel文件
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "错误码校对结果"

                # 设置表头
                headers = ["源文件键名", "匹配键名", "语言", "源文件错误码", "翻译错误码", "匹配类型"]
                ws.append(headers)

                # 设置表头样式
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                # 设置列宽
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 30
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 20

                # 定义颜色样式
                green_fill = PatternFill(start_color='8FBC8F', end_color='8FBC8F', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFF68F', end_color='FFF68F', fill_type='solid')
                blue_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
                red_fill = PatternFill(start_color='CD0000', end_color='CD0000', fill_type='solid')

                # 写入数据
                for match in matches:
                    row = [
                        match['source_key'],
                        match['matched_key'],
                        match['lang'],
                        match['source_value'],
                        match['trans_value'],
                        match['match_type']
                    ]
                    ws.append(row)

                    # 设置行样式
                    current_row = ws.max_row
                    match_type = match['match_type']

                    if '内容一致' in match_type:
                        for col in range(4, 6):
                            ws.cell(row=current_row, column=col).fill = green_fill
                    elif '内容不一致' in match_type:
                        for col in range(4, 6):
                            ws.cell(row=current_row, column=col).fill = yellow_fill
                    elif '缺失' in match_type:
                        for col in range(4, 6):
                            ws.cell(row=current_row, column=col).fill = red_fill

                        # 合并源文件键名和匹配键名的单元格
                        current_source_key = None
                        start_row = 2
                        end_row = 2

                        # 先对matches按source_key排序，确保同一个键的行是连续的
                        for i, match in enumerate(matches):
                            if current_source_key is None:
                                current_source_key = match['source_key']
                                start_row = match['row']

                            # 如果是最后一个元素或者下一个键不同，执行合并
                            if i == len(matches) - 1 or matches[i + 1]['source_key'] != current_source_key:
                                end_row = match['row']

                                if end_row > start_row:
                                    # 合并源文件键名列
                                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                                    cell = ws.cell(row=start_row, column=1)
                                    cell.font = Font(bold=True)
                                    cell.alignment = Alignment(vertical='center', horizontal='left')

                                    # 合并匹配键名列
                                    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                                    cell = ws.cell(row=start_row, column=2)
                                    cell.font = Font(bold=True)

                                    # 在匹配键名后添加匹配信息
                                    match_info = key_matches.get(current_source_key, {})
                                    if match_info.get('matched_key') != '未匹配':
                                        cell.font = Font(bold=True, color='006400')
                                        # 在H列添加详细匹配信息
                                        info_cell = ws.cell(row=start_row, column=8)
                                        info_cell.value = f"匹配方式: {match_info.get('match_type', '未知')}, 相似度: {match_info.get('similarity', 0)}%"
                                        info_cell.font = Font(italic=True, size=9)
                                    else:
                                        cell.font = Font(bold=True, color='FF0000')

                                # 重置，准备处理下一个键
                                if i < len(matches) - 1:
                                    current_source_key = matches[i + 1]['source_key']
                                    start_row = matches[i + 1]['row']


                        # 根据匹配类型设置颜色
                        if match_info['matched_key'] != '未匹配':
                            cell.font = Font(bold=True, color='006400')
                            # 在合并单元格的旁边添加匹配信息
                            info_cell = ws.cell(row=start_row, column=7)
                            info_cell.value = f"匹配方式: {match_info['match_type']}, 相似度: {match_info['similarity']}%"
                            info_cell.font = Font(italic=True, size=10)

                # 添加统计信息
                summary_row = ws.max_row + 3
                ws.cell(row=summary_row, column=1).value = "统计信息"
                ws.cell(row=summary_row, column=1).font = Font(bold=True, size=14)

                total_items = len(source_data) * len(selected_languages)
                matched_keys = sum(1 for info in key_matches.values() if info['matched_key'] != '未匹配')

                stats = [
                    f"总键数: {len(source_data)}",
                    f"匹配到的键数: {matched_keys}",
                    f"键匹配率: {(matched_keys / len(source_data) * 100):.1f}%",
                    f"总语言项数: {total_items}",
                    f"内容一致项: {exact_matches}",
                    f"内容不一致项: {content_matches}",
                    f"未匹配项: {unmatched}",
                    f"内容匹配率: {((exact_matches + content_matches) / total_items * 100):.1f}%",
                    f"匹配阈值: {threshold}%"
                ]

                for i, stat in enumerate(stats):
                    ws.cell(row=summary_row + 1 + i, column=1).value = stat

                # 保存文件
                output_filename = f"error_check_{timestamp}.xlsx"
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
                wb.save(output_path)

                tasks[task_id].update({
                    'progress': 100,
                    'status': 'completed',
                    'message': f'校对完成！总键数: {len(source_data)}, 匹配键数: {matched_keys}, 内容一致: {exact_matches}',
                    'filename': output_filename,
                    'download_url': f'/download/{output_filename}',
                    'result': {
                        'statistics': {
                            'total_keys': len(source_data),
                            'matched_keys': matched_keys,
                            'key_match_rate': round(matched_keys / len(source_data) * 100, 1),
                            'total_items': total_items,
                            'exact_matches': exact_matches,
                            'content_matches': content_matches,
                            'unmatched': unmatched,
                            'item_match_rate': round((exact_matches + content_matches) / total_items * 100, 1)
                        }
                    }
                })

            except Exception as e:
                logger.error(f"错误码校对失败: {e}")
                import traceback
                traceback.print_exc()
                tasks[task_id].update({
                    'progress': 0,
                    'status': 'error',
                    'message': str(e)
                })
            finally:
                # 清理临时文件
                try:
                    wb_source.close()
                    wb_trans.close()
                    if os.path.exists(source_path):
                        os.remove(source_path)
                    if os.path.exists(trans_path):
                        os.remove(trans_path)
                except:
                    pass

        thread = threading.Thread(target=process_error_task)
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'task_id': task_id,
            'message': '错误码校对任务已启动'
        })

    except Exception as e:
        logger.error(f"启动错误码校对任务失败: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    local_ip = get_local_ip()
    print(f"启动服务器...")
    print(f"访问地址: http://localhost:5000")
    print(f"局域网访问: http://{local_ip}:5000")
    print(f"临时文件目录: {app.config['UPLOAD_FOLDER']}")
    print(f"支持语言: {', '.join(LANGUAGE_ORDER[:5])}...")
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)