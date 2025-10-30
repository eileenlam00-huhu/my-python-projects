import openpyxl
from openpyxl.styles import PatternFill, Font
import tkinter as tk
from tkinter import filedialog, messagebox


def compare_firmware_translations(source_file, translation_file, output_file):
    """
    对比源文件和翻译文件，标记差异

    参数:
        source_file: 源文件路径
        translation_file: 翻译文件路径
        output_file: 输出文件路径
    """
    # 定义颜色
    RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    GREEN_FILL = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

    try:
        # 加载工作簿
        source_wb = openpyxl.load_workbook(source_file)
        translation_wb = openpyxl.load_workbook(translation_file)

        # 创建输出工作簿
        output_wb = openpyxl.Workbook()

        # 假设我们比较第一个工作表
        source_ws = source_wb.active
        translation_ws = translation_wb.active
        output_ws = output_wb.active
        output_ws.title = "Comparison Result"

        # 复制表头
        for col in range(1, source_ws.max_column + 1):
            output_ws.cell(row=1, column=col).value = source_ws.cell(row=1, column=col).value

        # 比较内容
        differences_found = False
        empty_cells_found = False

        for row in range(2, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                source_value = source_ws.cell(row=row, column=col).value
                trans_value = translation_ws.cell(row=row, column=col).value

                # 检查是否为空值
                if trans_value is None or str(trans_value).strip() == "":
                    output_ws.cell(row=row, column=col).value = "(空值)"
                    output_ws.cell(row=row, column=col).fill = RED_FILL
                    empty_cells_found = True
                    differences_found = True
                    continue

                # 写入输出文件
                output_ws.cell(row=row, column=col).value = trans_value

                # 检查源文件是否为空
                if source_value is None or str(source_value).strip() == "":
                    output_ws.cell(row=row, column=col).fill = RED_FILL
                    differences_found = True
                    continue

                # 标记差异
                if str(source_value) != str(trans_value):
                    output_ws.cell(row=row, column=col).fill = RED_FILL
                    differences_found = True
                else:
                    output_ws.cell(row=row, column=col).fill = GREEN_FILL

        # 保存输出文件
        output_wb.save(output_file)

        # 显示结果消息
        result_msg = "比较完成！\n"
        if differences_found:
            result_msg += "发现差异内容已用红色标记。\n"
            if empty_cells_found:
                result_msg += "注意：发现空值单元格已标记为红色。\n"
        else:
            result_msg += "所有翻译内容与源文件一致。\n"

        result_msg += f"结果已保存到: {output_file}"
        messagebox.showinfo("完成", result_msg)

    except Exception as e:
        messagebox.showerror("错误", f"发生错误: {str(e)}")


def select_file(title):
    """打开文件选择对话框"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    return file_path


def select_output_file():
    """选择输出文件位置"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    file_path = filedialog.asksaveasfilename(
        title="选择输出文件位置",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    return file_path


def main():
    # 创建GUI界面
    root = tk.Tk()
    root.title("固件版本多语言差异对比工具")
    root.geometry("400x300")

    # 添加说明标签
    tk.Label(root, text="固件版本多语言差异对比工具", font=('Arial', 14)).pack(pady=10)
    tk.Label(root, text="空值单元格也会被标记为红色", fg="red").pack()

    # 添加按钮
    def on_compare():
        source_file = select_file("选择源文件")
        if not source_file:
            return

        translation_file = select_file("选择翻译文件")
        if not translation_file:
            return

        output_file = select_output_file()
        if not output_file:
            return

        compare_firmware_translations(source_file, translation_file, output_file)

    compare_btn = tk.Button(root, text="开始对比", command=on_compare, height=2, width=20)
    compare_btn.pack(pady=20)

    # 退出按钮
    exit_btn = tk.Button(root, text="退出", command=root.quit, height=2, width=20)
    exit_btn.pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    main()