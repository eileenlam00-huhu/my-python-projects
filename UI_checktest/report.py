import os
from openpyxl import Workbook
from openpyxl.styles import Font
from .utils import ensure_dir_exists

try:
    from openpyxl.drawing.image import Image as OpenpyxlImage
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False
    print("警告: Pillow 未安装，图片嵌入功能将被禁用")


def create_report_workbook(excel_path):
    ensure_dir_exists(os.path.dirname(excel_path))
    wb = Workbook()
    ws = wb.active
    ws.title = 'UI对比报告'
    
    # 设置表头
    headers = [
        '语言ID',
        '语言名称',
        '页面类型',
        '一级页面',
        '子页面',
        '基准图',
        '测试图',
        'diff',
        '状态'
    ]
    ws.append(headers)

    # --- 修改点 1: 初始化时设置 F 列和 G 列的宽度 ---
    # F列(基准图)和G列(测试图)宽度设为 45，防止图片重叠
    # 使用 set width 确保设置生效
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 45
    
    return wb, ws


def append_image(ws, row, path, column):
    """
    插入图片到指定单元格。
    如果图片插入成功，调整行高并返回 True。
    如果图片插入失败，在单元格内写入错误提示并返回 False。
    """
    # 检查路径是否存在
    if not path or not os.path.exists(path):
        cell = ws[f'{column}{row}']
        cell.value = f'文件缺失: {os.path.basename(path)}'
        cell.font = Font(color='FF0000', italic=True)
        return False

    # 检查 Pillow 是否可用
    if not PILLOW_AVAILABLE:
        cell = ws[f'{column}{row}']
        cell.value = f'无Pillow: {os.path.basename(path)}'
        cell.font = Font(color='FF0000', italic=True)
        return False

    try:
        img = OpenpyxlImage(path)
        # 保持原有的图片尺寸设置
        img.width = 220
        img.height = 130
        img.anchor = f'{column}{row}'
        ws.add_image(img)
        ws.row_dimensions[row].height = 120
        
        # --- 修改点 2: 确保不再重复设置列宽，因为已在 create_report_workbook 中统一设置 ---
        # 删除了末尾的重复代码块
            
        return True

    except Exception as e:
        cell = ws[f'{column}{row}']
        cell.value = f'加载失败: {os.path.basename(path)}'
        cell.font = Font(color='FF0000', italic=True)
        print(f"Error loading image {path}: {e}")
        return False

# --- 删除了原文件中第 54-70 行的重复代码块 ---
