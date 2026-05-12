"""
UI 自动化检查主程序

这个模块是 UI 自动化测试的主入口程序。
功能包括：
- 多语言 UI 界面检查
- 截图捕获和对比
- Excel 报告生成
- ADB 设备控制

使用方法：
1. 连接 ADB 设备
2. 运行 UI_checktest run.py
3. 选择要测试的语言
4. 等待自动测试完成

注意：需要安装相关依赖包
"""

import os
import subprocess
import time
import cv2
import numpy as np
from datetime import datetime
from openpyxl.drawing.image import Image as OpenpyxlImage

# 添加项目根目录到路径，以便导入 UI_checktest 包
# 这修复了相对导入问题
import sys
import argparse
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

# 导入项目模块
# 改为绝对导入，避免相对导入错误
from UI_checktest.constants import (
    BASE_RESULT_DIR,      # 结果存放目录
    SCREEN_RENDER,        # 页面渲染等待时间
    LANG_RELOAD,          # 语言切换重载时间
    HOME_RESET_STEPS,     # 返回首页的点击步骤
    GOTO_LANGUAGE_PAGE,   # 进入语言设置的步骤
    LANG_PAGE_NEXT_BTN,   # 语言列表翻页按钮
)
from UI_checktest.login import init_persistent_shell, start_evtest_background  # 登录和初始化
from UI_checktest.click import click_with_evtest  # 点击控制
from UI_checktest.page import UI_STRUCTURE, reset_to_home, normalize_actions, get_action_coords, is_page_definition  # 页面结构和导航
from UI_checktest.multilanguage import ALL_LANGUAGES, select_languages, build_language_data, check_language_switch  # 多语言处理
from UI_checktest.screenshot import capture_screen, set_local_dir, safe_imread  # 截图功能
from UI_checktest.compare import analyze_ui_diff  # 图像对比
from UI_checktest.report import create_report_workbook, append_image  # 报告生成
from UI_checktest.html_report import create_html_report  # HTML报告生成
from UI_checktest.utils import logger  # 日志记录


def page_allowed(selected_pages, main_page, sub_name=None, sub2_name=None):
    if not selected_pages:
        return True

    if main_page in selected_pages:
        return True
    if sub_name and sub_name in selected_pages:
        return True
    if sub2_name and sub2_name in selected_pages:
        return True
    if sub_name and sub2_name and f'{sub_name}_{sub2_name}' in selected_pages:
        return True

    return False


def any_page_allowed(selected_pages, main_page, config):
    if not selected_pages:
        return True

    if main_page in selected_pages:
        return True

    for sub_name, actions in config.get('sub_pages', {}).items():
        if sub_name in selected_pages:
            return True

        if isinstance(actions, dict):
            for sub2_name in actions.get('sub_pages', {}):
                if sub2_name in selected_pages or f'{sub_name}_{sub2_name}' in selected_pages:
                    return True
        elif isinstance(actions, list):
            # 列表项页面直接通过子页面名判断
            if sub_name in selected_pages:
                return True

    return False


def estimate_progress(selected_pages):
    if not selected_pages:
        return sum(1 for _ in UI_STRUCTURE)

    count = 0
    for main_page, config in UI_STRUCTURE.items():
        if main_page in selected_pages:
            count += 1
        for sub_name, actions in config.get('sub_pages', {}).items():
            if sub_name in selected_pages:
                count += 1
            if isinstance(actions, dict):
                for sub2_name in actions.get('sub_pages', {}):
                    if sub2_name in selected_pages or f'{sub_name}_{sub2_name}' in selected_pages:
                        count += 1
            else:
                if sub_name in selected_pages:
                    count += 1
    return max(count, 1)


def launch_gui():
    from UI_checktest.ui_qt import main as ui_main
    ui_main()


def main(selected_langs=None, selected_pages=None, status_callback=None):
    def report_status(message, percent=None):
        logger(message)
        if status_callback:
            if percent is not None:
                status_callback(f'PROGRESS:{percent}:{message}')
            else:
                status_callback(message)

    run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
    current_run_dir = os.path.join(BASE_RESULT_DIR, f'Run_{run_id}')
    os.makedirs(current_run_dir, exist_ok=True)

    subprocess.run(['adb', 'kill-server'], capture_output=True)
    subprocess.run(['adb', 'start-server'], capture_output=True)
    time.sleep(2.0)

    check_conn = subprocess.run(['adb', 'get-state'], capture_output=True, text=True)
    if 'device' not in check_conn.stdout:
        logger('[ERROR] 未检测到有效的 ADB 设备，程序退出！')
        return

    init_persistent_shell()
    time.sleep(2.0)
    start_evtest_background()

    if selected_langs is not None:
        target_langs = build_language_data(selected_langs)
    else:
        target_langs = select_languages()

    if not target_langs:
        return

    page_count = estimate_progress(selected_pages)
    total_steps = max(1, len(target_langs) * page_count)
    completed_steps = 0

    # 确保中文（CN）必选作为基准语言，并移动到首位
    cn_index = next((i for i, (_, name, _) in enumerate(target_langs) if name == '中文（CN）'), None)
    if cn_index is None:
        cn_coord = ALL_LANGUAGES['1'][1]
        target_langs.insert(0, (1, '中文（CN）', cn_coord))
        logger('[INIT] 自动添加中文（CN）为基准语言')
    elif cn_index != 0:
        target_langs.insert(0, target_langs.pop(cn_index))
        logger('[INIT] 将中文（CN）移动到语言列表首位，作为基准语言')

    set_local_dir(current_run_dir)
    report_status(f'[INIT] 本次测试结果将存放在: {current_run_dir}', percent=0)

    excel_path = os.path.join(current_run_dir, f'UI_Report_{run_id}.xlsx')
    wb, ws = create_report_workbook(excel_path)

    # 创建HTML报告
    html_report = create_html_report(current_run_dir, run_id)
    html_report.set_total_languages(len(target_langs))

    # --- 修改点 1: 初始化基准图字典，用于存储中文（CN）的所有页面截图 ---
    baseline_images = {}

    try:
        # 遍历所有语言
        for lang_idx, (lang_id, lang_name, lang_coord) in enumerate(target_langs):
            # --- 修改点 2: 明确判断是否为基准语言（中文） ---
            is_baseline_lang = (lang_name == '中文（CN）')
            lang_percent = int((lang_idx / len(target_langs)) * 100)
            report_status(f'\n================ 当前准备切换的语言: {lang_name} (是否基准语言: {is_baseline_lang}) ================', percent=lang_percent)

            # 返回首页
            for rx, ry in HOME_RESET_STEPS:
                click_with_evtest(rx, ry, page_label='回首页')
            time.sleep(5)

            # 进入语言设置
            for gx, gy in GOTO_LANGUAGE_PAGE:
                click_with_evtest(gx, gy, page_label='进入语言设置步骤')
                time.sleep(10.0)

            # 点击语言列表或翻页
            if lang_id < 13:
                click_with_evtest(169, 49, page_label='语言唤醒')
            else:
                click_with_evtest(LANG_PAGE_NEXT_BTN[0], LANG_PAGE_NEXT_BTN[1], page_label='翻页')

            time.sleep(10)
            logger('[SCREEN] 语言切换前，开始截图1')
            current_img1 = capture_screen(lang_name, 'LANG_BEFORE')

            # 切换语言
            click_with_evtest(lang_coord[0], lang_coord[1], page_label=f'切换至_{lang_name}')
            logger('[WAIT] 等待语言刷新...')
            time.sleep(LANG_RELOAD)

            logger('[SCREEN] 语言切换后，开始截图2')
            current_img2 = capture_screen(lang_name, 'LANG_AFTER')

            # --- 修改点 3: 语言切换验证的基准图处理 ---
            # 如果是基准语言，保存当前截图作为语言切换的基准图
            if is_baseline_lang:
                baseline_images['LANG_SWITCH'] = current_img2
            
            # 从字典中获取语言切换的基准图
            lang_base_img = baseline_images.get('LANG_SWITCH')
            
            # 如果不是基准语言且找不到基准图，跳过验证
            if not lang_base_img and not is_baseline_lang:
                logger('[WARN] 未找到语言切换的基准图，跳过验证')
            else:
                # 执行对比
                ok1, r1 = check_language_switch(lang_base_img, current_img1)
                ok2, r2 = check_language_switch(lang_base_img, current_img2)
                ok = ok1 or ok2
                ratio = max(r1, r2)
                status = 'PASS' if ok else 'WARN'
                if current_img1 is None or current_img2 is None:
                    status = 'FAIL_NO_IMAGE'

                logger(f'[LANG CHECK] diff={ratio:.4f}')
                if not ok:
                    logger(f'[WARN] 语言可能未切换或已是当前语言: {lang_name} (diff={ratio:.4f})')
                else:
                    logger('✅ 语言切换成功')

                # 写入 Excel
                ws.append([
                    lang_id,
                    lang_name,
                    'LANG',
                    '语言设置',
                    '切换验证',
                    lang_base_img, # 使用动态获取的基准图
                    current_img2,
                    ratio,
                    status
                ])
                append_image(ws, ws.max_row, lang_base_img, 'F')
                append_image(ws, ws.max_row, current_img2, 'G')

                # 更新HTML报告
                html_report.add_language_test(lang_id, lang_name, status, ratio, current_img1, current_img2)

            # 返回首页
            for i in range(2):
                click_with_evtest(63, 617, page_label=f'语言切换后返回首页_{i + 1}')
                time.sleep(15.0)

            last_page_was_home = True
            subprocess.run(['adb', 'shell', 'sync; echo 3 > /proc/sys/vm/drop_caches'], capture_output=True)
            ui_status = 'UNKNOWN'
            ui_diff_ratio = 0

            # 遍历所有UI页面
            for main_page, config in UI_STRUCTURE.items():
                if not any_page_allowed(selected_pages, main_page, config):
                    continue

                page_percent = int(((completed_steps + 1) / total_steps) * 100)
                report_status(f'\n--- 正在处理一级页面: {main_page} ---', percent=page_percent)

                if main_page == '首页':
                    last_page_was_home = True
                else:
                    if not last_page_was_home:
                        reset_to_home(main_page)
                    ex, ey = config.get('entry', (None, None))
                    if ex is not None:
                        click_with_evtest(ex, ey, page_label=f'进入_{main_page}')
                        last_page_was_home = False

                # 遍历子页面
                for sub_name, actions in config.get('sub_pages', {}).items():
                    if not page_allowed(selected_pages, main_page, sub_name):
                        continue
                    completed_steps += 1
                    sub_percent = int((completed_steps / total_steps) * 100)
                    report_status(f'  [SUB] {sub_name}', percent=sub_percent)
                    # 构建当前页面的唯一键名
                    page_key = f'{main_page}_{sub_name}'
                    
                    if isinstance(actions, dict):
                        sub_entry = actions.get('entry', (None, None))
                        if sub_entry and sub_entry[0] is not None:
                            click_with_evtest(sub_entry[0], sub_entry[1], page_label=f'{main_page}_{sub_name}_进入')
                            time.sleep(SCREEN_RENDER)

                        img_path = capture_screen(lang_name, page_key)
                        if img_path:
                            # --- 修改点 4: 核心逻辑：根据是否为基准语言处理基准图 ---
                            if is_baseline_lang:
                                baseline_images[page_key] = img_path
                                logger(f'[DEBUG] 基准图存储: {page_key} -> {img_path}')
                            else:
                                # 非基准语言：从字典中取出基准图
                                current_base_img_path = baseline_images.get(page_key)
                                if not current_base_img_path:
                                    logger(f'[WARN] 未找到 {page_key} 的基准图，已有基准图: {list(baseline_images.keys())}')
                                    logger(f'[WARN] 跳过此页面的对比测试')
                                    continue  # 跳过当前页面的测试
                                else:
                                    logger(f'[DEBUG] 基准图取出: {page_key} -> {current_base_img_path}')

                                # 执行对比
                                ui_status = analyze_ui_diff(current_base_img_path, img_path)
                                try:
                                    diff_img = safe_imread(current_base_img_path, 0)
                                    test_img = safe_imread(img_path, 0)
                                    diff_ratio = -1
                                    if diff_img is not None and test_img is not None:
                                        diff_ratio = cv2.absdiff(diff_img, test_img)
                                        diff_ratio = np.count_nonzero(diff_ratio) / diff_ratio.size
                                except Exception:
                                    diff_ratio = -1
                                
                                # 写入 Excel
                                ws.append([
                                    lang_id,
                                    lang_name,
                                    'UI',
                                    main_page,
                                    sub_name,
                                    current_base_img_path, # 使用动态计算的基准图路径
                                    img_path,
                                    diff_ratio,
                                    ui_status
                                ])
                                append_image(ws, ws.max_row, current_base_img_path, 'F')
                                append_image(ws, ws.max_row, img_path, 'G')

                        # 处理二级子页面 (SUB2)
                        for sub2_name, sub2_actions in actions.get('sub_pages', {}).items():
                            if not page_allowed(selected_pages, main_page, sub_name, sub2_name):
                                continue
                            completed_steps += 1
                            sub2_percent = int((completed_steps / total_steps) * 100)
                            report_status(f'    [SUB2] {sub2_name}', percent=sub2_percent)
                            # 构建二级页面的唯一键名
                            page_key_sub2 = f'{main_page}_{sub_name}_{sub2_name}'
                            
                            ax, ay = None, None
                            if is_page_definition(sub2_actions):
                                ax, ay = sub2_actions.get('entry', (None, None))
                            else:
                                normalized_sub2_actions = normalize_actions(sub2_actions)
                                for act in normalized_sub2_actions:
                                    act_x, act_y = get_action_coords(act)
                                    if act_x is not None:
                                        ax, ay = act_x, act_y
                                        break
                            if ax is not None:
                                click_with_evtest(ax, ay, page_label=f'{sub_name}_{sub2_name}')
                            time.sleep(SCREEN_RENDER)

                            img_path = capture_screen(lang_name, page_key_sub2)
                            if not img_path:
                                logger('[SKIP] SUB2截图失败')
                                continue
                            
                            # --- 修改点 5: SUB2 的基准图处理 ---
                            if is_baseline_lang:
                                baseline_images[page_key_sub2] = img_path
                                continue
                            
                            current_base_img_path = baseline_images.get(page_key_sub2)
                            if not current_base_img_path:
                                logger(f'[WARN] 未找到 {page_key_sub2} 的基准图，跳过此页面的对比测试')
                                continue

                            # 执行对比
                            ui_status = analyze_ui_diff(current_base_img_path, img_path)
                            try:
                                diff_img = safe_imread(current_base_img_path, 0)
                                test_img = safe_imread(img_path, 0)
                                if diff_img is None or test_img is None:
                                    diff_ratio = -1
                                else:
                                    diff_ratio = cv2.absdiff(diff_img, test_img)
                                    diff_ratio = np.count_nonzero(diff_ratio) / diff_ratio.size
                            except Exception as e:
                                logger(f'[DIFF ERROR] {e}')
                                diff_ratio = -1

                            # 写入 Excel
                            ws.append([
                                lang_id,
                                lang_name,
                                'UI',
                                main_page,
                                f'{sub_name}_{sub2_name}',
                                current_base_img_path, # 使用动态计算的基准图路径
                                img_path,
                                diff_ratio,
                                ui_status
                            ])
                            append_image(ws, ws.max_row, current_base_img_path, 'F')
                            append_image(ws, ws.max_row, img_path, 'G')

                            # 更新HTML报告
                            html_report.add_ui_test(lang_name, main_page, f'{sub_name}_{sub2_name}', ui_status, diff_ratio, current_base_img_path, img_path)
                        continue

                    # 处理列表项页面
                    actions_list = normalize_actions(actions)

                    for i, act in enumerate(actions_list):
                        if not page_allowed(selected_pages, main_page, sub_name):
                            continue
                        completed_steps += 1
                        list_percent = int((completed_steps / total_steps) * 100)
                        report_status(f'    [LIST] {sub_name}_{i}', percent=list_percent)
                        ax, ay = get_action_coords(act)
                        if ax is not None:
                            click_with_evtest(ax, ay, page_label=f'{main_page}_{sub_name}')
                        time.sleep(SCREEN_RENDER)

                        # 构建列表项页面的唯一键名
                        page_key_list = f'{main_page}_{sub_name}_{i}'
                        
                        img_path = capture_screen(lang_name, page_key_list)
                        
                        # --- 修改点 6: 列表项的基准图处理 ---
                        if is_baseline_lang:
                            if img_path:
                                baseline_images[page_key_list] = img_path
                            continue
                        
                        current_base_img_path = baseline_images.get(page_key_list)
                        if not current_base_img_path:
                            logger(f'[WARN] 未找到 {page_key_list} 的基准图，跳过此页面的对比测试')
                            continue

                        # 执行对比
                        diff_ratio = -1
                        status = 'FAIL_NO_IMAGE'
                        if img_path:
                            status = analyze_ui_diff(current_base_img_path, img_path)
                            try:
                                diff_img = safe_imread(current_base_img_path, 0)
                                test_img = safe_imread(img_path, 0)
                                if diff_img is not None and test_img is not None:
                                    diff_img = cv2.absdiff(diff_img, test_img)
                                    diff_ratio = np.count_nonzero(diff_img) / diff_img.size
                            except Exception as e:
                                logger(f'[DIFF ERROR] {e}')
                                diff_ratio = -1
                        else:
                            logger(f'[WARN] 未获取到页面截图: {page_key_list}')

                        # 写入 Excel
                        ws.append([
                            lang_id,
                            lang_name,
                            'UI',
                            main_page,
                            f'{sub_name}_{i}',
                            current_base_img_path,
                            img_path,
                            diff_ratio,
                            status
                        ])
                        append_image(ws, ws.max_row, current_base_img_path, 'F')
                        append_image(ws, ws.max_row, img_path, 'G')

                        # 更新HTML报告
                        html_report.add_ui_test(lang_name, main_page, f'{sub_name}_{i}', status, diff_ratio, current_base_img_path, img_path)
                    logger(f'[LANG DONE] {lang_name}')
    finally:
        # --- 修改点 7: 统一保存 Excel，解决列宽问题 ---
        logger(f'\n[FINISHED] 任务结束。报告: {excel_path}')
        
        # 强制确保 F 列和 G 列的宽度为 45
        ws.column_dimensions['F'].width = 45
        ws.column_dimensions['G'].width = 45
        
        wb.save(excel_path)
        logger('[EXCEL] 报告已保存')

        # 标记HTML报告为完成状态
        html_report.mark_completed()
        logger('[HTML] 报告已完成')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='UI 检查工具入口')
    parser.add_argument('--cli', action='store_true', help='直接运行命令行测试，不启动图形界面')
    parser.add_argument('--langs', nargs='+', help='指定要运行的语言ID或名称，例如: 1 2 3')
    parser.add_argument('--pages', nargs='+', help='指定要运行的页面名称，例如: 首页 配置')
    args = parser.parse_args()

    if args.cli:
        main(selected_langs=args.langs, selected_pages=args.pages)
    else:
        launch_gui()
