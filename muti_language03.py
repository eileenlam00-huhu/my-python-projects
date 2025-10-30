import re
import openpyxl
from openpyxl.styles import PatternFill, Font
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from datetime import datetime
import os


class CompleteDualDisplayTool:
    def __init__(self, root):
        self.root = root
        self.progress = None
        self.progress_label = None
        self.selected_languages = []
        self.language_checkboxes = []
        self.setup_ui()

    def setup_ui(self):
        """设置用户界面"""
        self.root.title("多语言文件对比工具")
        self.root.geometry("850x750")

        # 样式
        pad_y = 10
        btn_width = 35

        # 标题
        tk.Label(self.root, text="多语言文件对比工具", font=('Arial', 16, 'bold')).pack(pady=15)
        tk.Label(self.root, text="结果将按指定语言顺序排列对比", fg="blue").pack()
        tk.Label(self.root,
                 text="差异内容标记红色 | 一致内容标记绿色",
                 fg="red").pack(pady=5)

        # 进度显示区域
        progress_frame = tk.Frame(self.root)
        progress_frame.pack(pady=15, fill=tk.X, padx=30)

        self.progress_label = tk.Label(progress_frame, text="准备就绪", anchor=tk.W, font=('Arial', 10))
        self.progress_label.pack(fill=tk.X)

        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, mode='determinate')
        self.progress.pack(fill=tk.X)

        # 语言选择部分
        lang_frame = tk.LabelFrame(self.root, text="选择对比语言", padx=10, pady=10)
        lang_frame.pack(pady=10, padx=20, fill=tk.X)

        # 标准语言顺序
        self.LANGUAGE_ORDER = [
            "中文（CN）",
            "英文（EN）English",
            "德语(DE)Deutsch",
            "西语（ES）Español",
            "法语(FR)Français",
            "意大利语(IT)Italiano",
            "巴西葡语(BR)Português",
            "俄语（Pyc）Русский",
            "土耳其语(TR)Turkish",
            "日语(JP)日本語",
            "韩语(KR)한국어",
            "繁体中文",
            "阿拉伯语عربية"
        ]

        # 创建多选按钮
        self.language_vars = []
        for i, lang in enumerate(self.LANGUAGE_ORDER):
            var = tk.IntVar(value=1)  # 默认全选
            self.language_vars.append(var)
            cb = tk.Checkbutton(lang_frame, text=lang, variable=var)
            cb.grid(row=i // 4, column=i % 4, sticky=tk.W, padx=5, pady=2)
            self.language_checkboxes.append(cb)

        # 全选/取消全选按钮
        btn_frame = tk.Frame(lang_frame)
        btn_frame.grid(row=(len(self.LANGUAGE_ORDER) // 4 + 1), column=0, columnspan=4, pady=5)
        tk.Button(btn_frame, text="全选", command=self.select_all_languages).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="取消全选", command=self.deselect_all_languages).pack(side=tk.LEFT, padx=5)

        # 代码转Excel部分
        frame_convert = tk.Frame(self.root)
        frame_convert.pack(pady=10)
        tk.Label(frame_convert, text="第一步: 将代码文件转换为Excel", font=('Arial', 12)).pack()
        tk.Button(
            frame_convert, text="转换代码文件为Excel",
            command=self.start_conversion_thread, width=btn_width, height=2
        ).pack(pady=5)

        # 对比部分
        frame_compare = tk.Frame(self.root)
        frame_compare.pack(pady=10)
        tk.Label(frame_compare, text="第二步: 对比Excel文件", font=('Arial', 12)).pack()

        # 对比顺序说明
        tk.Label(frame_compare,
                 text="对比顺序:\n1. 按键名匹配行\n2. 按指定语言顺序对比\n3. 标记差异(红)和一致(绿)",
                 justify=tk.LEFT).pack()

        tk.Button(
            frame_compare, text="开始对比",
            command=self.start_comparison_thread, width=btn_width, height=2
        ).pack(pady=5)

        # 退出按钮
        tk.Button(self.root, text="退出", command=self.root.quit, width=btn_width - 5).pack(pady=15)

    def select_all_languages(self):
        """全选所有语言"""
        for var in self.language_vars:
            var.set(1)

    def deselect_all_languages(self):
        """取消全选所有语言"""
        for var in self.language_vars:
            var.set(0)

    def get_selected_languages(self):
        """获取用户选择的语言列表"""
        selected = []
        for i, var in enumerate(self.language_vars):
            if var.get() == 1:
                selected.append(self.LANGUAGE_ORDER[i])
        return selected if selected else self.LANGUAGE_ORDER  # 默认全选

    def get_output_filename(self, suffix):
        """生成带日期时间的输出文件名"""
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{now}_{suffix}.xlsx"

    def update_progress(self, value, message):
        """更新进度条和标签"""
        self.progress['value'] = value
        self.progress_label['text'] = message
        self.root.update_idletasks()

    def start_conversion_thread(self):
        """启动转换线程"""
        thread = threading.Thread(target=self.convert_code_to_excel)
        thread.start()

    def start_comparison_thread(self):
        """启动对比线程"""
        thread = threading.Thread(target=self.compare_excel_files)
        thread.start()

    def convert_code_to_excel(self):
        """将C语言多语言代码转换为Excel"""
        try:
            self.update_progress(0, "正在选择代码文件...")

            code_file = filedialog.askopenfilename(
                title="选择代码文件",
                filetypes=[("C files", "*.c *.h"), ("All files", "*.*")]
            )
            if not code_file:
                self.update_progress(0, "操作已取消")
                return

            output_filename = self.get_output_filename("mutilanguagetest")
            output_excel = filedialog.asksaveasfilename(
                title="保存转换后的Excel文件",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_filename
            )
            if not output_excel:
                self.update_progress(0, "操作已取消")
                return

            self.update_progress(20, "正在读取代码文件...")
            with open(code_file, 'r', encoding='utf-8') as f:
                content = f.read()

            self.update_progress(30, "正在解析代码结构...")
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Translations"

            # 添加表头
            headers = ["Key"] + self.LANGUAGE_ORDER
            ws.append(headers)

            # 匹配多语言数组
            pattern = re.compile(
                r'const char \*(.*?)\[MAX_LANGUAGE\]\s*=\s*\{([^}]*)\};',
                re.DOTALL
            )

            matches = list(pattern.finditer(content))
            total_matches = len(matches)

            self.update_progress(40, f"发现{total_matches}个多语言条目...")

            for i, match in enumerate(matches):
                key = match.group(1).strip()
                values_block = match.group(2)

                # 提取各语言字符串（处理可能的多行情况）
                values = re.findall(r'"(.*?)"', values_block.replace("\n", ""))

                # 写入Excel行
                row = [key] + values[:len(headers) - 1]  # 确保不超过语言列数
                ws.append(row)

                progress = 40 + (i / total_matches) * 50
                self.update_progress(progress, f"正在处理 {i + 1}/{total_matches}...")

            self.update_progress(95, "正在保存Excel文件...")
            wb.save(output_excel)

            self.update_progress(100, "转换完成！")
            messagebox.showinfo("成功", f"转换成功:\n{os.path.basename(output_excel)}")

        except Exception as e:
            self.update_progress(0, f"错误: {str(e)}")
            messagebox.showerror("错误", f"转换失败:\n{str(e)}")
        finally:
            self.root.after(2000, lambda: self.update_progress(0, "准备就绪"))

    def compare_excel_files(self):
        """对比两个Excel文件，按指定语言顺序排列结果"""
        try:
            # 获取用户选择的语言
            selected_languages = self.get_selected_languages()
            print(f"调试: 选择的语言列表: {selected_languages}")
            self.update_progress(0, f"已选择语言: {', '.join(selected_languages)}")

            self.update_progress(5, "正在选择源文件...")
            source_file = filedialog.askopenfilename(
                title="选择源Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not source_file:
                self.update_progress(0, "操作已取消")
                return
            print(f"调试: 源文件: {source_file}")

            self.update_progress(10, "正在选择翻译文件...")
            trans_file = filedialog.askopenfilename(
                title="选择翻译Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not trans_file:
                self.update_progress(0, "操作已取消")
                return
            print(f"调试: 翻译文件: {trans_file}")

            output_filename = self.get_output_filename("comparison_result")
            output_file = filedialog.asksaveasfilename(
                title="保存对比结果",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_filename
            )
            if not output_file:
                self.update_progress(0, "操作已取消")
                return
            print(f"调试: 输出文件: {output_file}")

            self.update_progress(20, "正在加载文件...")

            # 加载工作簿
            source_wb = openpyxl.load_workbook(source_file)
            trans_wb = openpyxl.load_workbook(trans_file)
            output_wb = openpyxl.Workbook()

            source_ws = source_wb.active
            trans_ws = trans_wb.active
            output_ws = output_wb.active
            output_ws.title = "对比结果"

            print(f"调试: 源文件工作表 - 行数: {source_ws.max_row}, 列数: {source_ws.max_column}")
            print(f"调试: 翻译文件工作表 - 行数: {trans_ws.max_row}, 列数: {trans_ws.max_column}")

            # 定义颜色和样式
            RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            GREEN_FILL = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            YELLOW_FILL = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')  # 黄色用于缺失
            HEADER_FONT = Font(bold=True, size=12)
            DIFF_FONT = Font(bold=True, color='FF0000')
            MATCH_FONT = Font(color='006400')
            MISSING_FONT = Font(bold=True, color='FFA500')  # 橙色用于缺失
            HEADER_FILL = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')

            # 设置列宽
            output_ws.column_dimensions['A'].width = 25  # 键名
            output_ws.column_dimensions['B'].width = 20  # 语言
            output_ws.column_dimensions['C'].width = 30  # 源文件内容
            output_ws.column_dimensions['D'].width = 30  # 翻译文件内容
            output_ws.column_dimensions['E'].width = 15  # 对比结果

            # 添加表头
            headers = ["键名", "语言", "源文件内容", "翻译文件内容", "对比结果"]
            output_ws.append(headers)
            print("调试: 已添加表头")

            # 设置表头样式
            for col in range(1, 6):
                cell = output_ws.cell(row=1, column=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL

            # 1. 找到中文列的索引
            def find_chinese_col(ws, file_type):
                chinese_names = ["中文（CN）", "中文(CN)", "中文", "Chinese", "CN", "简体中文", "中文简体"]
                print(f"调试: 在{file_type}文件中查找中文列...")

                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col).value
                    header = str(header_value).strip() if header_value is not None else ""
                    print(f"调试: {file_type}文件第{col}列: '{header}'")

                    if any(name in header for name in chinese_names):
                        print(f"调试: 找到中文列: 第{col}列 '{header}'")
                        return col

                    if re.search(r'中[文国]|汉|简体', header):
                        print(f"调试: 通过正则找到中文列: 第{col}列 '{header}'")
                        return col

                print(f"调试: 未找到中文列，使用第1列")
                return 1

            source_cn_col = find_chinese_col(source_ws, "源")
            trans_cn_col = find_chinese_col(trans_ws, "翻译")
            print(f"调试: 源文件中文列: {source_cn_col}, 翻译文件中文列: {trans_cn_col}")

            # 2. 构建键映射 - 只构建源文件的键映射
            def build_key_map(ws, cn_col, file_type):
                key_map = {}
                print(f"调试: 构建{file_type}文件键映射，使用第{cn_col}列作为键")

                for row in range(2, ws.max_row + 1):
                    key_cell = ws.cell(row=row, column=cn_col)
                    key_value = key_cell.value

                    if key_value is not None:
                        try:
                            key_str = str(key_value).strip()
                            if key_str:
                                key_map[key_str] = row
                        except Exception as e:
                            print(f"调试: 处理{file_type}文件行{row}时出错: {e}")

                print(f"调试: {file_type}文件总共找到 {len(key_map)} 个有效键")
                return key_map

            # 只构建源文件的键映射，以源文件为主
            source_key_map = build_key_map(source_ws, source_cn_col, "源")
            trans_key_map = build_key_map(trans_ws, trans_cn_col, "翻译")

            # 3. 只使用源文件的键进行对比
            sorted_keys = sorted(source_key_map.keys(), key=lambda x: str(x).lower())
            print(f"调试: 源文件共有 {len(sorted_keys)} 个键需要对比")
            print(f"调试: 前5个键: {sorted_keys[:5]}")

            # 4. 定义语言列映射
            def build_lang_col_map(ws):
                lang_map = {}
                print(f"调试: 构建语言列映射...")
                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col).value
                    header = str(header_value).strip() if header_value is not None else f"列{col}"
                    lang_map[header] = col
                print(f"调试: 找到的语言列: {list(lang_map.keys())}")
                return lang_map

            source_lang_map = build_lang_col_map(source_ws)
            trans_lang_map = build_lang_col_map(trans_ws)

            # 5. 构建对比语言列表 - 只使用用户选择的语言
            print("调试: 开始构建对比语言列表...")

            comparison_langs = []
            for lang in selected_languages:
                # 检查语言列在源文件中是否存在
                if lang in source_lang_map:
                    comparison_langs.append(lang)
                else:
                    print(f"调试: 跳过语言 '{lang}' - 在源文件中不存在")

            print(f"调试: 最终对比语言列表: {comparison_langs}")
            print(f"调试: 将对比 {len(comparison_langs)} 种语言")

            # 6. 执行对比 - 以源文件为主
            differences_found = False
            output_row = 2  # 从第2行开始输出
            total_comparisons = 0
            missing_in_trans = 0  # 统计翻译文件中缺失的键

            print(f"调试: 开始对比 {len(sorted_keys)} 个键...")

            if not comparison_langs:
                output_ws.cell(row=2, column=1).value = "错误：未找到可对比的语言列"
                output_ws.cell(row=2, column=1).font = Font(bold=True, color='FF0000')
                print("调试: 错误：未找到可对比的语言列")
            else:
                for i, cn_key in enumerate(sorted_keys):
                    src_row = source_key_map.get(cn_key)
                    trans_row = trans_key_map.get(cn_key)

                    print(f"调试: 处理键 '{cn_key}' - 源文件行: {src_row}, 翻译文件行: {trans_row}")

                    # 记录这个键是否进行了任何对比
                    key_compared = False

                    for lang in comparison_langs:
                        # 检查语言列在源文件中是否存在
                        source_col = source_lang_map.get(lang)
                        if not source_col:
                            continue

                        print(f"调试: 对比语言: {lang}")

                        # 获取源文件值
                        src_val = ""
                        if src_row:
                            try:
                                src_cell = source_ws.cell(row=src_row, column=source_col)
                                src_val = str(src_cell.value) if src_cell.value is not None else ""
                            except Exception as e:
                                print(f"调试: 读取源文件值出错: {e}")
                                src_val = "【读取错误】"

                        # 获取翻译文件值
                        trans_val = ""
                        if trans_row:
                            try:
                                trans_col = trans_lang_map.get(lang)
                                if trans_col:
                                    trans_cell = trans_ws.cell(row=trans_row, column=trans_col)
                                    trans_val = str(trans_cell.value) if trans_cell.value is not None else ""
                                else:
                                    trans_val = "【语言列缺失】"
                            except Exception as e:
                                print(f"调试: 读取翻译文件值出错: {e}")
                                trans_val = "【读取错误】"
                        else:
                            trans_val = "【代码文件缺失】"
                            missing_in_trans += 1

                        print(f"调试: 语言 '{lang}' - 源值: '{src_val}', 翻译值: '{trans_val}'")

                        # 写入键名（只在该语言组的第一行写入）
                        if not key_compared:
                            output_ws.cell(row=output_row, column=1).value = cn_key
                            output_ws.cell(row=output_row, column=1).font = Font(bold=True)
                            key_compared = True

                        # 写入语言标签
                        output_ws.cell(row=output_row, column=2).value = lang

                        # 写入源内容
                        output_ws.cell(row=output_row, column=3).value = src_val

                        # 写入翻译内容
                        output_ws.cell(row=output_row, column=4).value = trans_val

                        # 对比并标记结果
                        if trans_val == "【代码文件缺失】":
                            # 翻译文件中缺失该键
                            output_ws.cell(row=output_row, column=5).value = "⚠ 缺失"
                            output_ws.cell(row=output_row, column=5).font = MISSING_FONT
                            output_ws.cell(row=output_row, column=3).fill = YELLOW_FILL
                            output_ws.cell(row=output_row, column=4).fill = YELLOW_FILL
                            differences_found = True
                            print(f"调试: 键 '{cn_key}' 语言 '{lang}' - 代码文件缺失")
                        elif src_val == trans_val:
                            # 内容一致
                            output_ws.cell(row=output_row, column=5).value = "✓ 一致"
                            output_ws.cell(row=output_row, column=5).font = MATCH_FONT
                            output_ws.cell(row=output_row, column=3).fill = GREEN_FILL
                            output_ws.cell(row=output_row, column=4).fill = GREEN_FILL
                            print(f"调试: 键 '{cn_key}' 语言 '{lang}' - 一致")
                        else:
                            # 内容差异
                            output_ws.cell(row=output_row, column=5).value = "✗ 差异"
                            output_ws.cell(row=output_row, column=5).font = DIFF_FONT
                            output_ws.cell(row=output_row, column=3).fill = RED_FILL
                            output_ws.cell(row=output_row, column=4).fill = RED_FILL
                            differences_found = True
                            print(f"调试: 键 '{cn_key}' 语言 '{lang}' - 发现差异")

                        output_row += 1
                        total_comparisons += 1

                    # 如果这个键进行了对比，添加空行分隔
                    if key_compared:
                        output_row += 1

                    # 更新进度
                    progress = 40 + (i / len(sorted_keys)) * 50
                    if i % 10 == 0 or i == len(sorted_keys) - 1:
                        self.update_progress(progress, f"正在对比 {i + 1}/{len(sorted_keys)}...")

            print(f"调试: 总共完成了 {total_comparisons} 次对比")
            print(f"调试: 发现差异: {differences_found}")
            print(f"调试: 翻译文件中缺失的键: {missing_in_trans} 个")
            print(f"调试: 输出表格总行数: {output_row}")

            # 添加总结信息
            if total_comparisons > 0:
                summary_row = output_row + 2
                output_ws.cell(row=summary_row, column=1).value = "对比总结"
                output_ws.cell(row=summary_row, column=1).font = HEADER_FONT

                summary_text = []
                if missing_in_trans > 0:
                    summary_text.append(f"翻译文件缺失: {missing_in_trans} 个键")
                if differences_found:
                    summary_text.append("发现差异内容")
                else:
                    summary_text.append("所有内容完全一致")

                output_ws.cell(row=summary_row + 1, column=1).value = " | ".join(summary_text)
                if differences_found:
                    output_ws.cell(row=summary_row + 1, column=1).font = DIFF_FONT
                else:
                    output_ws.cell(row=summary_row + 1, column=1).font = MATCH_FONT

                # 添加颜色说明
                output_ws.cell(row=summary_row + 3, column=1).value = "颜色说明:"
                output_ws.cell(row=summary_row + 3, column=1).font = Font(bold=True)

                # 绿色说明
                green_cell = output_ws.cell(row=summary_row + 4, column=1)
                green_cell.value = "绿色: 内容完全一致"
                green_cell.fill = GREEN_FILL

                # 红色说明
                red_cell = output_ws.cell(row=summary_row + 5, column=1)
                red_cell.value = "红色: 内容存在差异"
                red_cell.fill = RED_FILL

                # 黄色说明
                yellow_cell = output_ws.cell(row=summary_row + 6, column=1)
                yellow_cell.value = "黄色: 代码文件缺失该键"
                yellow_cell.fill = YELLOW_FILL

            else:
                output_ws.cell(row=2, column=1).value = "未找到可对比的数据"
                output_ws.cell(row=2, column=1).font = Font(bold=True, color='FF0000')

            self.update_progress(95, "正在保存结果...")
            output_wb.save(output_file)

            msg = "对比完成！\n\n"
            msg += "结果文件格式说明:\n"
            msg += "- 第1列: 键名\n"
            msg += "- 第2列: 语言标签\n"
            msg += "- 第3列: 源文件内容\n"
            msg += "- 第4列: 翻译文件内容\n"
            msg += "- 第5列: 对比结果\n\n"
            msg += f"结果已保存到:\n{os.path.basename(output_file)}"
            msg += f"\n\n统计信息:"
            msg += f"\n- 源文件键数: {len(sorted_keys)}"
            msg += f"\n- 完成对比: {total_comparisons} 次"
            msg += f"\n- 翻译文件缺失: {missing_in_trans} 个键"

            self.update_progress(100, "对比完成！")
            messagebox.showinfo("完成", msg)

        except Exception as e:
            print(f"调试: 发生异常: {str(e)}")
            import traceback
            traceback.print_exc()
            self.update_progress(0, f"错误: {str(e)}")
            messagebox.showerror("错误", f"对比失败:\n{str(e)}")
        finally:
            self.root.after(2000, lambda: self.update_progress(0, "准备就绪"))

if __name__ == "__main__":
    root = tk.Tk()
    app = CompleteDualDisplayTool(root)
    root.mainloop()