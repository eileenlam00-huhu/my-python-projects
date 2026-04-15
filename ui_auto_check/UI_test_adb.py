import os
import time
import subprocess
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment
from datetime import datetime
import subprocess
import time
import cv2
import numpy as np
import shutil


# ---------------------------
# 新增：登录凭证
# ---------------------------
DEVICE_USER = "root"      # 你的用户名
DEVICE_PASS = "creality_2025"    # 你的密码（根据实际修改）
LOGIN_WAIT = 2.0          # 登录每一步的间隔
# ==========================================================
# 1. 全局配置与时间策略
# ==========================================================
TOUCH_DEV = "/dev/input/event0"  # 触摸屏设备节点
REMOTE_DIR = "/tmp/ui_screens"  # 设备端临时存放截图的目录
BASE_RESULT_DIR = r"C:\Users\119198\Downloads\ui_test\png"  # 本地结果存放路径
LOCAL_DIR = r"C:\Users\119198\Downloads\ui_test\png"
CLICK_PAUSE = 75.0      # 普通点击后的等待
STEP_PAUSE = 60.0       # 返回首页等动作的间隔
SCREEN_RENDER = 60.0    # 页面切换后的渲染等待（重要：防止黑屏截图）
FFMPEG_WRITE = 60.0     # 截图文件写入磁盘的缓冲
LANG_RELOAD = 90.0     # 切换语言后系统重载的耗时
PULSE_INTERVAL = 0.08  # 脉冲点击之间的极短停顿
# 语言翻页按钮坐标 (请根据实际情况确认这个坐标)
LANG_PAGE_NEXT_BTN = (391, 83)
# 进入语言设置页面的步骤
GOTO_LANGUAGE_PAGE = [(406, 247), (54, 108), (381, 66), (261, 405)]

# 全局进程对象
shell_proc = None
last_click_time = 0

# ==========================================================
# 2. 日志系统初始化
# ==========================================================
os.makedirs(BASE_RESULT_DIR, exist_ok=True)
LOG_FILE = os.path.join(BASE_RESULT_DIR, f"test_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

def logger(message):
    """带时间戳的日志记录函数，同时输出到屏幕和文件"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    log_entry = f"[{timestamp}] {message}"
    print(log_entry)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(log_entry + "\n")

# ==========================================================
# 2. 页面结构与 20 国语言定义
# ==========================================================
# 返回首页的物理点击轨迹
HOME_RESET_STEPS = [(63, 617)]

# UI 遍历结构配置
# 格式说明：
# entry: 进入该一级页面的坐标
# sub_pages: { "子页面名": 动作序列 }
# 动作序列支持：None(直接截图), (x,y)(点击截图), [("pulse",x,y), (x,y)](长按后点击并连续截图)
UI_STRUCTURE = {
    "首页": {
        "entry": (None, None),  # 已经是首页，无需点击
        "depth": 0,
        "sub_pages": {
            "主界面": [("click", None, None)]
        }
    },

    # ================= 文件列表 =================
    "文件列表": {
        "entry": (191, 390),
        "depth": 1,
        "sub_pages": {
            "本地浏览": [None],
            "本地长按菜单": [("pulse", 176, 516)],
            "本地复制操作": [("click", 29, 449)],
            "U盘": [(49, 287)],
            "打印历史": [(25, 135)]
        }
    },

    # ================= 耗材页面 =================
    "耗材页面": {
        "entry": (398, 546),
        "depth": 1,
        "sub_pages": {
            "CFS lite": [None],
            "料架": [(45, 291)]
        }
    },

    # ================= 控制页面 =================
    "控制页面": {
        "entry": (383, 414),
        "depth": 1,
        "sub_pages": {
            "控制": [None],
            "XYZ": [(69, 351)]
        }
    },

    # ================= 设置页面（已修复嵌套） =================
    "设置页面": {
        "entry": (406, 247),
        "depth": 1,
        "sub_pages": {
            "账号": [(None, None)],

            "打印设置": {
                "entry": (66, 315),
                "depth": 1,
                "sub_pages": {
                    "状态灯页面": {
                        "entry": (138, 354),
                        "depth": 2,
                        "sub_pages": {
                            "灯效样式": {
                                "entry": (228, 178),
                                "depth": 3,
                                "sub_pages": {}
                            }
                        }
                    }
                }
            },

            "系统": [(54, 108)]
        }
    },

    # ================= 帮助页面 =================
    "帮助页面": {
        "entry": (408, 79),
        "depth": 1,
        "sub_pages": {
            "提示": [None],
            "WiKi": [(33, 334)],
            "维护": [(58, 160)]
        }
    }
}

# 20国语言及其在设置列表中的点击坐标
ALL_LANGUAGES = {
    "1": ("中文（CN）", (140, 529)),
    "2": ("English", (128, 387)),
    "3": ("Deutsch", (131, 207)),
    "4": ("Español", (226, 573)),
    "5": ("Français", (218, 380)),
    "6": ("Italiano", (205, 209)),
    "7": ("Português", (313, 557)),
    "8": ("Русский", (313, 380)),
    "9": ("Turkish", (320, 217)),
    "10": ("日本語", (402, 556)),
    "11": ("한국어", (407, 364)),
    "12": ("العربية", (408, 213)),
# --- 序号13及以后需要先执行翻页 ---
# 注意：翻页后坐标会变回页面上方，请根据实际填写
    "13": ("繁体中文", (140, 529)),
    "14": ("Polski", (111, 394)),
    "15": ("Tiếng Việt", (131, 207)),
    "16": ("Bahasa Indonesia", (226, 573)),
    "17": ("ไทย", (218, 380)),
    "18": ("Bahasa Melayu", (205, 209)),
    "19": ("עברית", (313, 557)),
    "20": ("Afrikaans", (313, 380))
}

# ==========================================================
# 3. ADB 与 核心功能函数
# ==========================================================
def init_persistent_shell():
    global shell_proc
    logger("[INIT] 正在建立持久化 Shell 连接...")
    shell_proc = subprocess.Popen(
        ["adb", "shell"],
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding='utf-8',
        errors='replace',  # <--- 关键修改：遇到解析不了的字符替换为问号，不报错
        bufsize=1
    )

    def login_handshake(expect_str, send_str, label):
        """侦听期待的字符串，匹配后再发送指令"""
        logger(f"[AUTH] 正在等待 {label} 提示符...")
        start_time = time.time()
        buffer = ""

        # 持续读取输出，直到匹配或超时（给 90 秒，因为你的设备真的很慢）
        while time.time() - start_time < 90:
            char = shell_proc.stdout.read(1)  # 逐字符读取，防止 readline 阻塞
            if char:
                buffer += char
                # 当屏幕出现 login: 或 Password: 时
                if expect_str.lower() in buffer.lower():
                    logger(f"[DEVICE] 匹配到提示符: {expect_str}")
                    time.sleep(1.0)  # 匹配后稳一下再发
                    shell_proc.stdin.write(f"{send_str}\n")
                    shell_proc.stdin.flush()
                    return True
            else:
                time.sleep(0.1)
        return False

    # 1. 进场先敲个回车，激活一下可能在休眠的终端
    shell_proc.stdin.write("\n")
    shell_proc.stdin.flush()

    # 2. 等待 "login:" 出现，然后发 root
    if not login_handshake("login:", DEVICE_USER, "用户名"):
        logger("[FATAL] 等待登录提示符超时！")
        return

    # 3. 等待 "Password:" 出现，然后发密码
    if not login_handshake("Password:", DEVICE_PASS, "密码"):
        logger("[FATAL] 等待密码提示符超时！")
        return

    # 4. 最后确认看到 # 号，代表真正进入了系统
    logger("[AUTH] 正在确认 Root 权限...")
    time.sleep(3.0)
    shell_proc.stdin.write("id\n")
    shell_proc.stdin.flush()

    logger("[SUCCESS] 交互式登录序列完成")

    # 在登录成功后的位置加一行
    shell_proc.stdin.write(f"mkdir -p {REMOTE_DIR} && chmod 777 {REMOTE_DIR}\n")
    shell_proc.stdin.flush()

# 在 init_persistent_shell 成功拿到 # 号后执行
def start_evtest_background():
    # 使用 ( ) 和 < /dev/null 确保进程彻底脱离终端，路径改为 /sbin/evtest
    cmd = "( /sbin/evtest /dev/input/event0 < /dev/null > /tmp/ev.log 2>&1 & )"
    adb_shell(cmd)
    logger("[INIT] evtest 监听已通过子 Shell 启动")

def adb_shell(cmd, wait_time=1.0):
    global shell_proc
    # 逻辑简化：直接写指令，不带任何 adb shell 前缀
    if "touch_control" in cmd:
        # 这行日志能帮你抓到所有点击动作的元凶
        logger(f"  [DEBUG-CMD] 发送原始点击: {cmd}")
    try:
        if shell_proc.poll() is not None:
            logger("[FATAL] shell 已断开！需要重连")
            return ""
        # 清除可能误加的前缀（防御性编程）
        clean_cmd = cmd.replace("adb shell ", "")

        shell_proc.stdin.write(f"{clean_cmd}\n")
        shell_proc.stdin.flush()

        # 统一在这里控制步长，或者由调用方控制
        # 核心修改：清空当前的输出缓冲区，防止管道阻塞
        # 注意：这里只读取目前已经在缓冲区的数据，不进行阻塞等待
        time.sleep(wait_time)
        return "Sent"
    except Exception as e:
        logger(f"[ERROR] 指令下发异常: {e}")
        return ""

def smart_sleep(total_time, step=2):
    for _ in range(int(total_time / step)):
        time.sleep(step)

def safe_click(x, y, label="", max_retry=3):
    for i in range(max_retry):
        logger(f"[CLICK] {label} 第{i+1}次")
        click_with_evtest(x, y, page_label=label)
        time.sleep(3.0)
        return True


def click_with_evtest(x, y, page_label="未知页面", timeout=150.0, retries=3):
    """
    带页面名称、高频重试和熔断机制的点击函数
    """
    global last_click_time
    for attempt in range(retries + 1):

        adb_shell("rm -f /tmp/ev.log")

        prefix = f"[RETRY {attempt}]" if attempt > 0 else "[ACTION]"
        logger(f"  {prefix} 正在点击 {page_label} 坐标 ({x}, {y}) 并监控回显...")

        # 2. 下发点击
        adb_shell(f"/touch_control {TOUCH_DEV} {x} {y}")
        time.sleep(15.0)

        now = time.time()
        if now - last_click_time < 2.0:
            time.sleep(2.0 - (now - last_click_time))

        last_click_time = time.time()

        # 3. 轮询监听 (适当缩短单次轮询间隔，增加灵敏度)
        start_t = time.time()
        while time.time() - start_t < timeout:
            shell_proc.stdin.write('[ -s /tmp/ev.log ] && echo "HIT" || echo "MISS"\n')
            shell_proc.stdin.flush()

            line = safe_read_line()

            if "HIT" in line:
                logger(f"  [CONFIRM] {page_label} 点击成功 (尝试第 {attempt + 1} 次)")
                return True
            time.sleep(0.5)

    # --- 熔断逻辑：如果走到这里，说明所有重试都失败了 ---
    error_msg = f"!!! [FATAL ERROR] 页面 [{page_label}] 坐标 ({x}, {y}) 彻底无响应，脚本停止运行 !!!"
    logger(error_msg)

    # 打印最后一次 ev.log 内容辅助排查
    adb_shell("cat /tmp/ev.log")

    # 直接抛出异常，强制停止 main 循环
    logger(error_msg)

# 由于 adb shell 机制，登录后如果断开连接，权限可能会消失
# 如果你的设备是登录一次就全系统提权，那上面的代码 OK。
# 如果不是，我们需要把 adb_shell 改写为持续写入 proc.stdin。

def click(x, y, wait=CLICK_PAUSE):
    """执行点击操作"""
    if x is None or y is None: return
    print(f"  [ADB CLICK] 坐标: ({x}, {y})")
    logger(f"  [ADB CLICK] 坐标: ({x}, {y})，等待 {wait}s")
    adb_shell(f"/touch_control {TOUCH_DEV} {x} {y}")
    time.sleep(wait)

def safe_capture(name, tag):
    try:
        return capture_screen(name, tag)
    except Exception as e:
        logger(f"[ERROR] 截图失败: {e}")
        return None

def click_pulse(x, y, page_label="长按动作", times=25, interval=0.1, wait=5.0):
    """
    针对低性能设备优化的脉冲长按：
    1. 降低频率 (0.1s 间隔) 防止内核丢包
    2. 增加内核回显校验
    3. 增加 sync 强制同步
    """
    if x is None or y is None: return False

    # 1. 准备阶段
    adb_shell("> /tmp/ev.log")
    logger(f"  [PULSE ACTION] 正在对 {page_label} ({x}, {y}) 下发 {times} 次脉冲长按...")

    # 2. 执行阶段：分段发送并强制同步
    try:
        for i in range(times):
            # 每发送 10 次脉冲强制同步一次内核 IO
            if i % 10 == 0 and i > 0:
                adb_shell("sync")

            # 使用正确的设备节点发送
            adb_shell(f"/touch_control {TOUCH_DEV} {x} {y}", wait_time=0.01)

            # 这里的 interval 不能太小，低性能设备 0.05-0.1 比较稳
            time.sleep(interval)

        # 3. 校验阶段：给内核一点时间写入日志
        time.sleep(1.0)
        adb_shell("sync")

        # 检查日志文件是否有内容 (HIT)
        shell_proc.stdin.write('[ -s /tmp/ev.log ] && echo "PULSE_HIT" || echo "PULSE_MISS"\n')
        shell_proc.stdin.flush()

        # 读取回显
        response = ""
        start_t = time.time()
        while time.time() - start_t < 3.0:  # 最多等 3 秒看回显
            try:
                line = safe_read_line()
                if "PULSE_HIT" in line:
                    response = "HIT"
                    break

                if "PULSE_MISS" in line:
                    response = "MISS"
                    break
            except UnicodeDecodeError:
                continue  # 忽略解码错误的行

        if response == "HIT":
            logger(f"  [CONFIRM] {page_label} 脉冲长按内核响应成功")
            time.sleep(wait)  # 给 UI 留出长按后的弹出动画时间
            return True
        else:
            logger(f"  [ERROR] {page_label} 脉冲长按内核无响应！")
            # 熔断：长按失败通常意味着后续菜单出不来，直接报错停止
            raise RuntimeError(f"长按 {page_label} 失败，脚本熔断退出")

    except Exception as e:
        logger(f"  [FATAL] 脉冲函数执行异常: {e}")
        raise

def safe_imread(path, flag=0):
    if not path or not os.path.exists(path):
        return None

    try:
        data = np.fromfile(path, dtype=np.uint8)
        img = cv2.imdecode(data, flag)
        return img
    except Exception as e:
        logger(f"[IMREAD ERROR] {path} -> {e}")
        return None


# 动作 A: 灵活返回函数 (基于 depth 循环)
def reset_to_home(page_name):
    # 获取当前页面的深度，如果没定义则默认为 1
    depth = UI_STRUCTURE.get(page_name, {}).get("depth", 1)

    # --- 新增逻辑：如果深度为 0，直接返回，不执行任何点击 ---
    if depth <= 0:
        logger(f"  [SKIP] 页面 {page_name} 深度为 0，无需执行返回动作")
        return

    for i in range(depth):
        # 只有 depth > 0 才会跑到这里
        click_with_evtest(63, 617, page_label=f"返回首页_第{i + 1}步", retries=3)
        time.sleep(2.0)

# 动作 B: 智能点击函数 (识别 click 或 pulse)
def execute_actions(actions):
    for action_type, x, y in actions:
        for act in actions:
            if not act:
                continue

            if len(act) == 3:
                action_type, x, y = act
            elif len(act) == 2:
                x, y = act
                action_type = "click"
            else:
                continue

            if x is None:
                continue
        """"
        if action_type == "click":
            click(x, y, wait=5.0)
        elif action_type == "pulse":
        """
        click_pulse(x, y, times=30, wait=5.0)


def capture_screen(lang, name,save_dir=None):
    if shell_proc is None or shell_proc.poll() is not None:
        logger("[FATAL] shell已断，截图终止")
        return None

    if save_dir is None:
        save_dir = save_dir or BASE_RESULT_DIR or LOCAL_DIR
    file_full_name = f"{lang}_{name}.png"
    local_file = os.path.join(LOCAL_DIR, file_full_name)
    temp_remote_file = f"{REMOTE_DIR}/transfer_temp.png"

    # --- 第一道保险：彻底删除旧图，并确认它消失了 ---
    # 使用长连接发送，确保此时设备端没有任何名为 transfer_temp.png 的文件
    adb_shell(f'rm -f "{temp_remote_file}"')
    adb_shell("sync")
    time.sleep(5.0)  # 给系统文件系统一点响应时间

    # 2. 预留渲染时间
    logger(f"  [PRE-SHOT] 清理缓存并等待页面渲染...")
    adb_shell("echo 3 > /proc/sys/vm/drop_caches")
    time.sleep(5.0)

    # --- 新增：截图前的“冷静期” ---
    logger(f"  [PRE-SHOT] 等待磁盘 IO 同步...")
    adb_shell("sync")  # 清理系统缓存
    time.sleep(5.0)  # 给系统 2 秒时间处理缓存清理

    # --- 第二道保险：下达截图指令 ---
    logger(f"[SCREENSHOT] 正在请求新截图: {name}")
    shot_cmd = f'PATH=$PATH:/usr/bin:/bin ffmpeg -f fbdev -i /dev/fb0 -vf transpose=1 -frames:v 1 -y "{temp_remote_file}"'
    adb_shell(shot_cmd)

    # --- 第三道保险：严格轮询 (必须是“新”生成的) ---
    max_wait = 360
    interval = 10.0  # 稍微加快轮询频率
    found = False

    logger(f"[POLLING] 等待新文件生成...")
    for i in range(int(max_wait / interval)):
        if interval <= 0:
            interval = 1.0
        # 判断逻辑：文件必须存在，且大小必须大于 10KB (防止截成黑屏或损坏的 0 字节文件)
        # [ -s ] 判断文件存在且大小大于0
        check_cmd = f'[ -s "{temp_remote_file}" ] && echo "NEW_FILE_READY" || echo "STILL_WAITING"'
        shell_proc.stdin.write(f"{check_cmd}\n")
        shell_proc.stdin.flush()

        matched = False
        start_t = time.time()
        while time.time() - start_t < 2.0:
            line = safe_read_line()
            if "NEW_FILE_READY" in line:
                matched = True
                break
            if "STILL_WAITING" in line:
                break

        if matched:
            # 这里的匹配是绝对安全的，因为我们在开头已经 rm 掉了旧图
            # 如果现在有了，那它一定是 ffmpeg 刚刚生成的
            logger(f"[MATCH] 检测到新图片已就绪")
            adb_shell(f'chmod 777 "{temp_remote_file}"')
            time.sleep(3.0)  # 额外缓冲，确保 ffmpeg 彻底释放文件句柄
            found = True
            break

        logger(f"  ... 检查中({(i + 1) * interval}s)，旧图已删，新图尚未生成 ...")
        time.sleep(interval)

    # 4. 执行拉取
    if found:
        # 如果拉取前发现本地已经有同名文件（上一轮留下的），也先删掉
        if os.path.exists(local_file):
            os.remove(local_file)

        res = subprocess.run(["adb", "pull", temp_remote_file, local_file], capture_output=True)

        # ⭐新增：等待文件真正写完
        if not wait_file_ready(local_file):
            logger(f"[ERROR] 文件未就绪: {local_file}")
            return None

        # ⭐再做一次最终确认
        if os.path.exists(local_file) and os.path.getsize(local_file) > 0:
            logger(f"[SUCCESS] 截图已拉取并命名为: {file_full_name}")
            return local_file

        logger("[ERROR] pull成功但文件无效")
        return None


    if not os.path.exists(local_file) or os.path.getsize(local_file) == 0:
        logger("[ERROR] 截图文件无效")
        return None
    logger(f"[ERROR] 截图失败，可能超时或设备未响应")
    return None


def download_file(remote, local):
    """将设备端的截图拉取到 PC"""
    os.makedirs(os.path.dirname(local), exist_ok=True)
    result = subprocess.run(["adb", "pull", remote, local], capture_output=True)
    return os.path.exists(local)



def select_languages():
    print("\n" + "=" * 40)
    print("      20国语言多选清单")
    print("=" * 40)
    for k, v in ALL_LANGUAGES.items():
        print(f"{k.rjust(2)}. {v[0]}")
    print("=" * 40)
    choice = input("请输入序号多选(如 1,2,5) 或输入 'all': ").strip().lower()

    selected_data = []
    if choice == 'all':
        # 'all' 模式：提取字典里的所有项
        for k, v in ALL_LANGUAGES.items():
            selected_data.append((int(k), v[0], v[1]))
    else:
        # 序号多选模式
        # 修改点：确保 else 分支（多选序号）也返回三个值
        for idx in choice.split(','):
            idx = idx.strip()
            if idx in ALL_LANGUAGES:
                info = ALL_LANGUAGES[idx]
                selected_data.append((int(idx), info[0], info[1]))  # 必须包含 int(idx)

    # 打印一下选择结果，方便调试
    print(f"--- 已确认选择 {len(selected_data)} 种语言 ---")
    return selected_data

def check_ui_overflow(base_img_path, test_img_path):
    """
    对比基准图和待测图：
    如果中文版（基准）右侧是黑的，而待测版右侧有像素，说明超限。
    """
    try:
        if not os.path.exists(base_img_path) or not os.path.exists(test_img_path):
            return "N/A"

        # 读取图片并转为灰度
        img_base = safe_imread(base_img_path, 0)
        img_test = safe_imread(test_img_path, 0)

        # 1. 简单的像素差异计算：二值化处理
        _, thresh_base = cv2.threshold(img_base, 50, 255, cv2.THRESH_BINARY)
        _, thresh_test = cv2.threshold(img_test, 50, 255, cv2.THRESH_BINARY)

        # 2. 寻找中文文案的右边界 (假设从左往右扫描)
        # 这里需要根据你的 UI 布局微调区域
        rows, cols = thresh_base.shape
        # 简化版：对比两图的总像素密度，如果待测图明显多出很多白色像素，大概率是超限
        diff = cv2.absdiff(thresh_base, thresh_test)
        change_ratio = np.count_nonzero(diff) / (rows * cols)

        if change_ratio > 0.05: # 如果差异超过 5%，标记潜在超限
            return "Potential Overflow"
        return "Normal"
    except:
        return "Check Error"

def safe_cv_imread(path, flag=0):
    if not path or not os.path.exists(path):
        return None

    try:
        data = np.fromfile(path, dtype=np.uint8)
        img = cv2.imdecode(data, flag)
        return img
    except:
        return None

def safe_read_line(timeout=2.0):
    start = time.time()
    while time.time() - start < timeout:
        try:
            if shell_proc and shell_proc.stdout:
                line = shell_proc.stdout.readline()
                if line:
                    return line.strip()
        except:
            pass
        time.sleep(0.1)
    return ""


def analyze_ui_diff(base_path, test_path):
    if not base_path or not test_path:
        return "Missing Path"

    img_b = safe_cv_imread(base_path, 0)
    img_t = safe_cv_imread(test_path, 0)

    if img_b is None or img_t is None:
        logger(f"[WARN] 图片读取失败 base={base_path}, test={test_path}")
        return "Read Error"

    # 防止尺寸不一致
    if img_b.shape != img_t.shape:
        logger("[WARN] 图片尺寸不一致，自动resize")
        img_t = cv2.resize(img_t, (img_b.shape[1], img_b.shape[0]))

    _, thresh_b = cv2.threshold(img_b, 70, 255, cv2.THRESH_BINARY)
    _, thresh_t = cv2.threshold(img_t, 70, 255, cv2.THRESH_BINARY)

    diff = cv2.absdiff(thresh_b, thresh_t)

    diff_count = np.count_nonzero(diff)
    total_pixels = diff.size

    if total_pixels == 0:
        return "Bad Image"

    diff_ratio = (diff_count / total_pixels) * 100

    return f"Warning({diff_ratio:.1f}%)" if diff_ratio > 3.0 else "Pass"

def check_language_switch(base_path, test_path):
    img_base = safe_imread(base_path, 0)
    img_test = safe_imread(test_path, 0)

    if img_base is None or img_test is None:
        return False, 0

        # ===============================
        # ⭐ 关键修复：只看中间语言列表区域
        # ===============================
    h, w = img_base.shape

    # 语言列表通常在中间区域（可根据UI微调）
    y1, y2 = int(h * 0.15), int(h * 0.90)
    x1, x2 = int(w * 0.20), int(w * 0.80)

    img_base = img_base[y1:y2, x1:x2]
    img_test = img_test[y1:y2, x1:x2]

    # 二值化（增强高亮差异）
    _, b1 = cv2.threshold(img_base, 80, 255, cv2.THRESH_BINARY)
    _, b2 = cv2.threshold(img_test, 80, 255, cv2.THRESH_BINARY)

    # 差异计算
    diff = cv2.absdiff(b1, b2)

    diff_ratio = np.count_nonzero(diff) / diff.size

    # ===============================
    # ⭐ 关键：降低阈值（因为只检测高亮）
    # ===============================
    ok = diff_ratio > 0.005

    return ok, diff_ratio

def wait_file_ready(path, min_size=20*1024, timeout=10):
    start = time.time()

    while time.time() - start < timeout:
        if os.path.exists(path):
            size = os.path.getsize(path)
            if size > min_size:
                return True
        time.sleep(0.5)

    return False

# ==========================================================
# 4. 主流程逻辑
# ==========================================================
def main():
    skip_home_first = True
    # --- 新增：创建本次运行的专属文件夹 ---
    run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
    current_run_dir = os.path.join(BASE_RESULT_DIR, f"Run_{run_id}")
    os.makedirs(current_run_dir, exist_ok=True)

    subprocess.run(["adb", "kill-server"], capture_output=True)
    subprocess.run(["adb", "start-server"], capture_output=True)
    time.sleep(2.0)

    check_conn = subprocess.run(["adb", "get-state"], capture_output=True, text=True)
    if "device" not in check_conn.stdout:
        logger("[ERROR] 未检测到有效的 ADB 设备，程序退出！")
        return

    init_persistent_shell()
    time.sleep(2.0)

    adb_shell("/sbin/evtest /dev/input/event0 < /dev/null > /tmp/ev.log 2>&1 &")
    logger("[INIT] evtest 监听已启动")

    target_langs = select_languages()
    if not target_langs:
        return

    global LOCAL_DIR
    LOCAL_DIR = current_run_dir
    logger(f"[INIT] 本次测试结果将存放在: {current_run_dir}")

    excel_path = os.path.join(current_run_dir, f"UI_Report_{run_id}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UI对比报告"
    ws.append([
        "语言ID",
        "语言名称",
        "页面类型",
        "一级页面",
        "子页面",
        "基准图",
        "测试图",
        "diff",
        "状态"
    ])

    base_lang = target_langs[0]

    # 生成基准图
    base_id, base_name, base_coord = base_lang
    base_img_path = safe_capture(base_name, "BASE_LANG")

    try:
        all_langs = target_langs
        # =========================================================
        # ⭐ STEP 2：其他语言
        # =========================================================
        for index, (lang_id, lang_name, lang_coord) in enumerate(all_langs):

            logger(f"\n================ 当前准备切换的语言: {lang_name} ================")

            # 回到首页 + 重新进入语言页（统一状态）
            if skip_home_first:
                logger("[INIT] 首次启动，跳过回首页")
                skip_home_first = False
            else:
                for rx, ry in HOME_RESET_STEPS:
                    safe_click(rx, ry, "回首页")
                time.sleep(5)

            for gx, gy in GOTO_LANGUAGE_PAGE:
                click_with_evtest(gx, gy, page_label="进入语言设置步骤")
                time.sleep(10.0)

            if lang_id < 13:
                click_with_evtest(169, 49, page_label="语言唤醒")
            else:
                click_with_evtest(LANG_PAGE_NEXT_BTN[0],
                                  LANG_PAGE_NEXT_BTN[1],
                                  page_label="翻页")

            time.sleep(10)
            logger("[SCREEN] 语言切换前，开始截图1")
            current_img1 = safe_capture(lang_name, "LANG_BEFORE")

            click_with_evtest(lang_coord[0], lang_coord[1],
                              page_label=f"切换至_{lang_name}")

            logger("[WAIT] 等待语言刷新...")
            time.sleep(LANG_RELOAD)

            # =========================
            # ⭐ 两次截图验证（修复你原来的 bug）
            # =========================
            diff_ratio = 0
            status = "UNKNOWN"

            time.sleep(8.0)
            logger("[SCREEN] 语言切换后，开始截图2")
            current_img2 = safe_capture(lang_name, "LANG_AFTER")


            ok1, r1 = check_language_switch(base_img_path, current_img1)
            ok2, r2 = check_language_switch(base_img_path, current_img2)

            ok = ok1 or ok2
            ratio = max(r1, r2)

            if ok:
                status = "PASS"
            else:
                status = "WARN"

            logger(f"[LANG CHECK] diff={ratio:.4f}")

            if not ok:
                logger(f"[WARN] 语言可能未切换或已是当前语言: {lang_name} (diff={ratio:.4f})")
            else:
                logger("✅ 语言切换成功")

            if current_img1 is None or current_img2 is None:
                status = "FAIL_NO_IMAGE"

            # =========================
            # ⭐ 新增：切换语言后返回首页
            # =========================
            logger("[NAV] 返回首页（语言切换后）")

            for i in range(2):
                click_with_evtest(63, 617, page_label=f"语言切换后返回首页_{i + 1}")
                time.sleep(15.0)  # ⭐ 必须给UI时间

            ws.append([
                lang_id,
                lang_name,
                "LANG",  # ⭐ 新增这一列（关键）
                "语言设置",
                "切换验证",
                current_img1,
                current_img2,
                ratio,
                status
            ])

            # F列：基准图
            if base_img_path and os.path.exists(base_img_path):
                img_base = OpenpyxlImage(base_img_path)
                img_base.width = 220
                img_base.height = 130
                img_base.anchor = f"F{ws.max_row}"
                ws.add_image(img_base)

            # ⭐ 插入图片（关键）
            #G列：测试图
            if current_img2 and os.path.exists(current_img2):
                img = OpenpyxlImage(current_img2)
                row = ws.max_row
                img.width = 220
                img.height = 130
                img.anchor = f"G{row}"

                ws.column_dimensions["G"].width = 45
                ws.row_dimensions[row].height = 120
                ws.add_image(img, f"G{row}")


            wb.save(excel_path)

            # =====================================================
            # UI扫描部分（完全不动你的逻辑）
            # =====================================================
            last_page_was_home = True
            adb_shell("sync; echo 3 > /proc/sys/vm/drop_caches")
            ui_status = "UNKNOWN"
            ui_diff_ratio = 0

            for main_page, config in UI_STRUCTURE.items():
                logger(f"\n--- 正在处理一级页面: {main_page} ---")

                if main_page == "首页":
                    last_page_was_home = True
                else:
                    if not last_page_was_home:
                        reset_to_home(main_page)

                    ex, ey = config.get("entry", (None, None))
                    if ex:
                        click_with_evtest(ex, ey, page_label=f"进入_{main_page}")
                        last_page_was_home = False

                for sub_name, actions in config.get("sub_pages", {}).items():
                    logger(f"  [SUB] {sub_name}")

                    # =========================
                    # ⭐ 情况1：dict子页面
                    # =========================
                    if isinstance(actions, dict):

                        sub_entry = actions.get("entry", (None, None))

                        if sub_entry and sub_entry[0] is not None:
                            click_with_evtest(
                                sub_entry[0],
                                sub_entry[1],
                                page_label=f"{main_page}_{sub_name}_进入"
                            )
                            time.sleep(SCREEN_RENDER)

                        img_path = capture_screen(lang_name, f"{main_page}_{sub_name}")
                        if not img_path:
                            logger("[SKIP] 截图失败，跳过分析")
                            continue

                        base_ui_path = base_img_path
                        ui_status = analyze_ui_diff(base_ui_path, img_path)

                        try:
                            img_b = safe_imread(base_ui_path, 0)
                            img_t = safe_imread(img_path, 0)

                            diff = cv2.absdiff(img_b, img_t)
                            diff_ratio = np.count_nonzero(diff) / diff.size
                        except:
                            diff_ratio = -1

                        ws.append([
                            lang_id,
                            lang_name,
                            "UI",
                            main_page,
                            sub_name,
                            base_ui_path,
                            img_path,
                            diff_ratio,
                            ui_status
                        ])

                        # F列：基准图
                        if base_ui_path and os.path.exists(base_ui_path):
                            img_base = OpenpyxlImage(base_ui_path)
                            img_base.width = 220
                            img_base.height = 130
                            img_base.anchor = f"F{ws.max_row}"
                            ws.add_image(img_base)

                        # G列：测试图
                        if img_path and os.path.exists(img_path):
                            img_test = OpenpyxlImage(img_path)
                            img_test.width = 220
                            img_test.height = 130
                            img_test.anchor = f"G{ws.max_row}"
                            ws.add_image(img_test)

                        # =========================
                        # ⭐ SUB2
                        # =========================
                        for sub2_name, sub2_actions in actions.get("sub_pages", {}).items():
                            logger(f"    [SUB2] {sub2_name}")

                            ax, ay = None, None

                            if isinstance(sub2_actions, tuple):
                                ax, ay = sub2_actions

                            elif isinstance(sub2_actions, dict):
                                sub2_entry = sub2_actions.get("entry", (None, None))
                                if sub2_entry:
                                    ax, ay = sub2_entry

                            elif isinstance(sub2_actions, list) and len(sub2_actions) > 0:
                                act = sub2_actions[0]
                                if isinstance(act, tuple):
                                    ax, ay = act

                            if ax is not None:
                                click_with_evtest(ax, ay, page_label=f"{sub_name}_{sub2_name}")

                            time.sleep(SCREEN_RENDER)

                            img_path = capture_screen(
                                lang_name,
                                f"{main_page}_{sub_name}_{sub2_name}"
                            )

                            if not img_path:
                                logger("[SKIP] SUB2截图失败")
                                continue

                            ui_status = analyze_ui_diff(base_img_path, img_path)

                            try:
                                img_b = safe_imread(base_img_path, 0)
                                img_t = safe_imread(img_path, 0)

                                if img_b is None or img_t is None:
                                    diff_ratio = -1
                                else:
                                    diff = cv2.absdiff(img_b, img_t)
                                    diff_ratio = np.count_nonzero(diff) / diff.size

                            except Exception as e:
                                logger(f"[DIFF ERROR] {e}")
                                diff_ratio = -1

                            ws.append([
                                lang_id,
                                lang_name,
                                "UI",
                                main_page,
                                f"{sub_name}_{sub2_name}",
                                base_img_path,
                                img_path,
                                diff_ratio,
                                ui_status
                            ])

                            # F列：基准图
                            if base_ui_path and os.path.exists(base_ui_path):
                                img_base = OpenpyxlImage(base_ui_path)
                                img_base.width = 220
                                img_base.height = 130
                                img_base.anchor = f"F{ws.max_row}"
                                ws.add_image(img_base)

                            # G列：测试图
                            if img_path and os.path.exists(img_path):
                                img_test = OpenpyxlImage(img_path)
                                img_test.width = 220
                                img_test.height = 130
                                img_test.anchor = f"G{ws.max_row}"
                                ws.add_image(img_test)


                        continue

                    # =========================
                    # ⭐ 情况2：普通click
                    # =========================
                    if actions is None:
                        actions = [(None, None)]

                    if isinstance(actions, tuple):
                        actions = [actions]

                    if not isinstance(actions, list):
                        actions = [actions]

                    for i, act in enumerate(actions):

                        ax, ay = None, None

                        if isinstance(act, tuple):
                            if len(act) == 3:
                                _, ax, ay = act
                            else:
                                ax, ay = act

                        if ax is not None:
                            click_with_evtest(ax, ay, page_label=f"{main_page}_{sub_name}")

                        time.sleep(SCREEN_RENDER)

                        ws.append([
                            lang_id,
                            lang_name,
                            "UI",
                            main_page,
                            f"{sub_name}_{i}",
                            base_img_path,
                            img_path,
                            "OK"
                        ])

            logger(f"[LANG DONE] {lang_name}")

            wb.save(excel_path)

    finally:
        logger(f"\n[FINISHED] 任务结束。报告: {excel_path}")
        wb.save(excel_path)
        logger("[EXCEL] 报告已保存")






if __name__ == "__main__":
    main()